# core/views.py
from datetime import timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from core.permissions import group_required
from itertools import chain
from decimal import Decimal
from django.db.models import Avg, Count
from core.services.pedagogie import sync_enseignant_groupe_from_matiere
from core.views_prof import _allowed_groupes as _allowed_groupes_for_user
from datetime import time
from django.db.models import Q
from core.utils_dates import month_start, add_months
from .forms import AnneeScolaireForm
from .models import AnneeScolaire,PaiementLigne, ProfGroupe
from .models import Degre
from .forms import DegreForm
from .models import Niveau
from .forms import NiveauForm
from .models import Groupe
from .forms import GroupeForm
from .models import Eleve
from .forms import EleveForm
from .models import AnneeScolaire, Niveau, FraisNiveau
from .models import Inscription
from .forms import InscriptionForm
from django.http import JsonResponse
from django.db.models import Sum, DecimalField
from django.db.models.functions import Coalesce
from .models import Paiement
from .forms import PaiementForm
from .models import Enseignant
from .forms import EnseignantForm
from .models import Seance
from .forms import SeanceForm
from .models import Absence
from .forms import AbsenceForm
from .models import Parent, ParentEleve
from .forms import ParentForm, ParentEleveFormSet
from . import pdf_utils
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import Matiere, Periode, Evaluation, Note
from .forms import MatiereForm, EvaluationForm
from .notes_utils import bulletin_data, rang_eleve
from .models import Recouvrement, Relance
from django.utils import timezone
from django.db.models import Sum, DecimalField, F, ExpressionWrapper
from .models import FraisNiveau, AnneeScolaire
from django.db.models import Count, Max
from datetime import date, datetime
from .models import EnseignantGroupe
from .forms import EnseignantGroupeForm
from .utils_users import get_or_create_user_with_group
import csv
from django.contrib.auth import get_user_model
from .models import Avis, SmsHistorique, Parent, ParentEleve, Eleve, Groupe, Niveau, Degre
from .forms_communication import AvisForm, SmsSendForm
from .services.sms_provider import normalize_phone, send_sms_via_bulksms_ma
import calendar
from datetime import date as date_cls
from datetime import date
from django.db.models import Sum, Count, F, Value, DecimalField
from django.db.models.functions import Coalesce, TruncMonth
from django.shortcuts import render, redirect
from core.models import Eleve, Groupe, Inscription, Paiement, AnneeScolaire
from django.db.models import (
    Sum, Count, F, Value, DecimalField,
)
from django.db.models import OuterRef, Subquery
from django.db.models.functions import Coalesce

from core.models import Niveau, Groupe, Inscription, FraisNiveau, TransactionFinance

# ⚠️ adapte ces imports si tes noms diffèrent
from core.models import TransactionLigne, RemboursementFinance, RemboursementFinanceLigne

from openpyxl.styles import Font, Alignment, PatternFill

from core.models import (
    AnneeScolaire, Eleve, Groupe, Paiement, Inscription, Absence
)
import json
from django.db.models import (
    Sum, Count, F, Q, Value, DecimalField, ExpressionWrapper
)
from django.db.models.functions import TruncMonth
from django.db import transaction
from django.shortcuts import redirect, render
from django.shortcuts import render
from django.db.models import Min, Max
from .models import AnneeScolaire, Degre, FraisNiveau
from django.db.models import Min, Max, Count  # ✅ ajoute Count
from accounts.decorators import group_required
from django.db.models import Q, Exists, OuterRef
from core.models import Eleve, Note, RelanceMensuelle  # ajuste si besoin
from .forms import InscriptionFullForm
from datetime import datetime
import uuid
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum, DecimalField
from django.http import JsonResponse, Http404
from .models import (
    AnneeScolaire, Niveau, Groupe, Periode,
    Inscription, Eleve, ParentEleve,
    EcheanceMensuelle, Paiement
)
from django.db.models import Q, Sum, F, DecimalField
from django.utils.dateparse import parse_date
from accounts.permissions import group_required
from django.db.models import Sum, F, DecimalField, ExpressionWrapper, Case, When, Value, BooleanField
from core.models import (
    AnneeScolaire, Periode, Niveau, Groupe, Inscription,
    TransactionFinance
)
from openpyxl import Workbook
from core.utils import mois_index_courant, mois_nom
from .models import Enseignant, AnneeScolaire, AbsenceProf
from .services_absences_profs import stats_mensuelles_prof
from core.models import (
    Enseignant, AnneeScolaire, Niveau, Groupe, Periode,
    EnseignantGroupe, Seance, AbsenceProf
)
from django.db.models.deletion import ProtectedError
from .models import Seance, Absence
from .models import AnneeScolaire, Groupe, Niveau, Enseignant, Seance
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
from .models import Parent, AnneeScolaire, Niveau, Groupe  # + ParentEleve pas obligatoire ici
from .pdf_utils import paiement_recu_pdf, paiement_recu_batch_pdf
from .models import (
    AnneeScolaire, Niveau, Groupe, Periode,
    Matiere, Enseignant, Evaluation, EnseignantGroupe
)
from django.shortcuts import get_object_or_404, render, redirect
from core.models import (
    Eleve, Inscription, AnneeScolaire, Periode,
    Note, Absence
)
from core import pdf_utils  # adapte si besoin
from .models import ParentEleve
from django.db.models import (
    Count, Sum, F, Value, DecimalField, IntegerField, Q
)
from .models import Eleve, Groupe, Inscription, Paiement, Niveau
from .utils_users import reset_password
from .models import TempPassword
from django.contrib.auth.models import Group
from .forms_users import UserCreateForm, UserUpdateForm, PasswordChangeForm
from .utils_users import generate_temp_password, reset_password
from django.shortcuts import render, get_object_or_404
from django.db import IntegrityError
from decimal import Decimal, InvalidOperation
from core.models import Degre, AnneeScolaire
from .models import EcheanceMensuelle
from core.models import Periode
from .views_prof import _allowed_groupes_for_user
from core.services.transport_echeances import sync_transport_echeances_for_inscription
from core.models import Inscription, EleveTransport, EcheanceTransportMensuelle
from django.shortcuts import get_object_or_404, redirect
from core.models import Eleve, EleveTransport, Inscription
from core.models import TransactionFinance, RemboursementFinance
from django.db import models
from core.services.notes_stats import moyenne_classe

from core.models import Depense
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def _as_datetime_safe(value):
    """
    Convertit value -> datetime (timezone-aware si possible).
    Accepte: datetime, date, str ISO, None.
    """
    if value is None:
        return timezone.make_aware(datetime.min) if timezone.is_naive(datetime.min) else datetime.min

    if isinstance(value, datetime):
        return timezone.make_aware(value) if timezone.is_naive(value) else value

    if isinstance(value, date):
        dt = datetime(value.year, value.month, value.day)
        return timezone.make_aware(dt)

    if isinstance(value, str):
        s = value.strip()
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        except Exception:
            return timezone.make_aware(datetime.min) if timezone.is_naive(datetime.min) else datetime.min

    return timezone.make_aware(datetime.min) if timezone.is_naive(datetime.min) else datetime.min

# ============================================================
# Helpers roles (si tu les as déjà ailleurs, garde les tiens)
# ============================================================
def _is_superadmin(user) -> bool:
    return bool(getattr(user, "is_superuser", False))

def _is_admin(user) -> bool:
    if _is_superadmin(user):
        return True
    # adapte si tu utilises Groups
    return user.groups.filter(name__in=["ADMIN", "SUPER_ADMIN"]).exists()


# ============================================================
# Dashboard Context — FINAL (Impayés mensuels + Encaissé non 0)
# - Impayés: basé sur EcheanceMensuelle (mois courant) + (option) inscription
# - Encaissé: Paiement + TransactionFinance (nouveau système)
# ============================================================


MONEY_FIELD = DecimalField(max_digits=12, decimal_places=2)
ZERO_MONEY = Value(Decimal("0.00"), output_field=MONEY_FIELD)


def _month_key(x):
    if isinstance(x, datetime):
        x = x.date()
    if isinstance(x, date):
        return date(x.year, x.month, 1)
    return None


def _model_has_field(model, field_name: str) -> bool:
    try:
        model._meta.get_field(field_name)
        return True
    except Exception:
        return False


def _last_12_months_labels(today: date):
    labels = []
    d = date(today.year, today.month, 1)
    for i in range(11, -1, -1):
        y = d.year + (d.month - 1 - i) // 12
        m = (d.month - 1 - i) % 12 + 1
        labels.append(date(y, m, 1))
    return labels


def _trend(cur, prev):
    cur = Decimal(cur or "0.00")
    prev = Decimal(prev or "0.00")

    # Aucun mouvement
    if cur == 0 and prev == 0:
        return 0

    # Nouveau chiffre ce mois-ci
    if prev == 0 and cur > 0:
        return None  # ⬅️ important

    # Chute totale
    if prev > 0 and cur == 0:
        return -100

    return int(((cur - prev) / prev) * 100)


# --- helper: champ élève dispo dans EcheanceMensuelle ---
def _echeance_eleve_values_key():
    """
    Retourne une clé utilisable dans QuerySet.values()
    IMPORTANT: doit être un champ RÉELLEMENT présent dans le modèle Django.
    """
    # EcheanceMensuelle a toujours "eleve" dans le modèle actuel
    return "eleve_id"


PERIOD_CHOICES = ("7d", "month", "year")

def _get_period(request):
    p = (request.GET.get("period") or "month").strip().lower()
    return p if p in PERIOD_CHOICES else "month"

def _get_range_for_period(period, today, annee_active_obj=None):
    """
    Retourne (date_from, date_to_excl) en date() pour filtrer :
    - 7d   : [today-6, tomorrow)
    - month: [1er du mois, 1er du mois prochain)
    - year : [annee.date_debut, annee.date_fin+1) si dispo, sinon fallback 01/01..01/01+1
    """
    if period == "7d":
        d_from = today - timedelta(days=6)
        d_to = today + timedelta(days=1)
        return d_from, d_to

    if period == "year":
        if annee_active_obj and getattr(annee_active_obj, "date_debut", None) and getattr(annee_active_obj, "date_fin", None):
            d_from = annee_active_obj.date_debut
            d_to = annee_active_obj.date_fin + timedelta(days=1)
            return d_from, d_to
        # fallback année civile
        d_from = date(today.year, 1, 1)
        d_to = date(today.year + 1, 1, 1)
        return d_from, d_to

    # month par défaut
    d_from = date(today.year, today.month, 1)
    d_to = add_months(d_from, 1)
    return d_from, d_to

def _date_to_dt_start(d):
    # d est date -> datetime début de journée
    return timezone.make_aware(timezone.datetime(d.year, d.month, d.day, 0, 0, 0))

def _date_to_dt_end_excl(d):
    # d est date "exclusive" -> datetime début de journée
    return timezone.make_aware(timezone.datetime(d.year, d.month, d.day, 0, 0, 0))

def _build_dashboard_context(period="month"):
    today = timezone.now().date()
    derniers_paiements = []

    # =========================
    # Helpers dates
    # =========================
    def _to_dt(d):
        return timezone.make_aware(timezone.datetime(d.year, d.month, d.day, 0, 0, 0))

    def _safe_year_range(annee_obj, today_):
        # année scolaire si dispo sinon année civile
        if annee_obj and getattr(annee_obj, "date_debut", None) and getattr(annee_obj, "date_fin", None):
            y_from = annee_obj.date_debut
            y_to_excl = annee_obj.date_fin + timedelta(days=1)
            return y_from, y_to_excl
        y_from = date(today_.year, 1, 1)
        y_to_excl = date(today_.year + 1, 1, 1)
        return y_from, y_to_excl

    def _period_range_dates(p):
        if p == "7d":
            d_from = today - timedelta(days=6)
            d_to = today + timedelta(days=1)
            return d_from, d_to, "7 jours"

        if p == "year":
            y_from, y_to_excl = _safe_year_range(annee_active_obj, today)
            return y_from, y_to_excl, "Année"

        # month
        d_from = date(today.year, today.month, 1)
        d_to = add_months(d_from, 1)
        return d_from, d_to, "Mois"

    def _month_start(d):
        return date(d.year, d.month, 1)

    def _month_end_excl(d):
        return add_months(_month_start(d), 1)

    def _iter_months(d_from, d_to_excl):
        cur = _month_start(d_from)
        end = _month_start(d_to_excl)
        out = []
        while cur < end:
            out.append(cur)
            cur = add_months(cur, 1)
        return out

    # =========================
    # Périodes month/last/next (pour taux mensuel + trend inscriptions)
    # =========================
    start_this_month = date(today.year, today.month, 1)
    start_last_month = add_months(start_this_month, -1)
    start_next_month = add_months(start_this_month, 1)

    # =========================
    # Année active
    # =========================
    annee_active_obj = AnneeScolaire.objects.filter(is_active=True).first()
    annee_active = annee_active_obj.nom if annee_active_obj else "—"

    if not annee_active_obj:
        return {
            "today": today,
            "current_date": today,
            "annee_active": annee_active,
            "period": period,
            "period_label": "",
            "date_from": today,
            "date_to": today,

            "nb_eleves": 0,
            "nb_groupes": 0,

            "total_net_mois": Decimal("0.00"),
            "total_net_annee": Decimal("0.00"),
            "total_depenses": Decimal("0.00"),
            "total_impayes": Decimal("0.00"),
            "encaisse_brut": Decimal("0.00"),
            "rembourse_total": Decimal("0.00"),

            "nb_impayes": 0,
            "impayes_mois": Decimal("0.00"),

            "taux_paiement": 0,
            "eleves_trend": 0,
            "groupes_trend": 0,
            "revenue_trend": 0,
            "payment_trend": 0,

            "derniers_paiements": [],
            "nouvelles_inscriptions": [],
            "impayes_recents": [],
            "repartition_eleves": [],
            "presences": {"presents": 0, "absents": 0, "retards": 0, "total": 0, "taux_presence": 0, "taux_absence": 0, "taux_retard": 0},

            # objectif annuel
            "objectif_annuel": Decimal("0.00"),
            "objectif_atteint": 0,
            "objectif_reste": Decimal("0.00"),

            # chart
            "chart_labels": json.dumps([]),
            "chart_net": json.dumps([]),
            "chart_depenses": json.dumps([]),
            "chart_impayes": json.dumps([]),

            "mois_courant": 0,
            "mois_courant_nom": "",
        }

    # =========================
    # Range selon filtre
    # =========================
    date_from, date_to_excl, period_label = _period_range_dates(period)
    dt_from = _to_dt(date_from)
    dt_to = _to_dt(date_to_excl)

    # =========================
    # Année scolaire range (pour annuel)
    # =========================
    year_from, year_to_excl = _safe_year_range(annee_active_obj, today)
    year_dt_from = _to_dt(year_from)
    year_dt_to = _to_dt(year_to_excl)

    # =========================
    # Helpers mois courant scolaire
    # =========================
    idx_courant = mois_index_courant(annee_active_obj, today)  # 1..10 (Sep..Jun)
    idx_courant = max(1, min(10, int(idx_courant or 1)))
    mois_courant_nom = mois_nom(idx_courant)

    # =========================
    # Global counts
    # =========================
    inscs_active = Inscription.objects.filter(annee_id=annee_active_obj.id, eleve__is_active=True)
    nb_eleves = inscs_active.values("eleve_id").distinct().count()
    nb_groupes = inscs_active.values("groupe_id").distinct().count()

    # =========================
    # ENCAISSÉ BRUT période (Paiement + Transaction)
    # =========================
    enc_old = (
        Paiement.objects
        .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
        .filter(date_paiement__gte=date_from, date_paiement__lt=date_to_excl)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    enc_new = (
        TransactionFinance.objects
        .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
        .filter(date_transaction__gte=dt_from, date_transaction__lt=dt_to)
        .aggregate(s=Coalesce(Sum("montant_total"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    encaisse_brut = enc_old + enc_new

    # REMBOURSEMENTS période
    rembourse_total = (
        RemboursementFinance.objects
        .filter(transaction__inscription__annee_id=annee_active_obj.id, transaction__inscription__eleve__is_active=True)
        .filter(created_at__gte=dt_from, created_at__lt=dt_to)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    # NET période
    total_net_periode = encaisse_brut - rembourse_total

    # =========================
    # NET ANNUEL (année scolaire)
    # =========================
    enc_year_old = (
        Paiement.objects
        .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
        .filter(date_paiement__gte=year_from, date_paiement__lt=year_to_excl)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    enc_year_new = (
        TransactionFinance.objects
        .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
        .filter(date_transaction__gte=year_dt_from, date_transaction__lt=year_dt_to)
        .aggregate(s=Coalesce(Sum("montant_total"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    remb_year = (
        RemboursementFinance.objects
        .filter(transaction__inscription__annee_id=annee_active_obj.id, transaction__inscription__eleve__is_active=True)
        .filter(created_at__gte=year_dt_from, created_at__lt=year_dt_to)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    total_net_annee = (enc_year_old + enc_year_new) - remb_year

    # =========================
    # DÉPENSES période
    # =========================
    total_depenses = (
        Depense.objects
        .filter(annee_id=annee_active_obj.id)
        .filter(date_depense__gte=date_from, date_depense__lt=date_to_excl)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    # =========================
    # 🎯 OBJECTIF ANNUEL (TA FORMULE)
    # =========================
    objectif_scolarite = (
        EcheanceMensuelle.objects
        .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
        .aggregate(s=Coalesce(Sum("montant_du"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    objectif_transport = Decimal("0.00")
    try:
        objectif_transport = (
            EcheanceTransportMensuelle.objects
            .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
            .aggregate(s=Coalesce(Sum("montant_du"), ZERO_MONEY))["s"]
            or Decimal("0.00")
        )
    except Exception:
        pass

    objectif_inscription = (
        Inscription.objects
        .filter(
            annee_id=annee_active_obj.id,
            eleve__is_active=True,
            date_inscription__gte=year_from,
            date_inscription__lt=year_to_excl,
        )
        .aggregate(s=Coalesce(Sum("frais_inscription_du"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    objectif_annuel = objectif_scolarite + objectif_transport + objectif_inscription
    objectif_atteint = int(min((total_net_annee / objectif_annuel) * 100, 100)) if objectif_annuel > 0 else 0
    objectif_reste = max(objectif_annuel - total_net_annee, Decimal("0.00"))

    # =========================
    # IMPAYÉS (selon période)
    # =========================
    sco_qs = (
        EcheanceMensuelle.objects
        .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
        .annotate(
            reste=ExpressionWrapper(
                Coalesce(F("montant_du"), ZERO_MONEY) - Coalesce(F("montant_paye"), ZERO_MONEY),
                output_field=MONEY_FIELD,
            )
        )
        .filter(reste__gt=0)
    )

    tr_qs = None
    try:
        tr_qs = (
            EcheanceTransportMensuelle.objects
            .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
            .annotate(
                reste=ExpressionWrapper(
                    Coalesce(F("montant_du"), ZERO_MONEY) - Coalesce(F("montant_paye"), ZERO_MONEY),
                    output_field=MONEY_FIELD,
                )
            )
            .filter(reste__gt=0)
        )
    except Exception:
        tr_qs = None

    ins_qs = (
        Inscription.objects
        .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
        .annotate(
            reste_ins=ExpressionWrapper(
                Coalesce(F("frais_inscription_du"), ZERO_MONEY) - Coalesce(F("frais_inscription_paye"), ZERO_MONEY),
                output_field=MONEY_FIELD,
            )
        )
        .filter(reste_ins__gt=0)
    )

    if period == "month":
        sco_qs = sco_qs.filter(mois_index=idx_courant)
        if tr_qs is not None:
            tr_qs = tr_qs.filter(mois_index=idx_courant)
        ins_qs = ins_qs.filter(date_inscription__gte=start_this_month, date_inscription__lt=start_next_month)
    else:
        if _model_has_field(EcheanceMensuelle, "date_echeance"):
            sco_qs = sco_qs.filter(date_echeance__gte=date_from, date_echeance__lt=date_to_excl)
        if tr_qs is not None and _model_has_field(EcheanceTransportMensuelle, "date_echeance"):
            tr_qs = tr_qs.filter(date_echeance__gte=date_from, date_echeance__lt=date_to_excl)
        ins_qs = ins_qs.filter(date_inscription__gte=date_from, date_inscription__lt=date_to_excl)

    total_impaye_sco = sco_qs.aggregate(s=Coalesce(Sum("reste"), ZERO_MONEY))["s"] or Decimal("0.00")
    total_impaye_tr = Decimal("0.00")
    if tr_qs is not None:
        total_impaye_tr = tr_qs.aggregate(s=Coalesce(Sum("reste"), ZERO_MONEY))["s"] or Decimal("0.00")
    total_impaye_ins = ins_qs.aggregate(s=Coalesce(Sum("reste_ins"), ZERO_MONEY))["s"] or Decimal("0.00")

    total_impayes = total_impaye_sco + total_impaye_tr + total_impaye_ins

    # impayes_mois (optionnel : utile si tu veux l’afficher quelque part)
    impayes_mois = Decimal("0.00")
    try:
        imp_m_sco = (
            EcheanceMensuelle.objects
            .filter(annee_id=annee_active_obj.id, mois_index=idx_courant, eleve__is_active=True)
            .annotate(
                reste=ExpressionWrapper(
                    Coalesce(F("montant_du"), ZERO_MONEY) - Coalesce(F("montant_paye"), ZERO_MONEY),
                    output_field=MONEY_FIELD,
                )
            )
            .filter(reste__gt=0)
            .aggregate(s=Coalesce(Sum("reste"), ZERO_MONEY))["s"] or Decimal("0.00")
        )
        imp_m_tr = Decimal("0.00")
        try:
            imp_m_tr = (
                EcheanceTransportMensuelle.objects
                .filter(annee_id=annee_active_obj.id, mois_index=idx_courant, eleve__is_active=True)
                .annotate(
                    reste=ExpressionWrapper(
                        Coalesce(F("montant_du"), ZERO_MONEY) - Coalesce(F("montant_paye"), ZERO_MONEY),
                        output_field=MONEY_FIELD,
                    )
                )
                .filter(reste__gt=0)
                .aggregate(s=Coalesce(Sum("reste"), ZERO_MONEY))["s"] or Decimal("0.00")
            )
        except Exception:
            pass
        imp_m_ins = (
            Inscription.objects
            .filter(annee_id=annee_active_obj.id, eleve__is_active=True, date_inscription__gte=start_this_month, date_inscription__lt=start_next_month)
            .annotate(
                reste_ins=ExpressionWrapper(
                    Coalesce(F("frais_inscription_du"), ZERO_MONEY) - Coalesce(F("frais_inscription_paye"), ZERO_MONEY),
                    output_field=MONEY_FIELD,
                )
            )
            .filter(reste_ins__gt=0)
            .aggregate(s=Coalesce(Sum("reste_ins"), ZERO_MONEY))["s"] or Decimal("0.00")
        )
        impayes_mois = imp_m_sco + imp_m_tr + imp_m_ins
    except Exception:
        impayes_mois = Decimal("0.00")

    # =========================
    # NB IMPAYÉS + IMPAYÉS RÉCENTS (anti double comptage)
    # =========================
    from django.db.models import OuterRef, Subquery

    def _sub_sum_paiements():
        return Coalesce(
            Subquery(
                Paiement.objects
                .filter(inscription_id=OuterRef("pk"))
                .values("inscription_id")
                .annotate(s=Sum("montant"))
                .values("s")[:1]
            ),
            ZERO_MONEY
        )

    def _sub_sum_transactions():
        return Coalesce(
            Subquery(
                TransactionFinance.objects
                .filter(inscription_id=OuterRef("pk"))
                .values("inscription_id")
                .annotate(s=Sum("montant_total"))
                .values("s")[:1]
            ),
            ZERO_MONEY
        )

    def _sub_sum_remboursements():
        return Coalesce(
            Subquery(
                RemboursementFinance.objects
                .filter(transaction__inscription_id=OuterRef("pk"))
                .values("transaction__inscription_id")
                .annotate(s=Sum("montant"))
                .values("s")[:1]
            ),
            ZERO_MONEY
        )

    impayes_global_qs = (
        Inscription.objects
        .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
        .annotate(p_old=_sub_sum_paiements())
        .annotate(p_new=_sub_sum_transactions())
        .annotate(remb=_sub_sum_remboursements())
        .annotate(
            p_total=ExpressionWrapper(F("p_old") + F("p_new") - F("remb"), output_field=MONEY_FIELD)
        )
        .annotate(
            reste_global=ExpressionWrapper(Coalesce(F("montant_total"), ZERO_MONEY) - F("p_total"), output_field=MONEY_FIELD)
        )
        .filter(reste_global__gt=0)
    )

    nb_eleves_impayes = impayes_global_qs.values("eleve_id").distinct().count()

    impayes_recents = (
        Inscription.objects
        .select_related("eleve", "groupe")
        .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
        .annotate(p_old=_sub_sum_paiements())
        .annotate(p_new=_sub_sum_transactions())
        .annotate(remb=_sub_sum_remboursements())
        .annotate(
            p_total=ExpressionWrapper(F("p_old") + F("p_new") - F("remb"), output_field=MONEY_FIELD)
        )
        .annotate(
            reste=ExpressionWrapper(Coalesce(F("montant_total"), ZERO_MONEY) - F("p_total"), output_field=MONEY_FIELD)
        )
        .filter(reste__gt=0)
        .order_by("-reste")[:8]
    )

    # =========================
    # Trends (NET) — annuel vs année précédente (range -12 mois)
    # =========================
    prev_year_from = add_months(year_from, -12)
    prev_year_to_excl = add_months(year_to_excl, -12)
    prev_dt_from = _to_dt(prev_year_from)
    prev_dt_to = _to_dt(prev_year_to_excl)

    enc_prev_old = (
        Paiement.objects
        .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
        .filter(date_paiement__gte=prev_year_from, date_paiement__lt=prev_year_to_excl)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    enc_prev_new = (
        TransactionFinance.objects
        .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
        .filter(date_transaction__gte=prev_dt_from, date_transaction__lt=prev_dt_to)
        .aggregate(s=Coalesce(Sum("montant_total"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    remb_prev = (
        RemboursementFinance.objects
        .filter(transaction__inscription__annee_id=annee_active_obj.id, transaction__inscription__eleve__is_active=True)
        .filter(created_at__gte=prev_dt_from, created_at__lt=prev_dt_to)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    net_prev_year = (enc_prev_old + enc_prev_new) - remb_prev
    revenue_trend = _trend(total_net_annee, net_prev_year)

    # Trends inscriptions (mensuel)
    eleves_this = Inscription.objects.filter(
        annee_id=annee_active_obj.id,
        date_inscription__gte=start_this_month,
        date_inscription__lt=start_next_month,
    ).count()
    eleves_last = Inscription.objects.filter(
        annee_id=annee_active_obj.id,
        date_inscription__gte=start_last_month,
        date_inscription__lt=start_this_month,
    ).count()
    eleves_trend = _trend(eleves_this, eleves_last)

    # =========================
    # TAUX PAIEMENT (mensuel)
    # =========================
    dt_m_from = _to_dt(start_this_month)
    dt_m_to = _to_dt(start_next_month)

    enc_m_old = (
        Paiement.objects
        .filter(inscription__annee_id=annee_active_obj.id)
        .filter(date_paiement__gte=start_this_month, date_paiement__lt=start_next_month)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    enc_m_new = (
        TransactionFinance.objects
        .filter(inscription__annee_id=annee_active_obj.id)
        .filter(date_transaction__gte=dt_m_from, date_transaction__lt=dt_m_to)
        .aggregate(s=Coalesce(Sum("montant_total"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    remb_m = (
        RemboursementFinance.objects
        .filter(transaction__inscription__annee_id=annee_active_obj.id)
        .filter(created_at__gte=dt_m_from, created_at__lt=dt_m_to)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    net_mois = (enc_m_old + enc_m_new) - remb_m

    attendu_sco_mois = (
        EcheanceMensuelle.objects
        .filter(annee_id=annee_active_obj.id, mois_index=idx_courant)
        .aggregate(s=Coalesce(Sum("montant_du"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    attendu_tr_mois = Decimal("0.00")
    try:
        attendu_tr_mois = (
            EcheanceTransportMensuelle.objects
            .filter(annee_id=annee_active_obj.id, mois_index=idx_courant)
            .aggregate(s=Coalesce(Sum("montant_du"), ZERO_MONEY))["s"] or Decimal("0.00")
        )
    except Exception:
        pass
    attendu_ins_mois = (
        Inscription.objects
        .filter(annee_id=annee_active_obj.id, date_inscription__gte=start_this_month, date_inscription__lt=start_next_month)
        .aggregate(s=Coalesce(Sum("frais_inscription_du"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    attendu_mois = attendu_sco_mois + attendu_tr_mois + attendu_ins_mois
    taux_paiement = int(min((net_mois / attendu_mois) * 100, 100)) if attendu_mois > 0 else 0

    # taux_last (mois précédent)
    dt_last_from = _to_dt(start_last_month)
    dt_last_to = _to_dt(start_this_month)

    enc_last_old = (
        Paiement.objects
        .filter(inscription__annee_id=annee_active_obj.id)
        .filter(date_paiement__gte=start_last_month, date_paiement__lt=start_this_month)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    enc_last_new = (
        TransactionFinance.objects
        .filter(inscription__annee_id=annee_active_obj.id)
        .filter(date_transaction__gte=dt_last_from, date_transaction__lt=dt_last_to)
        .aggregate(s=Coalesce(Sum("montant_total"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    remb_last = (
        RemboursementFinance.objects
        .filter(transaction__inscription__annee_id=annee_active_obj.id)
        .filter(created_at__gte=dt_last_from, created_at__lt=dt_last_to)
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    net_last = (enc_last_old + enc_last_new) - remb_last

    idx_last = max(1, min(10, idx_courant - 1))
    attendu_sco_last = (
        EcheanceMensuelle.objects
        .filter(annee_id=annee_active_obj.id, mois_index=idx_last)
        .aggregate(s=Coalesce(Sum("montant_du"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    attendu_tr_last = Decimal("0.00")
    try:
        attendu_tr_last = (
            EcheanceTransportMensuelle.objects
            .filter(annee_id=annee_active_obj.id, mois_index=idx_last)
            .aggregate(s=Coalesce(Sum("montant_du"), ZERO_MONEY))["s"] or Decimal("0.00")
        )
    except Exception:
        pass
    attendu_ins_last = (
        Inscription.objects
        .filter(annee_id=annee_active_obj.id, date_inscription__gte=start_last_month, date_inscription__lt=start_this_month)
        .aggregate(s=Coalesce(Sum("frais_inscription_du"), ZERO_MONEY))["s"] or Decimal("0.00")
    )
    attendu_last = attendu_sco_last + attendu_tr_last + attendu_ins_last
    taux_last = int(min((net_last / attendu_last) * 100, 100)) if attendu_last > 0 else 0
    payment_trend = _trend(taux_paiement, taux_last)

    # =========================
    # Derniers paiements (OLD + NEW)
    # =========================
    last_old = (
        Paiement.objects
        .select_related("inscription__eleve", "inscription__groupe")
        .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
        .order_by("-date_paiement", "-id")[:8]
    )

    old_items = []
    for p in last_old:
        old_items.append({
            "type": "OLD",
            "eleve_nom": getattr(p.inscription.eleve, "nom_complet", str(p.inscription.eleve)),
            "groupe_nom": getattr(p.inscription.groupe, "nom", "—"),
            "montant": p.montant,
            "date": p.date_paiement,
            "mode": getattr(p, "mode", None),
            "mode_label": getattr(p, "get_mode_display", lambda: "")(),
        })

    last_new = (
        TransactionFinance.objects
        .select_related("inscription__eleve", "inscription__groupe")
        .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
        .order_by("-date_transaction", "-id")[:8]
    )

    new_items = []
    for t in last_new:
        new_items.append({
            "type": "NEW",
            "eleve_nom": getattr(t.inscription.eleve, "nom_complet", str(t.inscription.eleve)),
            "groupe_nom": getattr(t.inscription.groupe, "nom", "—"),
            "montant": t.montant_total,
            "date": t.date_transaction,
            "mode": getattr(t, "mode", None),
            "mode_label": getattr(t, "get_mode_display", lambda: "Transaction")() or "Transaction",
        })

    derniers_paiements = sorted((old_items + new_items), key=lambda p: _as_datetime_safe(p.get("date")), reverse=True)[:8]

    # nouvelles inscriptions
    nouvelles_inscriptions = (
        Inscription.objects
        .select_related("eleve", "groupe", "groupe__niveau")
        .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
        .order_by("-date_inscription", "-id")[:8]
    )

    # Répartition par niveau
    repartition_qs = (
        Inscription.objects
        .filter(annee_id=annee_active_obj.id, eleve__is_active=True)
        .select_related("groupe__niveau")
        .values("groupe__niveau__nom")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    repartition_eleves = [{"nom": r["groupe__niveau__nom"] or "—", "total": r["total"], "couleur": None} for r in repartition_qs]

    # Présences
    absents = Absence.objects.filter(date=today, type="ABS", eleve__is_active=True).values("eleve_id").distinct().count()
    retards = Absence.objects.filter(date=today, type="RET", eleve__is_active=True).values("eleve_id").distinct().count()
    total = nb_eleves
    presents = max(total - absents, 0)

    presences = {
        "presents": presents,
        "absents": absents,
        "retards": retards,
        "total": total,
        "taux_presence": int((presents / total) * 100) if total else 0,
        "taux_absence": int((absents / total) * 100) if total else 0,
        "taux_retard": int((retards / total) * 100) if total else 0,
    }

    # =========================
    # Chart pilotage — ANNÉE SCOLAIRE (NET / DEPENSES / IMPAYES)
    # =========================
    months = _iter_months(year_from, year_to_excl)
    months_fr = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Aoû", "Sep", "Oct", "Nov", "Déc"]
    chart_labels = [months_fr[m.month - 1] for m in months]

    chart_net = []
    chart_depenses = []
    chart_impayes = []

    for m_start in months:
        m_end_excl = _month_end_excl(m_start)
        dt_m_from = _to_dt(m_start)
        dt_m_to = _to_dt(m_end_excl)

        enc_old_m = (
            Paiement.objects
            .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
            .filter(date_paiement__gte=m_start, date_paiement__lt=m_end_excl)
            .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
        )
        enc_new_m = (
            TransactionFinance.objects
            .filter(inscription__annee_id=annee_active_obj.id, inscription__eleve__is_active=True)
            .filter(date_transaction__gte=dt_m_from, date_transaction__lt=dt_m_to)
            .aggregate(s=Coalesce(Sum("montant_total"), ZERO_MONEY))["s"] or Decimal("0.00")
        )
        remb_m = (
            RemboursementFinance.objects
            .filter(transaction__inscription__annee_id=annee_active_obj.id, transaction__inscription__eleve__is_active=True)
            .filter(created_at__gte=dt_m_from, created_at__lt=dt_m_to)
            .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
        )
        net_m = (enc_old_m + enc_new_m) - remb_m

        dep_m = (
            Depense.objects
            .filter(annee_id=annee_active_obj.id)
            .filter(date_depense__gte=m_start, date_depense__lt=m_end_excl)
            .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"] or Decimal("0.00")
        )

        idx_m = mois_index_courant(annee_active_obj, m_start)
        idx_m = int(max(1, min(10, idx_m or 1)))

        imp_sco = (
            EcheanceMensuelle.objects
            .filter(annee_id=annee_active_obj.id, mois_index=idx_m, eleve__is_active=True)
            .annotate(
                reste=ExpressionWrapper(
                    Coalesce(F("montant_du"), ZERO_MONEY) - Coalesce(F("montant_paye"), ZERO_MONEY),
                    output_field=MONEY_FIELD,
                )
            )
            .filter(reste__gt=0)
            .aggregate(s=Coalesce(Sum("reste"), ZERO_MONEY))["s"] or Decimal("0.00")
        )

        imp_tr = Decimal("0.00")
        try:
            imp_tr = (
                EcheanceTransportMensuelle.objects
                .filter(annee_id=annee_active_obj.id, mois_index=idx_m, eleve__is_active=True)
                .annotate(
                    reste=ExpressionWrapper(
                        Coalesce(F("montant_du"), ZERO_MONEY) - Coalesce(F("montant_paye"), ZERO_MONEY),
                        output_field=MONEY_FIELD,
                    )
                )
                .filter(reste__gt=0)
                .aggregate(s=Coalesce(Sum("reste"), ZERO_MONEY))["s"] or Decimal("0.00")
            )
        except Exception:
            pass

        imp_ins = (
            Inscription.objects
            .filter(annee_id=annee_active_obj.id, eleve__is_active=True, date_inscription__gte=m_start, date_inscription__lt=m_end_excl)
            .annotate(
                reste_ins=ExpressionWrapper(
                    Coalesce(F("frais_inscription_du"), ZERO_MONEY) - Coalesce(F("frais_inscription_paye"), ZERO_MONEY),
                    output_field=MONEY_FIELD,
                )
            )
            .filter(reste_ins__gt=0)
            .aggregate(s=Coalesce(Sum("reste_ins"), ZERO_MONEY))["s"] or Decimal("0.00")
        )

        imp_m = imp_sco + imp_tr + imp_ins

        chart_net.append(float(net_m))
        chart_depenses.append(float(dep_m))
        chart_impayes.append(float(imp_m))

    # =========================
    # Context final
    # =========================
    return {
        "today": today,
        "current_date": today,
        "annee_active": annee_active,

        "period": period,
        "period_label": period_label,
        "date_from": date_from,
        "date_to": (date_to_excl - timedelta(days=1)),

        "nb_eleves": nb_eleves,
        "nb_groupes": nb_groupes,

        "total_net_mois": total_net_periode,
        "total_net_annee": total_net_annee,
        "total_depenses": total_depenses,
        "total_impayes": total_impayes,
        "encaisse_brut": encaisse_brut,
        "rembourse_total": rembourse_total,

        "nb_impayes": nb_eleves_impayes,
        "impayes_mois": impayes_mois,

        "taux_paiement": taux_paiement,
        "eleves_trend": eleves_trend,
        "groupes_trend": 0,
        "revenue_trend": revenue_trend,
        "payment_trend": payment_trend,

        "derniers_paiements": derniers_paiements,
        "nouvelles_inscriptions": nouvelles_inscriptions,
        "impayes_recents": impayes_recents,
        "repartition_eleves": repartition_eleves,
        "presences": presences,

        # ✅ Objectif annuel + reste
        "objectif_annuel": objectif_annuel,
        "objectif_atteint": objectif_atteint,
        "objectif_reste": objectif_reste,

        # Chart
        "chart_labels": json.dumps(chart_labels),
        "chart_net": json.dumps(chart_net),
        "chart_depenses": json.dumps(chart_depenses),
        "chart_impayes": json.dumps(chart_impayes),

        "mois_courant": idx_courant,
        "mois_courant_nom": mois_courant_nom,
    }

def _build_staff_dashboard_context():
    today = timezone.now().date()

    # périodes mois courant
    start_this_month = date(today.year, today.month, 1)
    start_next_month = add_months(start_this_month, 1)

    annee_active_obj = AnneeScolaire.objects.filter(is_active=True).first()
    annee_active = annee_active_obj.nom if annee_active_obj else "—"

    # ✅ SAFE si pas d'année active
    if not annee_active_obj:
        return {
            "today": today,
            "current_date": today,
            "annee_active": annee_active,
            "nb_eleves": 0,
            "nb_groupes": 0,
            "total_paye": Decimal("0.00"),
            "presences": {
                "presents": 0, "absents": 0, "retards": 0, "total": 0,
                "taux_presence": 0, "taux_absence": 0, "taux_retard": 0,
            },
        }

    inscs_active = Inscription.objects.filter(
        annee_id=annee_active_obj.id,
        eleve__is_active=True,
    )
    nb_eleves = inscs_active.values("eleve_id").distinct().count()
    nb_groupes = inscs_active.values("groupe_id").distinct().count()

    # ✅ Encaissement mensuel (comme admin) => Paiement + TransactionFinance
    pay_mois_old = (
        Paiement.objects
        .filter(
            inscription__annee_id=annee_active_obj.id,
            date_paiement__gte=start_this_month,
            date_paiement__lt=start_next_month,
        )
        .aggregate(s=Coalesce(Sum("montant"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    pay_mois_new = (
        TransactionFinance.objects
        .filter(
            inscription__annee_id=annee_active_obj.id,
            date_transaction__gte=start_this_month,
            date_transaction__lt=start_next_month,
        )
        .aggregate(s=Coalesce(Sum("montant_total"), ZERO_MONEY))["s"]
        or Decimal("0.00")
    )

    total_paye = (pay_mois_old or Decimal("0.00")) + (pay_mois_new or Decimal("0.00"))

    # Présences
    absents = Absence.objects.filter(date=today, type="ABS").values("eleve_id").distinct().count()
    retards = Absence.objects.filter(date=today, type="RET").values("eleve_id").distinct().count()
    total = nb_eleves
    presents = max(total - absents, 0)

    presences = {
        "presents": presents,
        "absents": absents,
        "retards": retards,
        "total": total,
        "taux_presence": int((presents / total) * 100) if total else 0,
        "taux_absence": int((absents / total) * 100) if total else 0,
        "taux_retard": int((retards / total) * 100) if total else 0,
    }

    return {
        "today": today,
        "current_date": today,
        "annee_active": annee_active,

        "nb_eleves": nb_eleves,
        "nb_groupes": nb_groupes,
        "total_paye": total_paye,

        "presences": presences,
    }

# ─────────────────────────────────────────
# Route "dashboard" -> redirige selon role
# ─────────────────────────────────────────
@login_required
def dashboard(request):
    user = request.user

    if _is_superadmin(user):
        return redirect("core:dashboard_superadmin")

    if _is_admin(user):
        return redirect("core:dashboard_admin")

    return redirect("core:dashboard_staff")


@login_required
def dashboard_staff(request):
    ctx = _build_staff_dashboard_context()
    ctx["dash_kind"] = "staff"
    ctx["dash_title"] = "Dashboard Direction"
    return render(request, "admin/Dashboard/staff.html", ctx)


@login_required
def dashboard_admin(request):
    ctx = _build_dashboard_context()
    ctx["dash_kind"] = "admin"
    ctx["dash_title"] = "Tableau de bord"
    return render(request, "admin/Dashboard/admin.html", ctx)


@login_required
def dashboard_superadmin(request):
    if not _is_superadmin(request.user):
        return redirect("core:dashboard_admin")

    period = (request.GET.get("period") or "month").strip().lower()
    if period not in ("7d", "month", "year"):
        period = "month"

    ctx = _build_dashboard_context(period=period)
    ctx["dash_kind"] = "superadmin"
    ctx["dash_title"] = "SuperAdmin Center"
    ctx["period"] = period  # ✅ pour template (chips)

    User = get_user_model()
    ctx["nb_users"] = User.objects.count()
    ctx["nb_staff"] = User.objects.filter(is_staff=True).count()
    ctx["nb_superusers"] = User.objects.filter(is_superuser=True).count()

    return render(request, "admin/Dashboard/superadmin.html", ctx)


def mois_index_courant(annee: AnneeScolaire, today: date_cls) -> int:
    """
    Retourne l’index du mois scolaire AZ (1..10)
    Septembre = 1, Octobre = 2, ..., Juin = 10
    """
    if not annee or not annee.date_debut:
        return 1

    mapping = {
        9: 1,   # Septembre
        10: 2,  # Octobre
        11: 3,  # Novembre
        12: 4,  # Décembre
        1: 5,   # Janvier
        2: 6,   # Février
        3: 7,   # Mars
        4: 8,   # Avril
        5: 9,   # Mai
        6: 10,  # Juin
    }

    mois_reel = today.month

    # hors année scolaire (été)
    if mois_reel not in mapping:
        return 10  # on considère fin d’année

    return mapping[mois_reel]


@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def annee_list(request):
    annees = AnneeScolaire.objects.all()
    return render(request, "admin/annees/list.html", {"annees": annees})


@login_required
@permission_required("core.add_anneescolaire", raise_exception=True)
def annee_create(request):
    if request.method == "POST":
        form = AnneeScolaireForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    annee = form.save()
                messages.success(request, "✅ Année scolaire ajoutée + semestres créés (S1/S2).")
                return redirect("core:annee_list")

            except Exception as e:
                messages.error(request, f"⚠️ Erreur lors de la création des semestres: {e}")
        else:
            messages.error(request, "⚠️ Formulaire invalide. Vérifie les champs.")
    else:
        form = AnneeScolaireForm()

    return render(request, "admin/annees/form.html", {"form": form, "mode": "create"})

@login_required
@permission_required("core.change_anneescolaire", raise_exception=True)
def annee_update(request, pk):
    annee = get_object_or_404(AnneeScolaire, pk=pk)

    if request.method == "POST":
        form = AnneeScolaireForm(request.POST, instance=annee)
        if form.is_valid():
            try:
                with transaction.atomic():
                    annee = form.save()
                messages.success(request, "✅ Année scolaire mise à jour + semestres vérifiés (S1/S2).")
                return redirect("core:annee_list")

            except Exception as e:
                messages.error(request, f"⚠️ Erreur lors de la vérification des semestres: {e}")
    else:
        form = AnneeScolaireForm(instance=annee)

    return render(request, "admin/annees/form.html", {
        "form": form,
        "mode": "update",
        "annee": annee
    })


@login_required
@permission_required("core.delete_anneescolaire", raise_exception=True)
def annee_delete(request, pk):
    annee = get_object_or_404(AnneeScolaire, pk=pk)

    if request.method == "POST":
        try:
            annee.delete()
            messages.success(request, "🗑️ Année supprimée avec succès.")
            return redirect("core:annee_list")

        except ProtectedError:
            messages.error(
                request,
                "❌ Suppression impossible : cette année contient des données "
                "(groupes, élèves, paiements, séances…). "
                "Désactive-la plutôt."
            )
            return redirect("core:annee_list")

    return render(request, "admin/annees/delete.html", {
        "annee": annee,
    })


@login_required
@permission_required("core.change_anneescolaire", raise_exception=True)
def annee_activer(request, pk):
    annee = get_object_or_404(AnneeScolaire, pk=pk)
    AnneeScolaire.objects.update(is_active=False)
    annee.is_active = True
    annee.save()
    messages.success(request, f"✅ Année active: {annee.nom}")
    return redirect("core:annee_list")

# ============================
# B1 — Degrés
# ============================


@login_required
@permission_required("core.view_degre", raise_exception=True)
def degre_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    degres = Degre.objects.all().order_by("ordre", "nom")

    resume_by_degre_id = {}
    cap_by_degre_id = {}

    if annee_active:
        # ✅ résumé prix existant
        qs = (
            FraisNiveau.objects
            .filter(annee=annee_active)
            .values("niveau__degre_id")
            .annotate(
                min_ins=Min("frais_inscription"),
                max_ins=Max("frais_inscription"),
            )
        )
        resume_by_degre_id = {
            row["niveau__degre_id"]: {
                "min": row["min_ins"] or Decimal("0.00"),
                "max": row["max_ins"] or Decimal("0.00"),
            }
            for row in qs
        }

        # ✅ capacité (= nb élèves) par degré
        cap_qs = (
            Inscription.objects
            .filter(annee=annee_active, eleve__is_active=True)
            .values("groupe__niveau__degre_id")
            .annotate(c=Count("id"))
        )
        cap_by_degre_id = {row["groupe__niveau__degre_id"]: row["c"] for row in cap_qs}


    return render(request, "admin/degres/list.html", {
        "degres": degres,
        "annee_active": annee_active,
        "resume_by_degre_id": resume_by_degre_id,
        "cap_by_degre_id": cap_by_degre_id,  # ✅ NEW
    })


@login_required
@permission_required("core.add_degre", raise_exception=True)
def degre_create(request):
    if request.method == "POST":
        form = DegreForm(request.POST)
        if form.is_valid():
            nom = form.cleaned_data["nom"].strip()
            # code auto (propre, sans accents -> simple)
            code = nom.upper().replace(" ", "_").replace("É", "E").replace("È", "E").replace("Ê", "E").replace("Ë", "E") \
                              .replace("À", "A").replace("Â", "A").replace("Ä", "A") \
                              .replace("Ç", "C") \
                              .replace("Î", "I").replace("Ï", "I") \
                              .replace("Ô", "O").replace("Ö", "O") \
                              .replace("Ù", "U").replace("Û", "U").replace("Ü", "U")

            Degre.objects.create(code=code, nom=nom, ordre=form.cleaned_data["ordre"])
            messages.success(request, "✅ Degré ajouté.")
            return redirect("core:degre_list")
    else:
        form = DegreForm()

    return render(request, "admin/degres/form.html", {"form": form, "mode": "create"})


@login_required
@permission_required("core.change_degre", raise_exception=True)
def degre_update(request, pk):
    degre = get_object_or_404(Degre, pk=pk)
    if request.method == "POST":
        form = DegreForm(request.POST, instance=degre)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Degré mis à jour.")
            return redirect("core:degre_list")
    else:
        form = DegreForm(instance=degre)

    return render(request, "admin/degres/form.html", {"form": form, "mode": "update", "degre": degre})


@login_required
@permission_required("core.delete_degre", raise_exception=True)
def degre_delete(request, pk):
    degre = get_object_or_404(Degre, pk=pk)

    if request.method == "POST":
        try:
            degre.delete()
            messages.success(request, "🗑️ Degré supprimé.")
        except ProtectedError:
            messages.error(
                request,
                "❌ Suppression impossible : ce degré contient des niveaux (ex: CE1, CM1...). "
                "Supprime d’abord les niveaux liés."
            )
        return redirect("core:degre_list")

    return render(request, "admin/degres/delete.html", {"degre": degre})
# ============================
# B2 — Niveaux
# ============================


@login_required
def niveau_prix_edit(request, niveau_id):
    niveau = get_object_or_404(Niveau, id=niveau_id)
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    if not annee_active:
        messages.error(request, "Aucune année scolaire active. Active une année avant de définir les prix.")
        return redirect("core:niveau_list")

    frais_niveau, _ = FraisNiveau.objects.get_or_create(
        annee=annee_active,
        niveau=niveau,
        defaults={
            "frais_inscription": Decimal("0.00"),
            "frais_scolarite_mensuel": Decimal("0.00"),
        },
    )

    if request.method == "POST":
        insc_str = (request.POST.get("frais_inscription") or "").replace(",", ".").strip()
        mens_str = (request.POST.get("frais_scolarite_mensuel") or "").replace(",", ".").strip()

        def parse_money(v: str) -> Decimal:
            if v == "":
                return Decimal("0.00")
            d = Decimal(v)
            if d < 0:
                raise ValueError("neg")
            return d

        try:
            frais_inscription = parse_money(insc_str)
            frais_mensuel = parse_money(mens_str)
        except Exception:
            messages.error(request, "Montant invalide. Exemple: 1500 ou 1500.00")
            return render(request, "admin/niveaux/prix_edit.html", {
                "niveau": niveau,
                "annee_active": annee_active,
                "frais_niveau": frais_niveau,
            })

        frais_niveau.frais_inscription = frais_inscription
        frais_niveau.frais_scolarite_mensuel = frais_mensuel
        frais_niveau.save(update_fields=["frais_inscription", "frais_scolarite_mensuel"])

        messages.success(
            request,
            f"✅ Prix enregistrés: Inscription={frais_inscription} MAD | Mensuel={frais_mensuel} MAD ({niveau.nom})"
        )
        return redirect("core:niveau_list")

    return render(request, "admin/niveaux/prix_edit.html", {
        "niveau": niveau,
        "annee_active": annee_active,
        "frais_niveau": frais_niveau,
    })

@login_required
@permission_required("core.view_niveau", raise_exception=True)
def niveau_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    annee_id = (request.GET.get("annee") or "").strip()
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    degre_selected = (request.GET.get("degre") or "").strip()

    degres = Degre.objects.all().order_by("ordre", "nom")
    annees = AnneeScolaire.objects.all().order_by("-id")

    niveaux_qs = Niveau.objects.select_related("degre").all()
    if degre_selected:
        niveaux_qs = niveaux_qs.filter(degre_id=degre_selected)
    niveaux_qs = niveaux_qs.order_by("degre__ordre", "ordre", "nom")

    # =========================
    # ✅ MAP FRAIS par niveau
    # =========================
    frais_map = {}
    if annee_id:
        frais_qs = FraisNiveau.objects.filter(annee_id=annee_id).select_related("niveau")
        frais_map = {f.niveau_id: f for f in frais_qs}

    # =========================
    # ✅ MAP CAPACITÉ (= nb élèves) par niveau
    # =========================
    cap_map = {}
    if annee_id:
        cap_qs = (
            Inscription.objects
            .filter(annee_id=annee_id, eleve__is_active=True)
            .values("groupe__niveau_id")
            .annotate(c=Count("id"))
        )
        cap_map = {row["groupe__niveau_id"]: row["c"] for row in cap_qs}


    rows = []
    for n in niveaux_qs:
        fn = frais_map.get(n.id)
        rows.append({
            "niveau": n,
            "degre": n.degre,
            "inscription": (fn.frais_inscription if fn else Decimal("0.00")),
            "mensuel": (fn.frais_scolarite_mensuel if fn else Decimal("0.00")),
            "capacite": cap_map.get(n.id, 0),  # ✅ nb élèves
        })

    return render(request, "admin/niveaux/list.html", {
        "annee_active": annee_active,
        "annees": annees,
        "annee_selected": annee_id,
        "degres": degres,
        "degre_selected": degre_selected,
        "rows": rows,
    })

@login_required
@permission_required("core.add_niveau", raise_exception=True)
def niveau_create(request):
    if request.method == "POST":
        form = NiveauForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Niveau ajouté.")
            return redirect("core:niveau_list")
    else:
        form = NiveauForm()

    return render(request, "admin/niveaux/form.html", {"form": form, "mode": "create"})

@login_required
@permission_required("core.change_niveau", raise_exception=True)
def niveau_update(request, pk):
    niveau = get_object_or_404(Niveau, pk=pk)
    if request.method == "POST":
        form = NiveauForm(request.POST, instance=niveau)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Niveau mis à jour.")
            return redirect("core:niveau_list")
    else:
        form = NiveauForm(instance=niveau)

    return render(request, "admin/niveaux/form.html", {"form": form, "mode": "update", "niveau": niveau})

@login_required
@permission_required("core.delete_niveau", raise_exception=True)
def niveau_delete(request, pk):
    niveau = get_object_or_404(Niveau, pk=pk)

    if request.method == "POST":
        force = (request.POST.get("force") == "1")

        if not force:
            try:
                niveau.delete()
                messages.success(request, "🗑️ Niveau supprimé.")
            except Exception as e:
                messages.error(request, f"❌ Suppression impossible : {e}")
            return redirect("core:niveau_list")

        # ✅ SUPPRESSION FORCÉE (débranche tout ce qui PROTECT)
        with transaction.atomic():
            # 1) Inscriptions concernées (toutes années)
            insc_ids = list(
                Inscription.objects.filter(groupe__niveau=niveau).values_list("id", flat=True)
            )

            if insc_ids:
                tx_ids = list(
                    TransactionFinance.objects.filter(inscription_id__in=insc_ids)
                    .values_list("id", flat=True)
                )

                if tx_ids:
                    # 2) Remboursements -> lignes remboursements
                    RemboursementFinanceLigne.objects.filter(remboursement__transaction_id__in=tx_ids).delete()
                    RemboursementFinance.objects.filter(transaction_id__in=tx_ids).delete()

                    # 3) Lignes de transaction
                    TransactionLigne.objects.filter(transaction_id__in=tx_ids).delete()

                    # 4) Transactions
                    TransactionFinance.objects.filter(id__in=tx_ids).delete()

                # 5) Inscriptions (échéances etc en cascade)
                Inscription.objects.filter(id__in=insc_ids).delete()

            # 6) Groupes
            Groupe.objects.filter(niveau=niveau).delete()

            # 7) Frais du niveau (toutes années)
            FraisNiveau.objects.filter(niveau=niveau).delete()

            # 8) Niveau
            niveau.delete()

        messages.success(
            request,
            "🗑️ Niveau supprimé (FORCÉ) : remboursements + transactions + inscriptions + groupes + frais supprimés."
        )
        return redirect("core:niveau_list")

    return render(request, "admin/niveaux/delete.html", {"niveau": niveau})

# ============================
# B3 — Groupes
# ============================

@login_required
@permission_required("core.view_groupe", raise_exception=True)
def groupe_list(request):
    q = request.GET.get("q", "").strip()
    niveau_id = request.GET.get("niveau", "")
    annee_id = request.GET.get("annee", "")

    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    groupes = (
        Groupe.objects
        .select_related("annee", "niveau", "niveau__degre")
        .annotate(
            nb_eleves=Count(
                "inscriptions",
                filter=Q(inscriptions__eleve__is_active=True),
                distinct=True
            )
        )
    )

    # ✅ par défaut : année active
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)

    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    if q:
        groupes = groupes.filter(
            Q(nom__icontains=q) |
            Q(niveau__nom__icontains=q) |
            Q(niveau__degre__nom__icontains=q)
        )

    niveaux = Niveau.objects.select_related("degre").all()
    annees = AnneeScolaire.objects.all()

    return render(
        request,
        "admin/groupes/list.html",
        {
            "groupes": groupes,
            "q": q,
            "niveaux": niveaux,
            "annees": annees,
            "niveau_selected": niveau_id,
            "annee_selected": annee_id,
        },
    )


@login_required
@permission_required("core.add_groupe", raise_exception=True)
def groupe_create(request):
    if request.method == "POST":
        form = GroupeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Groupe ajouté.")
            return redirect("core:groupe_list")
    else:
        form = GroupeForm()
    return render(request, "admin/groupes/form.html", {"form": form, "mode": "create"})


@login_required
@permission_required("core.change_groupe", raise_exception=True)
def groupe_update(request, pk):
    groupe = get_object_or_404(Groupe, pk=pk)
    if request.method == "POST":
        form = GroupeForm(request.POST, instance=groupe)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Groupe mis à jour.")
            return redirect("core:groupe_list")
    else:
        form = GroupeForm(instance=groupe)
    return render(request, "admin/groupes/form.html", {"form": form, "mode": "update", "groupe": groupe})


@login_required
@permission_required("core.delete_groupe", raise_exception=True)
def groupe_delete(request, pk):
    g = get_object_or_404(Groupe, pk=pk)

    if request.method == "POST":
        try:
            g.delete()
            messages.success(request, "🗑️ Groupe supprimé.")
        except ProtectedError:
            messages.error(
                request,
                "❌ Suppression impossible : ce groupe est utilisé (inscriptions, échéances, séances, cahier, PDF...). "
                "Supprime d’abord les éléments liés ou archive le groupe."
            )
        return redirect("core:groupe_list")

    return render(request, "admin/groupes/delete.html", {"groupe": g})

# ============================
# C1 — Élèves
# ============================


@login_required
@permission_required("core.view_eleve", raise_exception=True)
def eleve_list(request):
    q = request.GET.get("q", "").strip()
    statut = request.GET.get("statut", "")

    # ✅ nouveau filtre inscription
    # valeurs possibles: "", "inscrits", "non_inscrits"
    insc = request.GET.get("insc", "").strip()

    # ✅ filtres
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = request.GET.get("annee", "")
    niveau_id = request.GET.get("niveau", "")
    groupe_id = request.GET.get("groupe", "")
    periode_id = request.GET.get("periode", "")

    if statut == "inactifs":
        eleves = Eleve.objects.filter(is_active=False)
    else:
        statut = "actifs"
        eleves = Eleve.objects.filter(is_active=True)


    # ✅ recherche
    if q:
        eleves = eleves.filter(
            Q(matricule__icontains=q) |
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(telephone__icontains=q)
        )

    # ✅ Année par défaut = année active
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    # ===============================
    # ✅ LOGIQUE "INSCRIPTION" (IMPORTANT)
    # ===============================
    if annee_id:
        # Sous-requêtes existence
        insc_year = Inscription.objects.filter(eleve_id=OuterRef("pk"), annee_id=annee_id)
        insc_year_validee = Inscription.objects.filter(
            eleve_id=OuterRef("pk"),
            annee_id=annee_id,
            statut="VALIDEE",
        )
        insc_year_en_cours = Inscription.objects.filter(
            eleve_id=OuterRef("pk"),
            annee_id=annee_id,
            statut="EN_COURS",
        )

        eleves = eleves.annotate(
            has_insc_year=Exists(insc_year),
            has_insc_validee=Exists(insc_year_validee),
            has_insc_en_cours=Exists(insc_year_en_cours),
        )

        # 🔹 cas 1: afficher seulement "inscrits" (inscription validée)
        if insc == "inscrits":
            eleves = eleves.filter(has_insc_validee=True)

        # 🔹 cas 2: afficher "non inscrits" = (aucune inscription) OU (inscription EN_COURS)
        elif insc == "non_inscrits":
            eleves = eleves.filter(
                Q(has_insc_year=False) | Q(has_insc_en_cours=True)
            )

        # 🔹 cas défaut (insc == "") : comportement normal = élèves ayant une inscription sur l'année
        else:
            eleves = eleves.filter(has_insc_year=True)

        # ✅ Filtrage niveau/groupe/période :
        # - si insc == non_inscrits => ces filtres ne peuvent s'appliquer qu'aux élèves qui ont une inscription (EN_COURS)
        # - donc on les applique via la relation inscriptions (ça exclura ceux qui n'ont aucune inscription, ce qui est logique)
        if niveau_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__groupe__niveau_id=niveau_id)

        if groupe_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__groupe_id=groupe_id)

        if periode_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__periode_id=periode_id)

    eleves = eleves.distinct()

    # ✅ periodes pour dropdown
    periodes = Periode.objects.all()
    if annee_id:
        periodes = periodes.filter(annee_id=annee_id)

    # ✅ dropdowns
    annees = AnneeScolaire.objects.all()

    niveaux = Niveau.objects.all()
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    groupes = Groupe.objects.select_related("niveau", "annee").all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    return render(request, "admin/eleves/list.html", {
        "eleves": eleves,
        "q": q,
        "statut": statut,

        # ✅ nouveau
        "insc": insc,

        # ✅ context filtres
        "annee_active": annee_active,
        "annees": annees,
        "annee_selected": annee_id,
        "niveaux": niveaux,
        "niveau_selected": niveau_id,
        "groupes": groupes,
        "groupe_selected": groupe_id,
        "periodes": periodes,
        "periode_selected": periode_id,
    })

@login_required
@permission_required("core.view_eleve", raise_exception=True)
def eleves_pdf_view(request):
    # =========================
    # GET params (mêmes que la page)
    # =========================
    q = (request.GET.get("q") or "").strip()
    statut = (request.GET.get("statut") or "").strip()

    # "", "inscrits", "non_inscrits"
    insc = (request.GET.get("insc") or "").strip()

    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    periode_id = (request.GET.get("periode") or "").strip()

    # ✅ année par défaut = active (comme page)
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    # =========================
    # Base queryset élèves (comme page)
    # =========================
    if statut == "inactifs":
        statut = "inactifs"
        eleves = Eleve.objects.filter(is_active=False)
    else:
        statut = "actifs"
        eleves = Eleve.objects.filter(is_active=True)

    if q:
        eleves = eleves.filter(
            Q(matricule__icontains=q) |
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(telephone__icontains=q)
        )

    # ===============================
    # LOGIQUE "INSCRIPTION" (IDENTIQUE À eleve_list)
    # ===============================
    if annee_id:
        insc_year = Inscription.objects.filter(eleve_id=OuterRef("pk"), annee_id=annee_id)
        insc_year_validee = Inscription.objects.filter(
            eleve_id=OuterRef("pk"),
            annee_id=annee_id,
            statut="VALIDEE",
        )
        insc_year_en_cours = Inscription.objects.filter(
            eleve_id=OuterRef("pk"),
            annee_id=annee_id,
            statut="EN_COURS",
        )

        eleves = eleves.annotate(
            has_insc_year=Exists(insc_year),
            has_insc_validee=Exists(insc_year_validee),
            has_insc_en_cours=Exists(insc_year_en_cours),
        )

        # 🔹 cas 1: inscrits = validée
        if insc == "inscrits":
            eleves = eleves.filter(has_insc_validee=True)

        # 🔹 cas 2: non_inscrits = (aucune inscription) OU (inscription EN_COURS)
        elif insc == "non_inscrits":
            eleves = eleves.filter(Q(has_insc_year=False) | Q(has_insc_en_cours=True))

        # 🔹 défaut = élèves ayant une inscription sur l'année
        else:
            eleves = eleves.filter(has_insc_year=True)

        # ✅ filtres niveau/groupe/période via inscriptions
        if niveau_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__groupe__niveau_id=niveau_id)
        if groupe_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__groupe_id=groupe_id)
        if periode_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__periode_id=periode_id)

    eleves = eleves.distinct().order_by("nom", "prenom")

    title = (
        f"Filtres: q={q or '—'} / statut={statut or '—'} / insc={insc or '—'} "
        f"/ annee={annee_id or '—'} / niveau={niveau_id or '—'} / groupe={groupe_id or '—'} / periode={periode_id or '—'}"
    )

    return pdf_utils.eleves_list_pdf(title, eleves)

@login_required
@permission_required("core.view_eleve", raise_exception=True)
def eleves_excel_export(request):
    # =========================
    # GET params (mêmes que la page)
    # =========================
    q = (request.GET.get("q") or "").strip()
    statut = (request.GET.get("statut") or "").strip()

    # "", "inscrits", "non_inscrits"
    insc = (request.GET.get("insc") or "").strip()

    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    periode_id = (request.GET.get("periode") or "").strip()

    # ✅ année par défaut = active (comme page)
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    # =========================
    # Base queryset élèves (comme page)
    # =========================
    if statut == "inactifs":
        eleves = Eleve.objects.filter(is_active=False)
    else:
        statut = "actifs"
        eleves = Eleve.objects.filter(is_active=True)

    if q:
        eleves = eleves.filter(
            Q(matricule__icontains=q) |
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(telephone__icontains=q)
        )

    # ===============================
    # LOGIQUE "INSCRIPTION" (IDENTIQUE À eleve_list)
    # ===============================
    if annee_id:
        insc_year = Inscription.objects.filter(eleve_id=OuterRef("pk"), annee_id=annee_id)
        insc_year_validee = Inscription.objects.filter(
            eleve_id=OuterRef("pk"),
            annee_id=annee_id,
            statut="VALIDEE",
        )
        insc_year_en_cours = Inscription.objects.filter(
            eleve_id=OuterRef("pk"),
            annee_id=annee_id,
            statut="EN_COURS",
        )

        eleves = eleves.annotate(
            has_insc_year=Exists(insc_year),
            has_insc_validee=Exists(insc_year_validee),
            has_insc_en_cours=Exists(insc_year_en_cours),
        )

        if insc == "inscrits":
            eleves = eleves.filter(has_insc_validee=True)
        elif insc == "non_inscrits":
            eleves = eleves.filter(Q(has_insc_year=False) | Q(has_insc_en_cours=True))
        else:
            eleves = eleves.filter(has_insc_year=True)

        if niveau_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__groupe__niveau_id=niveau_id)
        if groupe_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__groupe_id=groupe_id)
        if periode_id:
            eleves = eleves.filter(inscriptions__annee_id=annee_id, inscriptions__periode_id=periode_id)

    eleves = eleves.distinct().order_by("nom", "prenom")

    # =========================
    # Précharger les inscriptions utiles pour afficher "Groupe"
    # On récupère toutes les inscriptions de l'année (pour les élèves exportés)
    # =========================
    eleve_ids = list(eleves.values_list("id", flat=True))
    insc_qs = (
        Inscription.objects
        .select_related("groupe", "groupe__niveau")
        .filter(eleve_id__in=eleve_ids)
    )
    if annee_id:
        insc_qs = insc_qs.filter(annee_id=annee_id)

    # Priorité: VALIDEE > EN_COURS > (reste)
    # On trie pour que la 1ère inscription rencontrée soit la meilleure
    statut_rank = {"VALIDEE": 0, "EN_COURS": 1}
    insc_qs = insc_qs.order_by("eleve_id", "statut", "-id")

    insc_by_eleve = {}
    for insc_obj in insc_qs:
        # on force la priorité manuellement (plus sûr)
        eid = insc_obj.eleve_id
        if eid not in insc_by_eleve:
            insc_by_eleve[eid] = insc_obj
        else:
            cur = insc_by_eleve[eid]
            cur_r = statut_rank.get(cur.statut, 9)
            new_r = statut_rank.get(insc_obj.statut, 9)
            if new_r < cur_r:
                insc_by_eleve[eid] = insc_obj

    # =========================
    # Excel PRO
    # =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eleves"

    headers = ["#", "Nom", "Prenom", "Groupe", "Téléphone"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    i = 0
    for e in eleves:
        i += 1
        insc_obj = insc_by_eleve.get(e.id)

        if insc_obj and insc_obj.groupe:
            g_nom = insc_obj.groupe.nom or ""
            niv_nom = insc_obj.groupe.niveau.nom if getattr(insc_obj.groupe, "niveau", None) else ""
            groupe_str = f"{g_nom} ({niv_nom})" if niv_nom else g_nom
        else:
            groupe_str = ""

        ws.append([
            i,
            e.nom or "",
            e.prenom or "",
            groupe_str,
            e.telephone or "",
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # largeur auto
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="eleves_pro.xlsx"'
    wb.save(response)
    return response

@login_required
@permission_required("core.change_eleve", raise_exception=True)
def eleves_excel_import(request):
    """
    Import Excel format:
    matricule | nom | prenom | telephone | is_active(1/0)
    """
    errors = []
    created = 0
    updated = 0

    if request.method == "POST":
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "⚠️ Aucun fichier uploadé.")
            return redirect("core:eleve_list")

        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
        except Exception:
            messages.error(request, "⚠️ Fichier Excel invalide.")
            return redirect("core:eleve_list")

        # headers
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        expected = ["matricule", "nom", "prenom", "telephone", "is_active"]
        if headers[:5] != expected:
            messages.error(request, f"⚠️ Colonnes attendues: {expected}")
            return redirect("core:eleve_list")

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            matricule, nom, prenom, telephone, is_active = row[:5]

            if not matricule or not nom or not prenom:
                errors.append(f"Ligne {idx}: matricule/nom/prenom obligatoire.")
                continue

            matricule = str(matricule).strip()
            nom = str(nom).strip()
            prenom = str(prenom).strip()
            telephone = str(telephone).strip() if telephone else ""
            is_active = str(is_active).strip() if is_active is not None else "1"
            active_bool = True if is_active in ["1", "true", "True", "OUI", "oui", "YES", "yes"] else False

            obj, created_flag = Eleve.objects.update_or_create(
                matricule=matricule,
                defaults={
                    "nom": nom,
                    "prenom": prenom,
                    "telephone": telephone,
                    "is_active": active_bool,
                }
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        if errors:
            messages.warning(request, f"⚠️ Import terminé avec erreurs ({len(errors)}).")
        messages.success(request, f"✅ Import terminé: {created} créés / {updated} mis à jour.")
        # On affiche erreurs dans une page dédiée
        return render(request, "admin/eleves/import_result.html", {
            "errors": errors,
            "created": created,
            "updated": updated,
        })

    # GET -> page upload
    return render(request, "admin/eleves/import.html")

@login_required
def eleve_detail(request, pk):
    eleve = get_object_or_404(Eleve, pk=pk)


    liens_parents = eleve.liens_parents.select_related("parent").all()

    inscriptions = (
        Inscription.objects
        .filter(eleve=eleve)
        .select_related("annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "periode")
        .prefetch_related( "paiements")
        .order_by("-annee__date_debut", "-id")
    )

    inscription_active = inscriptions.first()

    # ✅ Résumé "classe actuelle" (Degré / Niveau / Groupe)
    classe_active = None
    if inscription_active and inscription_active.groupe_id:
        g = inscription_active.groupe
        n = g.niveau
        d = n.degre
        classe_active = {
            "annee": inscription_active.annee,
            "periode": inscription_active.periode,
            "degre": d,
            "niveau": n,
            "groupe": g,
        }

    # =========================
    # FINANCE (inscription active)
    # =========================
    finance = None
    if inscription_active:
        echeances = (
                EcheanceMensuelle.objects
                .filter(eleve=eleve, annee=inscription_active.annee)
                .order_by("date_echeance", "mois_index")
            )

        total_mensuel_du = sum((e.montant_du or Decimal("0.00")) for e in echeances)
        total_mensuel_paye = sum((e.montant_paye or Decimal("0.00")) for e in echeances)
        total_mensuel_reste = total_mensuel_du - total_mensuel_paye

        total_insc_du = inscription_active.frais_inscription_du or Decimal("0.00")
        total_insc_paye = inscription_active.frais_inscription_paye or Decimal("0.00")
        total_insc_reste = total_insc_du - total_insc_paye

        paiements = (
            Paiement.objects
            .filter(inscription=inscription_active)
            .select_related("echeance")
            .order_by("-date_paiement", "-id")
        )

        finance = {
            "inscription": inscription_active,
            "echeances": echeances,
            "paiements": paiements,
            "kpi": {
                "insc_du": total_insc_du,
                "insc_paye": total_insc_paye,
                "insc_reste": total_insc_reste,
                "mensuel_du": total_mensuel_du,
                "mensuel_paye": total_mensuel_paye,
                "mensuel_reste": total_mensuel_reste,
                "global_reste": total_insc_reste + total_mensuel_reste,
            }
        }

    return render(request, "admin/eleves/detail.html", {
        "eleve": eleve,
        "liens_parents": liens_parents,
        "inscriptions": inscriptions,
        "inscription_active": inscription_active,
        "classe_active": classe_active,
        "finance": finance,
    })

@login_required
@permission_required("core.add_eleve", raise_exception=True)
def eleve_create(request):
    if request.method == "POST":
        form = EleveForm(request.POST, request.FILES)
        if form.is_valid():
            eleve = form.save()  # ✅ matricule généré via save()

        try:
            # ✅ 1) Créer le compte ÉLÈVE (username = matricule élève)
            user, pwd, created = get_or_create_user_with_group(eleve.matricule, "ELEVE")

            # ✅ 2) Lier le user au profil élève
            if getattr(eleve, "user_id", None) != user.id:
                eleve.user = user
                eleve.save(update_fields=["user"])

            # ✅ IMPORTANT : ✅ NE PAS créer de parent automatiquement ici.
            # (Le lien parent <-> élève sera fait uniquement via InscriptionFullForm
            #  ou une page dédiée "Affecter parent".)

            if created:
                messages.success(
                    request,
                    f"✅ Élève ajouté: {eleve.matricule} | Compte ÉLÈVE créé | MDP temporaire: {pwd}"
                )
            else:
                messages.info(request, f"ℹ️ Compte ÉLÈVE existe déjà pour {eleve.matricule}.")

        except Exception as e:
            messages.warning(request, f"⚠️ Élève créé mais compte ÉLÈVE non créé: {e}")

            return redirect("core:eleve_detail", pk=eleve.pk)
    else:
        form = EleveForm()

    return render(request, "admin/eleves/form.html", {"form": form, "mode": "create"})

@login_required
@permission_required("core.change_eleve", raise_exception=True)
def eleve_update(request, pk):
    eleve = get_object_or_404(Eleve, pk=pk)

    if request.method == "POST":
        # ✅ IMPORTANT : request.FILES pour update photo
        form = EleveForm(request.POST, request.FILES, instance=eleve)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Élève mis à jour.")
            return redirect("core:eleve_detail", pk=eleve.pk)
    else:
        form = EleveForm(instance=eleve)

    return render(request, "admin/eleves/form.html", {"form": form, "mode": "update", "eleve": eleve})

def _eleve_hard_delete(eleve: Eleve) -> None:
    """
    Supprime TOUT ce qui bloque (paiements, recouvrement, echeances, inscriptions, etc.)
    ⚠️ À utiliser seulement si tu veux une suppression définitive.
    """
    # 1) Liens directs à l'élève
    eleve.liens_parents.all().delete()      # ParentEleve (FK eleve CASCADE chez toi, mais safe)
    eleve.absences.all().delete()           # Absence (PROTECT -> on delete avant)
    Note.objects.filter(eleve=eleve).delete()

    # 2) Inscriptions + dépendances
    for insc in eleve.inscriptions.all():
        # relances mensuelles liées aux échéances de l'inscription (optionnel)
        RelanceMensuelle.objects.filter(echeance__inscription=insc).delete()

        # recouvrement + relances (si existe)
        if hasattr(insc, "recouvrement"):
            insc.recouvrement.relances.all().delete()
            insc.recouvrement.delete()

        # paiements (PROTECT -> delete avant inscription)
        insc.paiements.all().delete()

        # échéances (et paiements.echeance PROTECT -> ok car paiements déjà supprimés)
        insc.echeances.all().delete()

        # delete inscription
        insc.delete()

    # 3) Delete final
    eleve.delete()

@login_required
@permission_required("core.delete_eleve", raise_exception=True)
def eleve_delete(request, pk):
    eleve = get_object_or_404(Eleve, pk=pk)

    nb_inscriptions = eleve.inscriptions.count()
    nb_paiements = Paiement.objects.filter(inscription__eleve=eleve).count()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()

        # ✅ 1) ARCHIVER
        if action == "archive":
            if not eleve.is_active:
                messages.info(request, "Cet élève est déjà archivé.")
                return redirect("core:eleve_detail", pk=eleve.pk)

            eleve.archive(by_user=request.user)
            messages.success(request, "Élève archivé avec succès.")
            return redirect("core:eleve_detail", pk=eleve.pk)

        # ✅ 2) RESTAURER
        if action == "restore":
            if eleve.is_active:
                messages.info(request, "Cet élève est déjà actif.")
                return redirect("core:eleve_detail", pk=eleve.pk)

            eleve.restore()
            messages.success(request, "Élève restauré avec succès.")
            return redirect("core:eleve_detail", pk=eleve.pk)

        # ⚠️ 3) HARD DELETE (définitif) — double sécurité
        if action == "hard":
            confirm = (request.POST.get("confirm") or "").strip()
            if confirm != eleve.matricule:
                messages.error(request, f"Confirmation invalide. Tape exactement: {eleve.matricule}")
                return redirect("core:eleve_delete", pk=eleve.pk)

            try:
                with transaction.atomic():
                    parents_ids = list(eleve.liens_parents.values_list("parent_id", flat=True).distinct())
                    matricule = eleve.matricule

                    _eleve_hard_delete(eleve)

                    # cleanup parents orphelins
                    User = get_user_model()
                    parents = Parent.objects.filter(id__in=parents_ids).select_related("user")

                    for parent in parents:
                        if parent.liens.exists():
                            continue
                        user = parent.user
                        parent.delete()

                        if user:
                            if Enseignant.objects.filter(user=user).exists():
                                user.is_active = False
                                user.save(update_fields=["is_active"])
                            else:
                                user.delete()

                messages.success(request, f"Élève {matricule} supprimé définitivement.")
                return redirect("core:eleve_list")

            except ProtectedError:
                # fallback: archive
                eleve.archive(by_user=request.user)
                messages.error(request, "Suppression bloquée (dépendances protégées). Élève archivé.")
                return redirect("core:eleve_detail", pk=eleve.pk)

        messages.error(request, "Action invalide.")
        return redirect("core:eleve_detail", pk=eleve.pk)

    return render(request, "admin/eleves/delete.html", {
        "eleve": eleve,
        "nb_inscriptions": nb_inscriptions,
        "nb_paiements": nb_paiements,
    })


# ============================
# C3 — Inscriptions
# ============================


@login_required
@permission_required("core.add_inscription", raise_exception=True)
def inscription_full_create(request):
    """
    1 seul écran:
    - Eleve + Parent + Inscription
    - Lien ParentEleve
    - Compte ELEVE créé automatiquement
    - Compte parent : optionnel (pas auto)
    - Photo élève supportée (request.FILES)
    """
    if request.method == "POST":
        # ✅ IMPORTANT: toujours passer request.FILES
        form = InscriptionFullForm(request.POST, request.FILES)

        if form.is_valid():
            with transaction.atomic():
                eleve, parent, insc = form.save()

                # ✅ Créer COMPTE ELEVE (username = matricule élève)
                try:
                    user, pwd, created = get_or_create_user_with_group(eleve.matricule, "ELEVE")

                    if eleve.user_id != user.id:
                        eleve.user = user
                        eleve.save(update_fields=["user"])

                    if created:
                        messages.success(
                            request,
                            f"✅ Inscription créée: {eleve.matricule} — Groupe: {insc.groupe.nom} | "
                            f"Compte ÉLÈVE créé | MDP: {pwd}"
                        )
                    else:
                        messages.success(
                            request,
                            f"✅ Inscription créée: {eleve.matricule} — Groupe: {insc.groupe.nom} | "
                            f"Compte ÉLÈVE déjà existant"
                        )

                except Exception as e:
                    messages.warning(
                        request,
                        f"✅ Inscription créée: {eleve.matricule} — Groupe: {insc.groupe.nom} | "
                        f"⚠️ Compte ÉLÈVE non créé: {e}"
                    )

            return redirect("core:eleve_detail", pk=eleve.pk)

    else:
        active = AnneeScolaire.objects.filter(is_active=True).first()
        form = InscriptionFullForm(initial={"annee": active} if active else None)

    return render(request, "admin/inscriptions/full_form.html", {"form": form})

@login_required
@permission_required("core.add_inscription", raise_exception=True)
def inscription_create(request):
    if request.method == "POST":
        form = InscriptionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Inscription créée.")
            return redirect("core:inscription_list")
    else:
        form = InscriptionForm()
    return render(request, "admin/inscriptions/form.html", {"form": form, "mode": "create"})


@login_required
@permission_required("core.view_inscription", raise_exception=True)
def inscription_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    # ✅ nouveaux filtres
    q = request.GET.get("q", "").strip()
    statut = request.GET.get("statut", "").strip()

    annee_id = request.GET.get("annee", "")
    niveau_id = request.GET.get("niveau", "")
    groupe_id = request.GET.get("groupe", "")
    periode_id = request.GET.get("periode", "")
    date_str = request.GET.get("date", "").strip()

    inscriptions = Inscription.objects.select_related(
        "eleve", "annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "periode"
    )

    # ✅ par défaut : année active
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    if annee_id:
        inscriptions = inscriptions.filter(annee_id=annee_id)
    if niveau_id:
        inscriptions = inscriptions.filter(groupe__niveau_id=niveau_id)
    if groupe_id:
        inscriptions = inscriptions.filter(groupe_id=groupe_id)
    if periode_id:
        inscriptions = inscriptions.filter(periode_id=periode_id)

    # ✅ date inscription
    if date_str:
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            inscriptions = inscriptions.filter(date_inscription=d)
        except ValueError:
            inscriptions = inscriptions.none()

    # ✅ filtre statut
    if statut in ("EN_COURS", "VALIDEE"):
        inscriptions = inscriptions.filter(statut=statut)

    # ✅ recherche (élève)
    if q:
        inscriptions = inscriptions.filter(
            Q(eleve__matricule__icontains=q) |
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__telephone__icontains=q) |
            Q(groupe__nom__icontains=q) |
            Q(groupe__niveau__nom__icontains=q) |
            Q(groupe__niveau__degre__nom__icontains=q)
        )

    # ✅ dropdowns
    annees = AnneeScolaire.objects.all()

    niveaux = Niveau.objects.all()
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    groupes = Groupe.objects.all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    periodes = Periode.objects.all()
    if annee_id:
        periodes = periodes.filter(annee_id=annee_id)

    return render(request, "admin/inscriptions/list.html", {
        "inscriptions": inscriptions,
        "annee_active": annee_active,

        "annees": annees,
        "annee_selected": annee_id,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "periodes": periodes,
        "periode_selected": periode_id,

        "date_selected": date_str,

        # ✅ NEW
        "q": q,
        "statut_selected": statut,
    })


@login_required
@permission_required("core.add_eleve", raise_exception=True)
def inscription_create_for_eleve(request, eleve_id):
    eleve = get_object_or_404(Eleve, pk=eleve_id)

    if request.method == "POST":
        form = InscriptionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Inscription créée pour cet élève.")
            return redirect("core:eleve_detail", pk=eleve.id)
    else:
        form = InscriptionForm(initial={"eleve": eleve})

    return render(
        request,
        "admin/inscriptions/form.html",
        {"form": form, "mode": "create", "eleve": eleve},
    )


@login_required
@permission_required("core.change_inscription", raise_exception=True)
def inscription_update(request, pk):
    insc = get_object_or_404(Inscription, pk=pk)
    if request.method == "POST":
        form = InscriptionForm(request.POST, instance=insc)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Inscription mise à jour.")
            return redirect("core:inscription_list")
    else:
        form = InscriptionForm(instance=insc)

    return render(request, "admin/inscriptions/form.html", {"form": form, "mode": "update", "insc": insc})


@login_required
@permission_required("core.delete_inscription", raise_exception=True)
def inscription_delete(request, pk):
    insc = get_object_or_404(Inscription, pk=pk)
    if request.method == "POST":
        insc.delete()
        messages.success(request, "🗑️ Inscription supprimée.")
        return redirect("core:inscription_list")
    return render(request, "admin/inscriptions/delete.html", {"insc": insc})

@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def groupes_par_annee(request):
    """
    Retourne la liste des groupes pour une année donnée (JSON).
    URL: /core/api/groupes/?annee_id=XX
    """
    annee_id = request.GET.get("annee_id")
    if not annee_id:
        return JsonResponse({"results": []})

    groupes = (
        Groupe.objects
        .filter(annee_id=annee_id)
        .select_related("niveau", "niveau__degre")
        .order_by("niveau__degre__ordre", "niveau__ordre", "nom")
    )

    data = []
    for g in groupes:
        label = f"{g.niveau.degre.nom} / {g.niveau.nom} / {g.nom}"
        data.append({"id": g.id, "label": label})

    return JsonResponse({"results": data})

# =========================
# PAIEMENTS — LIST + FILTRES
# =========================


# =========================================================
# AJAX — fratrie
# =========================================================
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@login_required
def ajax_fratrie(request):
    eleve_id_raw = (request.GET.get("eleve") or "").strip()
    if not eleve_id_raw.isdigit():
        return JsonResponse({"ok": False, "error": "eleve manquant"}, status=400)
    eleve_id = int(eleve_id_raw)

    parent_ids = ParentEleve.objects.filter(eleve_id=eleve_id).values_list("parent_id", flat=True)
    if not parent_ids:
        return JsonResponse({"ok": True, "items": []})

    freres = (
        Eleve.objects
        .filter(liens_parents__parent_id__in=list(parent_ids))
        .exclude(id=eleve_id)
        .filter(is_active=True)  # ✅ فقط actifs (change le champ si besoin)
        .distinct()
        .order_by("nom", "prenom")
    )

    items = [
        {"id": e.id, "matricule": e.matricule, "nom": e.nom, "prenom": e.prenom}
        for e in freres
    ]
    return JsonResponse({"ok": True, "items": items})

# =========================================================
# LIST
# =========================================================

from collections import OrderedDict

def _safe_payeur_label(tx):
    """
    Essaie de construire un label "Parent" si tu as des champs payeur_*.
    Sinon fallback: "Paiement parent" + token.
    ✅ Ne casse jamais.
    """
    nom = getattr(tx, "payeur_nom", "") or ""
    tel = getattr(tx, "payeur_telephone", "") or getattr(tx, "payeur_tel", "") or ""
    nom = str(nom).strip()
    tel = str(tel).strip()

    if nom and tel:
        return f"{nom} — {tel}"
    if nom:
        return nom
    if tel:
        return tel
    # fallback sur token / id
    if tx.batch_token:
        return f"Parent (batch {tx.batch_token})"
    return f"Parent (TX-{tx.id})"

from django.core.exceptions import FieldError
from core.models import ParentEleve  # tu l'as dans ton projet

def _format_full_name(nom: str, prenom: str) -> str:
    nom = (nom or "").strip()
    prenom = (prenom or "").strip()
    return (f"{nom} {prenom}").strip()

def _extract_parent_label_from_obj(obj) -> str:
    if not obj:
        return ""

    # cas Parent(nom, prenom)
    nom = getattr(obj, "nom", "") or ""
    prenom = getattr(obj, "prenom", "") or ""
    full = (f"{nom} {prenom}").strip()
    if full:
        return full

    # cas User(first_name, last_name)
    first = getattr(obj, "first_name", "") or ""
    last = getattr(obj, "last_name", "") or ""
    full = (f"{last} {first}").strip()   # "NOM Prénom"
    if full:
        return full

    return ""

def _build_parent_map_for_txs(txs_qs):
    """
    Retourne dict: { eleve_id: "NOM Prénom" }
    ✅ Compatible avec TON ParentEleve (parent, eleve, created_by, updated_by)
    """
    eleve_ids = list(
        txs_qs.values_list("inscription__eleve_id", flat=True).distinct()
    )
    if not eleve_ids:
        return {}

    # ✅ on charge parent + eleve (les seuls champs FK utiles)
    qs = (
        ParentEleve.objects
        .filter(eleve_id__in=eleve_ids)
        .select_related("parent", "eleve")
    )

    parent_map = {}

    for pe in qs:
        # ParentEleve.parent => peut être un modèle Parent (nom/prenom)
        # OU un User (first_name/last_name) selon ton implémentation
        parent_obj = getattr(pe, "parent", None)

        label = _extract_parent_label_from_obj(parent_obj)

        # fallback: si ParentEleve stocke nom/prenom direct (rare) :
        if not label:
            label = _extract_parent_label_from_obj(pe)

        if label:
            parent_map[pe.eleve_id] = label

    return parent_map

def _payeur_label_for_tx(tx, parent_map: dict) -> str:
    """
    Affiche Nom Prénom du parent si trouvé via l'élève.
    Sinon fallback batch / TX.
    """
    eleve_id = getattr(tx.inscription, "eleve_id", None)
    if eleve_id and eleve_id in parent_map:
        return parent_map[eleve_id]

    if tx.batch_token:
        return f"Parent (batch {tx.batch_token})"
    return f"Parent (TX-{tx.id})"


# =========================================================
# LISTE PAR PARENT (group by batch_token)
# =========================================================


@login_required
@permission_required("core.view_paiement", raise_exception=True)
def paiement_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    q = request.GET.get("q", "").strip()
    mode = request.GET.get("mode", "").strip()
    periode_id = request.GET.get("periode", "").strip()
    annee_id = request.GET.get("annee", "").strip()
    niveau_id = request.GET.get("niveau", "").strip()
    groupe_id = request.GET.get("groupe", "").strip()

    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    # ===== base qs =====
    txs = (
        TransactionFinance.objects
        .select_related(
            "parent",
            "inscription",
            "inscription__eleve",
            "inscription__annee",
            "inscription__groupe",
            "inscription__groupe__niveau",
            "inscription__groupe__niveau__degre",
            "inscription__periode",
        )
        .prefetch_related("remboursements")
        .annotate(
            montant_rembourse=Coalesce(
                Sum("remboursements__montant"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
        )
        .annotate(
            has_annulation=Exists(
                RemboursementFinance.objects.filter(
                    transaction_id=OuterRef("pk"),
                    is_annulation=True
                )
            )
        )
        .annotate(
            is_rembourse=Case(
                When(
                    montant_total__gt=Value(Decimal("0.00")),
                    montant_rembourse__gte=F("montant_total"),
                    then=Value(True),
                ),
                default=Value(False),
                output_field=BooleanField(),
            ),
            is_rembourse_partiel=Case(
                When(
                    montant_total__gt=Value(Decimal("0.00")),
                    montant_rembourse__gt=Value(Decimal("0.00")),
                    montant_rembourse__lt=F("montant_total"),
                    then=Value(True),
                ),
                default=Value(False),
                output_field=BooleanField(),
            ),
            is_annulee_zero=Case(
                When(
                    montant_total=Value(Decimal("0.00")),
                    has_annulation=True,
                    then=Value(True),
                ),
                default=Value(False),
                output_field=BooleanField(),
            ),
            can_refund=Case(
                # tx=0 => remboursable si pas déjà annulée
                When(
                    montant_total=Value(Decimal("0.00")),
                    has_annulation=False,
                    then=Value(True),
                ),
                # tx>0 => remboursable si reste
                When(
                    montant_total__gt=Value(Decimal("0.00")),
                    montant_rembourse__lt=F("montant_total"),
                    then=Value(True),
                ),
                default=Value(False),
                output_field=BooleanField(),
            ),
        )
    )

    # ===== filtres =====
    if annee_id:
        txs = txs.filter(inscription__annee_id=annee_id)
    if niveau_id:
        txs = txs.filter(inscription__groupe__niveau_id=niveau_id)
    if groupe_id:
        txs = txs.filter(inscription__groupe_id=groupe_id)
    if periode_id:
        txs = txs.filter(inscription__periode_id=periode_id)

    d_from = parse_date(date_from) if date_from else None
    d_to = parse_date(date_to) if date_to else None
    if d_from:
        txs = txs.filter(date_transaction__date__gte=d_from)
    if d_to:
        txs = txs.filter(date_transaction__date__lte=d_to)

    if mode:
        txs = txs.filter(mode=mode)

    if q:
        txs = txs.filter(
            Q(inscription__eleve__matricule__icontains=q) |
            Q(inscription__eleve__nom__icontains=q) |
            Q(inscription__eleve__prenom__icontains=q) |
            Q(inscription__eleve__telephone__icontains=q) |
            Q(inscription__groupe__nom__icontains=q) |
            Q(inscription__groupe__niveau__nom__icontains=q) |
            Q(inscription__groupe__niveau__degre__nom__icontains=q)
        )

    txs = txs.order_by("-date_transaction", "-id")

    # ✅ txs actifs pour les totaux globaux (APRÈS filtres)
    txs_active = txs.filter(inscription__eleve__is_active=True)

    # ===== group by batch_token ou TX-id =====
    groups = OrderedDict()
    parent_map = _build_parent_map_for_txs(txs)

    for tx in txs:
        group_key = str(tx.batch_token) if tx.batch_token else f"TX-{tx.id}"

        if group_key not in groups:
            groups[group_key] = {
                "key": group_key,
                "batch_token": str(tx.batch_token) if tx.batch_token else "",
                "date": tx.date_transaction,
                "mode": tx.mode,
                "mode_label": tx.get_mode_display(),
                "payeur_label": _payeur_label_for_tx(tx, parent_map),

                "total_brut": Decimal("0.00"),
                "total_rembourse": Decimal("0.00"),
                "total_net": Decimal("0.00"),

                "has_partiel": False,
                "all_annulee": True,
                "all_rembourse": True,

                "refund_tx_id": None,
                "can_refund_any": False,

                "txs": [],
            }

        g = groups[group_key]
        g["txs"].append(tx)

        mt = (tx.montant_total or Decimal("0.00"))
        mr = (getattr(tx, "montant_rembourse", None) or Decimal("0.00"))
        g["total_brut"] += mt
        g["total_rembourse"] += mr

        if g["mode"] != tx.mode:
            g["mode"] = "MIXED"
            g["mode_label"] = "Multiple"

        if not getattr(tx, "is_annulee_zero", False):
            g["all_annulee"] = False

        if not getattr(tx, "is_rembourse", False):
            g["all_rembourse"] = False

        if getattr(tx, "is_rembourse_partiel", False):
            g["has_partiel"] = True

        if (g["refund_tx_id"] is None) and getattr(tx, "can_refund", False):
            g["refund_tx_id"] = tx.id
            g["can_refund_any"] = True

    for g in groups.values():
        g["total_net"] = (g["total_brut"] or Decimal("0.00")) - (g["total_rembourse"] or Decimal("0.00"))

        if g["all_annulee"] and g["total_brut"] == Decimal("0.00"):
            g["status_code"] = "ANNULE"
            g["status_label"] = "Annulé"
        elif g["all_rembourse"] and g["total_brut"] > Decimal("0.00"):
            g["status_code"] = "REMBOURSE"
            g["status_label"] = "Remboursé"
        elif g["has_partiel"]:
            g["status_code"] = "PARTIEL"
            g["status_label"] = f"Partiel ({g['total_rembourse']} MAD)"
        else:
            g["status_code"] = "PAYE"
            g["status_label"] = "Payé"

    totaux = txs_active.aggregate(
        total_brut=Coalesce(Sum("montant_total"), Value(Decimal("0.00")),
                            output_field=DecimalField(max_digits=10, decimal_places=2)),
        total_rembourse=Coalesce(Sum("montant_rembourse"), Value(Decimal("0.00")),
                                 output_field=DecimalField(max_digits=10, decimal_places=2)),
    )
    total_brut = totaux["total_brut"]
    total_rembourse = totaux["total_rembourse"]
    total_net = (total_brut or Decimal("0.00")) - (total_rembourse or Decimal("0.00"))

    annees = AnneeScolaire.objects.all().order_by("-date_debut")
    periodes = Periode.objects.all().order_by("nom")
    if annee_id:
        periodes = periodes.filter(annee_id=annee_id)

    niveaux = Niveau.objects.all().order_by("nom")
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    groupes_qs = Groupe.objects.select_related("niveau", "annee").all().order_by("nom")
    if annee_id:
        groupes_qs = groupes_qs.filter(annee_id=annee_id)
    if niveau_id:
        groupes_qs = groupes_qs.filter(niveau_id=niveau_id)

    modes = TransactionFinance._meta.get_field("mode").choices

    # ✅ année affichée dans table : si filtre année -> on affiche cette année
    annee_selected_obj = None
    if annee_id and annee_id.isdigit():
        annee_selected_obj = AnneeScolaire.objects.filter(id=int(annee_id)).first()

    return render(request, "admin/paiements/list.html", {
        "annee_active": annee_selected_obj or annee_active,

        "annees": annees,
        "annee_selected": annee_id,

        "total_brut": total_brut,
        "total_rembourse": total_rembourse,
        "total_net": total_net,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,
        "groupes": groupes_qs,
        "groupe_selected": groupe_id,
        "periodes": periodes,
        "periode_selected": periode_id,

        "q": q,
        "modes": modes,
        "mode_selected": mode,

        "date_from_selected": date_from,
        "date_to_selected": date_to,

        "groups": list(groups.values()),
    })



@login_required
@permission_required("core.view_paiement", raise_exception=True)
def paiements_excel_export(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    # mêmes filtres que ta page
    q = (request.GET.get("q") or "").strip()
    mode = (request.GET.get("mode") or "").strip()
    periode_id = (request.GET.get("periode") or "").strip()
    annee_id = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    txs = (
        TransactionFinance.objects
        .select_related(
            "parent",
            "inscription",
            "inscription__eleve",
            "inscription__annee",
            "inscription__groupe",
            "inscription__groupe__niveau",
            "inscription__groupe__niveau__degre",
            "inscription__periode",
        )
        .prefetch_related("lignes", "lignes__echeance", "lignes__echeance_transport", "remboursements")
        .annotate(
            montant_rembourse=Coalesce(
                Sum("remboursements__montant"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
        )
    )

    # filtres
    if annee_id:
        txs = txs.filter(inscription__annee_id=annee_id)
    if niveau_id:
        txs = txs.filter(inscription__groupe__niveau_id=niveau_id)
    if groupe_id:
        txs = txs.filter(inscription__groupe_id=groupe_id)
    if periode_id:
        txs = txs.filter(inscription__periode_id=periode_id)

    d_from = parse_date(date_from) if date_from else None
    d_to = parse_date(date_to) if date_to else None
    if d_from:
        txs = txs.filter(date_transaction__date__gte=d_from)
    if d_to:
        txs = txs.filter(date_transaction__date__lte=d_to)

    if mode:
        txs = txs.filter(mode=mode)

    if q:
        txs = txs.filter(
            Q(inscription__eleve__matricule__icontains=q) |
            Q(inscription__eleve__nom__icontains=q) |
            Q(inscription__eleve__prenom__icontains=q) |
            Q(inscription__eleve__telephone__icontains=q) |
            Q(inscription__groupe__nom__icontains=q) |
            Q(inscription__groupe__niveau__nom__icontains=q) |
            Q(inscription__groupe__niveau__degre__nom__icontains=q)
        )

    txs = txs.order_by("-date_transaction", "-id")

    # -------------------------
    # helpers
    # -------------------------
    def _mode_label(tx):
        fn = getattr(tx, "get_mode_display", None)
        return str(fn() if callable(fn) else (getattr(tx, "mode", "") or "—"))

    def _type_label(tx):
        # si tu as get_type_transaction_display
        fn = getattr(tx, "get_type_transaction_display", None)
        if callable(fn):
            return str(fn() or "—")
        return str(getattr(tx, "type_transaction", None) or "—")

    def _ref_paiement(tx):
        # référence “4321” etc
        return str(getattr(tx, "reference", "") or "—")

    def _tel_parent(tx):
        p = getattr(tx, "parent", None)
        return (
            getattr(p, "telephone", None)
            or getattr(p, "phone", None)
            or getattr(p, "tel", None)
            or "—"
        )

    def _recu_label(tx):
        # ✅ reçu = receipt_seq (identique dans un batch)
        dt = getattr(tx, "date_transaction", None) or timezone.now()
        year = int(dt.year)
        seq = getattr(tx, "receipt_seq", None)
        if seq:
            return f"AZ-PAY-{year}-{int(seq):04d}"
        return f"AZ-PAY-{year}-TEMP"

    def _months_for_tx(tx):
        months = []
        has_inscription = False

        lignes = tx.lignes.all().order_by("id")  # ordre stable

        for ln in lignes:
            e1 = getattr(ln, "echeance", None)
            e2 = getattr(ln, "echeance_transport", None)

            if e1 and getattr(e1, "mois_nom", None):
                months.append(str(e1.mois_nom))
                continue

            if e2 and getattr(e2, "mois_nom", None):
                months.append(str(e2.mois_nom))
                continue

            # ✅ ligne sans échéance => inscription/pack/autre
            has_inscription = True

        # unique + stable
        out = []
        seen = set()
        for m in months:
            if m and m not in seen:
                seen.add(m)
                out.append(m)

        if has_inscription:
            out.insert(0, "Inscription")

        return ", ".join(out) if out else ("Inscription" if has_inscription else "—")

    # -------------------------
    # Excel
    # -------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paiements"

    headers = [
        "#",
        "Type",
        "Nom",
        "Prenom",
        "Niveau",

        "Inscription (MAD)",
        "Scolarité (mois)",
        "Transport (mois)",

        "Montant (MAD)",
        "Mode",
        "Réf Reçu",
        "Réf Paiement",
        "Tél Parent",
    ]
    ws.append(headers)

    # style header
    header_fill = PatternFill("solid", fgColor="EEF2FF")
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


    def _split_for_tx(tx):
        """
        Retourne:
        - insc_amount: Decimal (somme des lignes sans échéance)
        - sco_months:  "Septembre, Octobre..."
        - tr_months:   "Septembre, Octobre..."
        """
        insc_amount = Decimal("0.00")
        sco = []
        tr = []

        lignes = tx.lignes.all().order_by("id")  # ordre stable

        for ln in lignes:
            e1 = getattr(ln, "echeance", None)
            e2 = getattr(ln, "echeance_transport", None)

            if e1 and getattr(e1, "mois_nom", None):
                sco.append(str(e1.mois_nom))
                continue

            if e2 and getattr(e2, "mois_nom", None):
                tr.append(str(e2.mois_nom))
                continue

            # ✅ pas d'échéance => inscription / pack / autre (sans mois)
            try:
                insc_amount += (getattr(ln, "montant", None) or Decimal("0.00"))
            except Exception:
                pass

        # unique + stable
        def _uniq(lst):
            out, seen = [], set()
            for m in lst:
                if m and m not in seen:
                    seen.add(m)
                    out.append(m)
            return out

        sco = _uniq(sco)
        tr = _uniq(tr)

        return insc_amount, ", ".join(sco) if sco else "", ", ".join(tr) if tr else ""

    from collections import defaultdict

    totaux_mode = defaultdict(Decimal)
    # rows
    i = 0
    for tx in txs:
        eleve = tx.inscription.eleve
        groupe = tx.inscription.groupe
        niveau = getattr(groupe, "niveau", None)

        nom = getattr(eleve, "nom", "") or ""
        prenom = getattr(eleve, "prenom", "") or ""
        niveau_nom = getattr(niveau, "nom", None) or "—"

        brut = tx.montant_total or Decimal("0.00")
        remb = getattr(tx, "montant_rembourse", None) or Decimal("0.00")
        net = brut - remb
        
        mode_lbl = _mode_label(tx)  # "Espèces", "Chèque", "Virement", etc.
        totaux_mode[mode_lbl] += (net or Decimal("0.00"))

        insc_amount, sco_months, tr_months = _split_for_tx(tx)

        i += 1
        ws.append([
            i,
            _type_label(tx),
            nom,
            prenom,
            niveau_nom,

            float(insc_amount),
            sco_months,
            tr_months,

            float(net),
            _mode_label(tx),
            _recu_label(tx),
            _ref_paiement(tx),
            _tel_parent(tx),
        ])

    # -------------------------
    # Totaux par mode (en bas)
    # -------------------------
    ws.append([])  # ligne vide

    title_row = ws.max_row + 1
    ws.append(["", "", "", "", "TOTAUX PAR MODE", "", "", "", "", "", "", "", ""])

    # style titre
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=title_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F8FAFC")

    def _mode_total_row(label, amount):
        r = ws.max_row + 1
        ws.append(["", "", "", "", label, "", "", "", float(amount), "", "", "", ""])
        ws.cell(row=r, column=5).font = Font(bold=True)
        ws.cell(row=r, column=9).font = Font(bold=True)
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="right")

    # ✅ Tes 3 totaux demandés (même si à 0)
    total_especes = totaux_mode.get("Espèces", Decimal("0.00"))
    total_cheque  = totaux_mode.get("Chèque",  Decimal("0.00"))
    total_virement = totaux_mode.get("Virement", Decimal("0.00"))

    _mode_total_row("Total Espèces", total_especes)
    _mode_total_row("Total Chèque", total_cheque)
    _mode_total_row("Total Virement", total_virement)

    total_general = total_especes + total_cheque + total_virement
    _mode_total_row("TOTAL GÉNÉRAL", total_general)

    # auto width
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="paiements.xlsx"'
    wb.save(response)
    return response

# =========================================================
# DETAILS PAIEMENT PARENT (batch)
# =========================================================
@login_required
@permission_required("core.view_paiement", raise_exception=True)
def paiement_parent_detail(request, group_key):
    """
    group_key = UUID string (batch_token) OU "TX-{id}" si paiement solo
    """
    # cas solo
    if group_key.startswith("TX-"):
        try:
            tx_id = int(group_key.replace("TX-", "").strip())
        except ValueError:
            raise Http404
        txs = TransactionFinance.objects.filter(id=tx_id)
    else:
        txs = TransactionFinance.objects.filter(batch_token=group_key)

    txs = (
        txs.select_related(
            "inscription",
            "inscription__eleve",
            "inscription__annee",
            "inscription__groupe",
            "inscription__groupe__niveau",
            "inscription__groupe__niveau__degre",
            "inscription__periode",
        )
        .prefetch_related("remboursements")
        .annotate(
            montant_rembourse=Coalesce(
                Sum("remboursements__montant"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
        )
        .order_by("inscription__eleve__nom", "id")
    )

    if not txs.exists():
        raise Http404

    first = txs.first()
    parent_map = _build_parent_map_for_txs(txs)
    first = txs.first()
    payeur_label = _payeur_label_for_tx(first, parent_map)


    total_brut = sum((tx.montant_total or Decimal("0.00")) for tx in txs)
    total_rembourse = sum((tx.montant_rembourse or Decimal("0.00")) for tx in txs)
    total_net = total_brut - total_rembourse

    return render(request, "admin/paiements/parent_detail.html", {
        "group_key": group_key,
        "batch_token": str(first.batch_token) if first.batch_token else "",
        "date": first.date_transaction,
        "payeur_label": payeur_label,
        "txs": txs,

        "total_brut": total_brut,
        "total_rembourse": total_rembourse,
        "total_net": total_net,
    })

# =========================================================
# CREATE (form unique) — multi-mois via payload
# =========================================================
@login_required
@permission_required("core.add_paiement", raise_exception=True)
def paiement_create(request):
    niveaux = Niveau.objects.select_related("degre").order_by("degre__ordre", "ordre", "nom")

    if request.method == "POST":
        form = PaiementForm(request.POST)

        if not form.is_valid():
            return render(request, "admin/paiements/form.html", {"form": form, "niveaux": niveaux})

        cd = form.cleaned_data
        insc = cd["inscription"]
        nature = (cd.get("nature") or "SCOLARITE").upper()

        # ==========================
        # ✅ MULTI-MOIS (SCOLARITE)
        # ==========================
        ids_int = cd.get("payload_selected_ids") or []
        prices = cd.get("payload_prices") or {}
        is_multi = bool(ids_int)

        if nature == "SCOLARITE" and is_multi:
            batch_token = str(uuid.uuid4())

            echeances = list(
                EcheanceMensuelle.objects.filter(
                    id__in=ids_int,
                    eleve_id=insc.eleve_id,
                    annee_id=insc.annee_id,
                ).order_by("mois_index")
            )

            if len(echeances) != len(ids_int):
                form.add_error(None, "Certaines échéances ne correspondent pas à cet élève / année.")
                return render(request, "admin/paiements/form.html", {"form": form, "niveaux": niveaux})

            try:
                with transaction.atomic():
                    for e in echeances:
                        key = str(e.id)
                        montant = Decimal(str(prices.get(key, "0")).replace(",", "."))

                        if montant <= 0:
                            raise ValidationError(f"Montant invalide pour {e.mois_nom}.")

                        # ✅ interdit si déjà payé
                        if e.statut == "PAYE":
                            raise ValidationError(f"{e.mois_nom} est déjà réglé.")

                        # ✅ Règle A : on enregistre le nouveau prix DU mois
                        e.montant_du = montant

                        # ✅ full payment (pas de partiel)
                        e.montant_paye = montant
                        e.statut = "PAYE"
                        e.save(update_fields=["montant_du", "montant_paye", "statut"])

                        Paiement.objects.create(
                            inscription=insc,
                            nature="SCOLARITE",
                            echeance=e,
                            montant=montant,
                            mode=cd.get("mode"),
                            reference=cd.get("reference") or "",
                            note=cd.get("note") or "",
                            batch_token=batch_token,
                        )


            except ValidationError as ve:
                form.add_error(None, str(ve))
                return render(request, "admin/paiements/form.html", {"form": form, "niveaux": niveaux})

            messages.success(request, "✅ Paiement multi-mois enregistré.")
            return redirect("core:paiement_recu_batch", batch_token=batch_token)

        # ==========================
        # ✅ SINGLE (INSCRIPTION ou SCOLARITE single)
        # ==========================
        p = form.save(commit=False)
        p.inscription = insc
        p.nature = nature
        p.batch_token = ""
        p.save()

        messages.success(request, "✅ Paiement enregistré.")
        return redirect("core:paiement_recu", pk=p.pk)

    # GET
    form = PaiementForm()
    return render(request, "admin/paiements/form.html", {"form": form, "niveaux": niveaux})


# =========================================================
# CREATE for inscription (pré-remplie + verrouillée)
# =========================================================
@login_required
@permission_required("core.add_inscription", raise_exception=True)
def paiement_create_for_inscription(request, inscription_id):
    insc = get_object_or_404(
        Inscription.objects.select_related("eleve", "annee", "groupe", "groupe__niveau", "groupe__niveau__degre"),
        pk=inscription_id
    )

    niveaux = Niveau.objects.select_related("degre").order_by("degre__ordre", "ordre", "nom")

    if request.method == "POST":
        form = PaiementForm(request.POST, inscription=insc)  # ✅ verrouillage ici

        if not form.is_valid():
            return render(request, "admin/paiements/form.html", {"form": form, "niveaux": niveaux, "insc": insc})

        cd = form.cleaned_data
        nature = (cd.get("nature") or "SCOLARITE").upper()

        ids_int = cd.get("payload_selected_ids") or []
        prices = cd.get("payload_prices") or {}
        is_multi = bool(ids_int)

        if nature == "SCOLARITE" and is_multi:
            batch_token = str(uuid.uuid4())

            echeances = list(
                EcheanceMensuelle.objects.filter(
                    id__in=ids_int,
                    eleve_id=insc.eleve_id,
                    annee_id=insc.annee_id,
                ).order_by("mois_index")
            )

            if len(echeances) != len(ids_int):
                form.add_error(None, "Certaines échéances ne correspondent pas à cet élève / année.")
                return render(request, "admin/paiements/form.html", {"form": form, "niveaux": niveaux, "insc": insc})

            try:
                with transaction.atomic():
                    for e in echeances:
                        key = str(e.id)
                        montant = Decimal(str(prices.get(key, "0")))

                        Paiement.objects.create(
                            inscription=insc,
                            nature="SCOLARITE",
                            echeance=e,
                            montant=montant,
                            mode=cd.get("mode"),
                            reference=cd.get("reference") or "",
                            note=cd.get("note") or "",
                            batch_token=batch_token,
                        )

            except ValidationError as ve:
                form.add_error(None, str(ve))
                return render(request, "admin/paiements/form.html", {"form": form, "niveaux": niveaux, "insc": insc})

            messages.success(request, "✅ Paiement multi-mois enregistré.")
            return redirect("core:paiement_recu_batch", batch_token=batch_token)

        # SINGLE
        p = form.save(commit=False)
        p.inscription = insc
        p.nature = nature
        p.batch_token = ""
        p.save()

        messages.success(request, "✅ Paiement enregistré.")
        return redirect("core:paiement_recu", pk=p.pk)

    # GET
    nat = (request.GET.get("nature") or "SCOLARITE").strip().upper()
    initial = {"inscription": insc, "nature": nat}

    form = PaiementForm(initial=initial, inscription=insc)  # ✅ verrouillage ici aussi
    return render(request, "admin/paiements/form.html", {"form": form, "niveaux": niveaux, "insc": insc})


# =========================================================
# RECU single
# =========================================================
@login_required
@permission_required("core.view_paiement", raise_exception=True)
def paiement_recu(request, pk):
    p = get_object_or_404(
        Paiement.objects.select_related(
            "inscription", "inscription__eleve", "inscription__annee",
            "inscription__groupe", "inscription__groupe__niveau", "inscription__groupe__niveau__degre",
            "echeance"
        ),
        pk=pk
    )

    total_paye = p.inscription.paiements.aggregate(
        s=Coalesce(Sum("montant"), Decimal("0.00"), output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"]

    reste = (p.inscription.montant_total or Decimal("0.00")) - (total_paye or Decimal("0.00"))

    return render(request, "admin/paiements/recu.html", {
        "p": p,
        "total_paye": total_paye,
        "reste": reste,
    })


# =========================================================
# RECU batch multi
# =========================================================
@login_required
@permission_required("core.view_paiement", raise_exception=True)
def paiement_recu_batch(request, batch_token):
    qs = (
        Paiement.objects
        .select_related("inscription", "inscription__eleve", "inscription__annee", "inscription__groupe", "echeance")
        .filter(batch_token=batch_token)
        .order_by("echeance__mois_index", "id")
    )
    if not qs.exists():
        raise Http404

    total = sum((p.montant or Decimal("0.00")) for p in qs)
    first = qs.first()

    return render(request, "admin/paiements/recu_batch.html", {
        "paiements": qs,
        "first": first,
        "total": total,
        "batch_token": batch_token,
    })


@login_required
@permission_required("core.view_echeancemensuelle", raise_exception=True)
def impayes_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    annee_id = request.GET.get("annee", "").strip()
    niveau_id = request.GET.get("niveau", "").strip()
    groupe_id = request.GET.get("groupe", "").strip()
    periode_id = request.GET.get("periode", "").strip()
    q = request.GET.get("q", "").strip()

    inscriptions = Inscription.objects.select_related(
        "eleve", "annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "periode"
    )

    # ✅ défaut: année active
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    # ✅ filtres de base (AVANT annotations)
    if annee_id:
        inscriptions = inscriptions.filter(annee_id=annee_id)

    if niveau_id:
        inscriptions = inscriptions.filter(groupe__niveau_id=niveau_id)

    if groupe_id:
        inscriptions = inscriptions.filter(groupe_id=groupe_id)

    if periode_id:
        inscriptions = inscriptions.filter(periode_id=periode_id)

    if q:
        inscriptions = inscriptions.filter(
            Q(eleve__matricule__icontains=q) |
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(groupe__nom__icontains=q) |
            Q(groupe__niveau__nom__icontains=q)
        )

    # ✅ total payé
    inscriptions = inscriptions.annotate(
        total_paye=Coalesce(
            Sum("paiements__montant"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        )
    )

    # ✅ reste
    inscriptions = inscriptions.annotate(
        reste=ExpressionWrapper(
            Coalesce(F("montant_total"), Decimal("0.00")) - F("total_paye"),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        )
    )

    # ✅ vrais impayés
    inscriptions = inscriptions.filter(reste__gt=Decimal("0.00"))

    # ✅ perf: dossier recouvrement (si tu as OneToOne)
    inscriptions = inscriptions.select_related("recouvrement")

    # ✅ KPIs
    total_du = inscriptions.aggregate(
        s=Coalesce(Sum("montant_total"), Decimal("0.00"),
                   output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"] or Decimal("0.00")

    total_encaisse = inscriptions.aggregate(
        s=Coalesce(Sum("total_paye"), Decimal("0.00"),
                   output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"] or Decimal("0.00")

    total_impaye = inscriptions.aggregate(
        s=Coalesce(Sum("reste"), Decimal("0.00"),
                   output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"] or Decimal("0.00")

    # ✅ dropdowns
    annees = AnneeScolaire.objects.all()

    niveaux = Niveau.objects.all()
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    groupes = Groupe.objects.all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    periodes = Periode.objects.all()
    if annee_id:
        periodes = periodes.filter(annee_id=annee_id)

    return render(request, "admin/impayes/list.html", {
        "inscriptions": inscriptions,
        "annees": annees,
        "annee_selected": annee_id,
        "q": q,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,
        "groupes": groupes,
        "groupe_selected": groupe_id,

        "periodes": periodes,
        "periode_selected": periode_id,

        "total_du": total_du,
        "total_encaisse": total_encaisse,
        "total_impaye": total_impaye,
    })


# =========================
# IMPAYÉS MENSUELS (NEW) — scolarité + transport + inscription
# Règle: afficher impayés du mois courant OU avant (mois_index <= mois_courant)
# =========================


@login_required
@permission_required("core.view_echeancemensuelle", raise_exception=True)
def impayes_mensuels_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    if not annee_active:
        return render(request, "admin/impayes/mensuels.html", {
            "annee_active": None,
            "annees": [],
            "niveaux": [],
            "groupes": [],
            "periodes": [],
            "mois_list": list(range(1, 11)),
            "mois_selected": "",
            "type_selected": "ALL",
            "rows": [],
            "kpi": {
                "sco_du": Decimal("0.00"), "sco_paye": Decimal("0.00"), "sco_reste": Decimal("0.00"),
                "tr_du": Decimal("0.00"),  "tr_paye": Decimal("0.00"),  "tr_reste": Decimal("0.00"),
                "ins_du": Decimal("0.00"), "ins_paye": Decimal("0.00"), "ins_reste": Decimal("0.00"),
                "du": Decimal("0.00"), "encaisse": Decimal("0.00"), "impaye": Decimal("0.00"),
            },
            "annee_selected": "",
            "niveau_selected": "",
            "groupe_selected": "",
            "periode_selected": "",
            "q": "",
            "mois_courant": 0,
            "mois_limit": 0,
            "mois_courant_nom": "",
            "mois_limit_nom": "",
        })

    today = timezone.now().date()

    # =========================
    # Filters
    # =========================
    annee_id = (request.GET.get("annee") or "").strip() or str(annee_active.id)
    type_selected = (request.GET.get("type") or "ALL").strip().upper()
    mois_selected = (request.GET.get("mois") or "").strip()
    q = (request.GET.get("q") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    periode_id = (request.GET.get("periode") or "").strip()

    mois_list = list(range(1, 11))
    if mois_selected and (not mois_selected.isdigit() or int(mois_selected) not in mois_list):
        mois_selected = ""

    # année effective
    annee_obj = AnneeScolaire.objects.filter(id=annee_id).first() or annee_active

    # mois courant (1..10)
    idx_courant = mois_index_courant(annee_obj, today)

    # limite : Auto => mois courant, sinon mois choisi
    idx_limit = idx_courant
    if mois_selected and mois_selected.isdigit():
        idx_limit = int(mois_selected)
    idx_limit = max(1, min(10, idx_limit))

    # =========================
    # Dropdowns
    # =========================
    annees = AnneeScolaire.objects.all().order_by("-date_debut")

    niveaux = (
        Niveau.objects.select_related("degre")
        .order_by("degre__ordre", "ordre", "nom")
    )

    groupes_qs = (
        Groupe.objects.select_related("niveau", "niveau__degre")
        .filter(annee_id=annee_obj.id)
    )

    if niveau_id:
        groupes_qs = groupes_qs.filter(niveau_id=niveau_id)
    groupes = groupes_qs.order_by("niveau__degre__ordre", "niveau__ordre", "nom")

    periodes = Periode.objects.filter(annee_id=annee_obj.id).order_by("ordre")

    # =========================
    # Base inscriptions
    # =========================
    inscs = (
        Inscription.objects
        .select_related("eleve", "annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "periode")
        .filter(annee_id=annee_obj.id)
    )

    if niveau_id:
        inscs = inscs.filter(groupe__niveau_id=niveau_id)
    if groupe_id:
        inscs = inscs.filter(groupe_id=groupe_id)
    if periode_id:
        inscs = inscs.filter(periode_id=periode_id)
    if q:
        inscs = inscs.filter(
            Q(eleve__matricule__icontains=q) |
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(groupe__nom__icontains=q) |
            Q(groupe__niveau__nom__icontains=q)
        )

    insc_by_eleve = {i.eleve_id: i for i in inscs}
    eleve_ids = list(insc_by_eleve.keys())

    # =========================
    # Scolarité impayée
    # =========================
    sco_rows = []
    sco_du = sco_paye = sco_reste = Decimal("0.00")

    if type_selected in ["ALL", "SCOLARITE"]:
        sco_qs = (
    EcheanceMensuelle.objects
    .select_related("eleve", "groupe", "annee")
    .filter(annee_id=annee_obj.id, eleve_id__in=eleve_ids, eleve__is_active=True)
    .filter(mois_index__lte=idx_limit)
    .order_by("eleve__matricule", "mois_index")
)

        for e in sco_qs:
            du = e.montant_du or Decimal("0.00")
            paye = e.montant_paye or Decimal("0.00")
            reste = max(du - paye, Decimal("0.00"))
            if reste <= 0:
                continue

            eleve = e.eleve
            insc = insc_by_eleve.get(e.eleve_id)

            sco_rows.append({
                "type": "SCOLARITE",
                "eleve": eleve,
                "inscription": insc,
                "groupe": (insc.groupe if insc else None),
                "mois_index": int(e.mois_index),
                "mois_nom": mois_nom(int(e.mois_index)),
                "date_echeance": e.date_echeance,
                "du": du,
                "paye": paye,
                "reste": reste,
                "statut": e.statut,
                "target_id": e.id,
            })

            sco_du += du
            sco_paye += paye
            sco_reste += reste


    # =========================
    # Transport impayé
    # =========================
    tr_rows = []
    tr_du = tr_paye = tr_reste = Decimal("0.00")

    if type_selected in ["ALL", "TRANSPORT"]:
        tr_qs = (
            EcheanceTransportMensuelle.objects
            .select_related("eleve", "groupe", "annee")
            .filter(annee_id=annee_obj.id, eleve_id__in=eleve_ids, eleve__is_active=True)
            .filter(mois_index__lte=idx_limit)
            .order_by("eleve__matricule", "mois_index")
        )

        for e in tr_qs:
            du = e.montant_du or Decimal("0.00")
            paye = e.montant_paye or Decimal("0.00")
            reste = max(du - paye, Decimal("0.00"))
            if reste <= 0:
                continue

            eleve = e.eleve
            insc = insc_by_eleve.get(e.eleve_id)

            tr_rows.append({
                "type": "TRANSPORT",
                "eleve": eleve,
                "inscription": insc,
                "groupe": (insc.groupe if insc else None),  # on garde ton affichage groupe via inscription
                "mois_index": int(e.mois_index),
                "mois_nom": mois_nom(int(e.mois_index)),
                "date_echeance": e.date_echeance,
                "du": du,
                "paye": paye,
                "reste": reste,
                "statut": e.statut,
                "target_id": e.id,
            })

            tr_du += du
            tr_paye += paye
            tr_reste += reste

    # =========================
    # Inscription impayée
    # =========================
    ins_rows = []
    ins_du = ins_paye = ins_reste = Decimal("0.00")
    if type_selected in ["ALL", "INSCRIPTION"]:
        for insc in inscs.select_related("eleve"):
            du = insc.frais_inscription_du or Decimal("0.00")
            paye = insc.frais_inscription_paye or Decimal("0.00")
            reste = max(du - paye, Decimal("0.00"))
            if reste <= 0:
                continue

            eleve = insc.eleve
            is_active = bool(getattr(eleve, "is_active", True))

            ins_rows.append({
                "type": "INSCRIPTION",
                "eleve": eleve,
                "inscription": insc,
                "groupe": insc.groupe,
                "mois_index": None,
                "mois_nom": "Inscription",
                "date_echeance": insc.date_inscription,
                "du": du,
                "paye": paye,
                "reste": reste,
                "statut": "A_PAYER" if paye <= 0 else ("PARTIEL" if paye < du else "PAYE"),
                "target_id": insc.id,
                "is_active_eleve": is_active,  # ✅ flag pour template
            })

            # ✅ KPI : on n’additionne QUE si élève actif
            if is_active:
                ins_du += du
                ins_paye += paye
                ins_reste += reste


    rows = sco_rows + tr_rows + ins_rows
    def sort_key(r):
        type_rank = {"INSCRIPTION": 0, "SCOLARITE": 1, "TRANSPORT": 2}
        mi = r["mois_index"] if r["mois_index"] is not None else 0
        matricule = (getattr(r["eleve"], "matricule", "") or "")
        return (matricule, type_rank.get(r["type"], 9), mi)


    rows.sort(key=sort_key)

    kpi = {
        "sco_du": sco_du, "sco_paye": sco_paye, "sco_reste": sco_reste,
        "tr_du": tr_du, "tr_paye": tr_paye, "tr_reste": tr_reste,
        "ins_du": ins_du, "ins_paye": ins_paye, "ins_reste": ins_reste,
        "du": (sco_du + tr_du + ins_du),
        "encaisse": (sco_paye + tr_paye + ins_paye),
        "impaye": (sco_reste + tr_reste + ins_reste),
    }

    return render(request, "admin/impayes/mensuels.html", {
        "annee_active": annee_active,
        "annees": annees,
        "niveaux": niveaux,
        "groupes": groupes,
        "periodes": periodes,
        "rows": rows,
        "kpi": kpi,
        "annee_selected": str(annee_obj.id),
        "mois_selected": mois_selected,
        "mois_list": mois_list,
        "type_selected": type_selected,
        "niveau_selected": niveau_id,
        "groupe_selected": groupe_id,
        "periode_selected": periode_id,
        "q": q,
        "mois_courant": idx_courant,
        "mois_limit": idx_limit,
        "mois_courant_nom": mois_nom(idx_courant),
        "mois_limit_nom": mois_nom(idx_limit),
    })


@login_required
@permission_required("core.view_echeancemensuelle", raise_exception=True)
def impayes_mensuels_excel_export(request):
    """
    Export Excel PRO (1 ligne par élève) :
    #, Nom, Prenom, Num Parent, Inscription, M1..M10, Prix total

    ✅ Affiche UNIQUEMENT les impayés (cellules vides si 0)
    ✅ Cache automatiquement les colonnes mois 100% vides
    ✅ Filtres identiques à l'écran
    ✅ Elèves actifs uniquement
    """
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    if not annee_active:
        return HttpResponse("Aucune année active", status=400)

    today = timezone.now().date()

    # =========================
    # Filters (comme page)
    # =========================
    annee_id = (request.GET.get("annee") or "").strip() or str(annee_active.id)
    type_selected = (request.GET.get("type") or "ALL").strip().upper()
    mois_selected = (request.GET.get("mois") or "").strip()
    q = (request.GET.get("q") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    periode_id = (request.GET.get("periode") or "").strip()

    annee_obj = AnneeScolaire.objects.filter(id=annee_id).first() or annee_active

    # mois courant (1..10)
    idx_courant = mois_index_courant(annee_obj, today)

    # limite : mois courant par défaut, sinon mois choisi
    idx_limit = idx_courant
    if mois_selected and mois_selected.isdigit():
        idx_limit = int(mois_selected)
    idx_limit = max(1, min(10, idx_limit))

    # =========================
    # Inscriptions filtrées (base) + élèves actifs
    # =========================
    inscs = (
        Inscription.objects
        .select_related("eleve", "annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "periode")
        .filter(annee_id=annee_obj.id, eleve__is_active=True)
    )

    if niveau_id:
        inscs = inscs.filter(groupe__niveau_id=niveau_id)
    if groupe_id:
        inscs = inscs.filter(groupe_id=groupe_id)
    if periode_id:
        inscs = inscs.filter(periode_id=periode_id)
    if q:
        inscs = inscs.filter(
            Q(eleve__matricule__icontains=q) |
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(groupe__nom__icontains=q) |
            Q(groupe__niveau__nom__icontains=q)
        )

    eleve_ids = list(inscs.values_list("eleve_id", flat=True).distinct())
    if not eleve_ids:
        return HttpResponse("Aucune donnée à exporter (filtres vides).", status=200)

    # =========================
    # Helper: numéro parent (1er parent lié)
    # =========================
    def get_parent_phone(eleve_id: int) -> str:
        lien = (
            ParentEleve.objects
            .select_related("parent")
            .filter(eleve_id=eleve_id)
            .order_by("id")
            .first()
        )
        if not lien or not lien.parent:
            return ""
        p = lien.parent
        # fallback champs possibles
        for attr in ("telephone", "tel", "phone", "gsm", "numero", "numero_tel"):
            val = getattr(p, attr, None)
            if val:
                return str(val)
        return ""

    # =========================
    # AGRÉGATION PRO (1 ligne / élève)
    # =========================
    agg = {}

    def ensure_bucket(eleve):
        if eleve.id not in agg:
            agg[eleve.id] = {
                "eleve": eleve,
                "parent_phone": get_parent_phone(eleve.id),
                "inscription": Decimal("0.00"),
                "mois": {i: Decimal("0.00") for i in range(1, 11)},
                "total": Decimal("0.00"),
            }
        return agg[eleve.id]

    # =========================
    # SCOLARITE (reste par mois)
    # =========================
    if type_selected in ("ALL", "SCOLARITE"):
        sco_qs = (
            EcheanceMensuelle.objects
            .select_related("eleve")
            .filter(
                annee_id=annee_obj.id,
                eleve_id__in=eleve_ids,
                eleve__is_active=True,
                mois_index__lte=idx_limit,
            )
            .order_by("eleve__matricule", "mois_index")
        )

        for e in sco_qs:
            du = e.montant_du or Decimal("0.00")
            paye = e.montant_paye or Decimal("0.00")
            reste = max(du - paye, Decimal("0.00"))
            if reste <= 0:
                continue

            b = ensure_bucket(e.eleve)
            mi = int(e.mois_index)
            if 1 <= mi <= 10:
                b["mois"][mi] += reste
                b["total"] += reste

    # =========================
    # TRANSPORT (reste par mois)
    # =========================
    if type_selected in ("ALL", "TRANSPORT"):
        tr_qs = (
            EcheanceTransportMensuelle.objects
            .select_related("eleve")
            .filter(
                annee_id=annee_obj.id,
                eleve_id__in=eleve_ids,
                eleve__is_active=True,
                mois_index__lte=idx_limit,
            )
            .order_by("eleve__matricule", "mois_index")
        )

        for e in tr_qs:
            du = e.montant_du or Decimal("0.00")
            paye = e.montant_paye or Decimal("0.00")
            reste = max(du - paye, Decimal("0.00"))
            if reste <= 0:
                continue

            b = ensure_bucket(e.eleve)
            mi = int(e.mois_index)
            if 1 <= mi <= 10:
                b["mois"][mi] += reste
                b["total"] += reste

    # =========================
    # INSCRIPTION (reste global)
    # =========================
    if type_selected in ("ALL", "INSCRIPTION"):
        # ici inscs est déjà filtré sur élèves actifs + filtres page
        for insc in inscs.select_related("eleve"):
            du = insc.frais_inscription_du or Decimal("0.00")
            paye = insc.frais_inscription_paye or Decimal("0.00")
            reste = max(du - paye, Decimal("0.00"))
            if reste <= 0:
                continue

            b = ensure_bucket(insc.eleve)
            b["inscription"] += reste
            b["total"] += reste

    # =========================
    # Excel PRO
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Impayes (PRO)"

    headers = (
        ["#", "Nom", "Prenom", "Num Parent", "Inscription"]
        + [f"M{i} ({mois_nom(i)})" for i in range(1, 11)]
        + ["Prix total"]
    )
    ws.append(headers)

    # Style header
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    def to_cell(v: Decimal):
        v = v or Decimal("0.00")
        return "" if v == Decimal("0.00") else float(v)

    # tri nom/prenom
    items = list(agg.items())
    items.sort(key=lambda kv: (
        (getattr(kv[1]["eleve"], "nom", "") or "").lower(),
        (getattr(kv[1]["eleve"], "prenom", "") or "").lower(),
    ))

    i = 0
    for _eid, data in items:
        if (data["total"] or Decimal("0.00")) <= 0:
            continue

        i += 1
        eleve = data["eleve"]

        ws.append(
            [
                i,
                getattr(eleve, "nom", "") or "",
                getattr(eleve, "prenom", "") or "",
                data["parent_phone"] or "",
                to_cell(data["inscription"]),
            ]
            + [to_cell(data["mois"][m]) for m in range(1, 11)]
            + [to_cell(data["total"])]
        )

    # UX Excel
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Largeurs
    widths = {1: 5, 2: 18, 3: 18, 4: 16, 5: 14}
    for col in range(6, 16):  # M1..M10 (col 6..15)
        widths[col] = 14
    widths[16] = 14  # total
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Format nombre seulement si cellule non vide
    for r in range(2, ws.max_row + 1):
        for c in range(5, len(headers) + 1):  # Inscription -> total
            cell = ws.cell(row=r, column=c)
            if cell.value not in ("", None):
                cell.number_format = '#,##0.00'

    # ✅ cacher colonnes mois 100% vides
    first_data_row = 2
    last_row = ws.max_row
    # colonnes mois : M1..M10 = 6..15
    for col in range(6, 16):
        all_empty = True
        for r in range(first_data_row, last_row + 1):
            val = ws.cell(row=r, column=col).value
            if val not in ("", None, 0, 0.0):
                all_empty = False
                break
        if all_empty:
            ws.column_dimensions[get_column_letter(col)].hidden = True

    # Response
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="impayes_groupe_pro.xlsx"'
    wb.save(resp)
    return resp


# ============================
# F2 — Recouvrement (MVP)
# ============================

@login_required
@permission_required("core.view_recouvrement", raise_exception=True)
def recouvrement_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = request.GET.get("annee", "")
    statut = request.GET.get("statut", "")
    q = request.GET.get("q", "").strip()

    sans_relance = request.GET.get("sans_relance", "") 
    
    date_limite = timezone.now().date() - timedelta(days=30)
    
    dossiers = (
        Recouvrement.objects
        .select_related("inscription", "inscription__eleve", "inscription__annee", "inscription__groupe")
        .annotate(
            relances_count=Count("relances", distinct=True),
            last_relance_at=Max("relances__created_at"),
        )
    )
    # défaut: année active
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    if annee_id:
        dossiers = dossiers.filter(inscription__annee_id=annee_id)

    if statut:
        dossiers = dossiers.filter(statut=statut)

    if q:
        dossiers = dossiers.filter(
            Q(inscription__eleve__matricule__icontains=q) |
            Q(inscription__eleve__nom__icontains=q) |
            Q(inscription__eleve__prenom__icontains=q) |
            Q(inscription__groupe__nom__icontains=q)
        )

    if sans_relance == "1":
        dossiers = dossiers.filter(relances_count=0)
    # KPIs
    total_dossiers = dossiers.count()

    annees = AnneeScolaire.objects.all()

    return render(request, "admin/recouvrements/list.html", {
        "dossiers": dossiers,
        "annees": annees,
        "annee_selected": annee_id,
        "statut_selected": statut,
        "q": q,
        "total_dossiers": total_dossiers,
        "STATUT_CHOICES": Recouvrement.STATUT_CHOICES,
    })


@login_required
@permission_required("core.add_inscription", raise_exception=True)
def recouvrement_create_for_inscription(request, inscription_id):
    insc = get_object_or_404(
        Inscription.objects.select_related("eleve", "annee", "groupe"),
        pk=inscription_id
    )

    # Bloquer si déjà existant
    if hasattr(insc, "recouvrement"):
        messages.info(request, "ℹ️ Un dossier de recouvrement existe déjà pour cette inscription.")
        return redirect("core:recouvrement_detail", pk=insc.recouvrement.pk)

    # Calcul solde via annotation simple
    total_paye = insc.paiements.aggregate(
        s=Coalesce(Sum("montant"), Decimal("0.00"), output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"] or Decimal("0.00")
    solde = (insc.montant_total or Decimal("0.00")) - total_paye

    if solde <= Decimal("0.00"):
        messages.warning(request, "✅ Cette inscription est déjà réglée. Aucun dossier nécessaire.")
        return redirect("core:impayes_list")

    # Création dossier
    dossier = Recouvrement.objects.create(
        inscription=insc,
        statut="EN_COURS",
        date_ouverture=timezone.now().date(),
    )

    messages.success(request, "✅ Dossier de recouvrement créé.")
    return redirect("core:recouvrement_detail", pk=dossier.pk)


@login_required
@permission_required("core.view_recouvrement", raise_exception=True)
def recouvrement_detail(request, pk):
    dossier = get_object_or_404(
        Recouvrement.objects.select_related(
            "inscription", "inscription__eleve", "inscription__annee", "inscription__groupe"
        ),
        pk=pk
    )

    # Totaux
    total_paye = dossier.total_paye
    solde = dossier.solde

    # Auto réglé si solde = 0
    dossier.refresh_statut_si_regle(save=True)

    relances = dossier.relances.select_related("created_by").all()
    paiements = dossier.inscription.paiements.all().order_by("-date_paiement", "-id")

    # Parents liés
    liens_parents = ParentEleve.objects.select_related("parent").filter(eleve=dossier.inscription.eleve)

    return render(request, "admin/recouvrements/detail.html", {
        "dossier": dossier,
        "relances": relances,
        "paiements": paiements,
        "liens_parents": liens_parents,
        "total_paye": total_paye,
        "solde": solde,
        "TYPE_CHOICES": Relance.TYPE_CHOICES,
    })


@login_required
@permission_required("core.add_relance", raise_exception=True)
def relance_create(request, dossier_id):
    dossier = get_object_or_404(Recouvrement, pk=dossier_id)

    if request.method == "POST":
        type_ = (request.POST.get("type") or "SMS").strip().upper()
        message_txt = (request.POST.get("message") or "").strip()

        # sécurité type
        allowed = [t[0] for t in Relance.TYPE_CHOICES]
        if type_ not in allowed:
            type_ = "SMS"

        Relance.objects.create(
            recouvrement=dossier,
            type=type_,
            message=message_txt
        )

        # passage en relance si pas réglé/clôturé
        if dossier.statut not in ["REGLE", "CLOTURE"]:
            dossier.statut = "EN_RELANCE"
            dossier.save(update_fields=["statut", "updated_at", "updated_by"])

        messages.success(request, "✅ Relance ajoutée.")
        return redirect("core:recouvrement_detail", pk=dossier.pk)

    # GET => renvoie vers detail (on utilise le form intégré dans la page)
    return redirect("core:recouvrement_detail", pk=dossier.pk)


@login_required
@permission_required("core.change_recouvrement", raise_exception=True)
def recouvrement_cloturer(request, pk):
    dossier = get_object_or_404(Recouvrement, pk=pk)

    if request.method == "POST":
        dossier.statut = "CLOTURE"
        dossier.date_cloture = timezone.now().date()
        dossier.save(update_fields=["statut", "date_cloture", "updated_at", "updated_by"])
        messages.success(request, "✅ Dossier clôturé.")
        return redirect("core:recouvrement_detail", pk=dossier.pk)

    return redirect("core:recouvrement_detail", pk=dossier.pk)

# ============================
# E1 — Enseignants
# ============================
# core/views.py
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.forms import modelformset_factory
from django.shortcuts import get_object_or_404, redirect, render

from .models import Enseignant, EnseignantGroupe
from .forms import EnseignantForm, EnseignantGroupeForm


@login_required
@permission_required("core.view_enseignant", raise_exception=True)
def enseignant_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    matieres = Matiere.objects.filter(is_active=True).order_by("nom")

    # --- GET (strings) ---
    annee_id  = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    periode_id = (request.GET.get("periode") or "").strip()
    matiere_id = (request.GET.get("matiere") or "").strip()  # ✅ optionnel si tu ajoutes un filtre matière

    q = (request.GET.get("q") or "").strip()
    statut = (request.GET.get("statut") or "").strip()  # "actifs" / "inactifs" / ""

    # --- année par défaut (pour UI) ---
    if not annee_id.isdigit():
        annee_id = str(annee_active.id) if annee_active else ""

    enseignants = Enseignant.objects.all()

    # ✅ Statut
    if statut == "actifs":
        enseignants = enseignants.filter(is_active=True)
    elif statut == "inactifs":
        enseignants = enseignants.filter(is_active=False)

    # ✅ Recherche
    if q:
        enseignants = enseignants.filter(
            Q(matricule__icontains=q) |
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(telephone__icontains=q) |
            Q(email__icontains=q)
        )

    # =========================================================
    # ✅ FILTRES PEDAGOGIQUES (via EnseignantGroupe)
    # =========================================================
    # On filtre via affectations seulement si l'utilisateur a choisi un filtre pédagogique
    filter_via_affectations = False
    aff_qs = EnseignantGroupe.objects.all()

    if annee_id.isdigit():
        aff_qs = aff_qs.filter(annee_id=int(annee_id))

    if niveau_id.isdigit():
        aff_qs = aff_qs.filter(groupe__niveau_id=int(niveau_id))
        filter_via_affectations = True

    if groupe_id.isdigit():
        aff_qs = aff_qs.filter(groupe_id=int(groupe_id))
        filter_via_affectations = True

    if matiere_id.isdigit():
        aff_qs = aff_qs.filter(matiere_fk_id=int(matiere_id))
        filter_via_affectations = True

    # Si on a déclenché un filtrage pédagogique → on restreint les enseignants via affectations
    if filter_via_affectations:
        enseignants = enseignants.filter(id__in=aff_qs.values("enseignant_id"))

    # =========================================================
    # ✅ FILTRE PERIODE (optionnel) : via AbsenceProf ou Seance
    # =========================================================
    # ⚠️ Ce filtre est "activité" et peut exclure des enseignants sans absences/séances.
    periode_obj = None
    if periode_id.isdigit():
        periode_obj = Periode.objects.filter(id=int(periode_id)).first()
        if periode_obj and periode_obj.date_debut and periode_obj.date_fin and annee_id.isdigit():
            # Ici je filtre via AbsenceProf (plus logique que Absence eleves)
            enseignants = enseignants.filter(
                absences_profs__annee_id=int(annee_id),
                absences_profs__date__gte=periode_obj.date_debut,
                absences_profs__date__lte=periode_obj.date_fin,
            )

    enseignants = enseignants.distinct().order_by("nom", "prenom")

    # =========================================================
    # ✅ Listes pour filtres UI
    # =========================================================
    annees = AnneeScolaire.objects.all().order_by("-date_debut")

    niveaux = Niveau.objects.select_related("degre").all().order_by("degre__ordre", "ordre", "nom")
    if annee_id.isdigit():
        niveaux = niveaux.filter(groupes__annee_id=int(annee_id)).distinct()

    groupes = Groupe.objects.select_related("niveau", "niveau__degre", "annee").all().order_by(
        "niveau__degre__ordre", "niveau__ordre", "nom"
    )
    if annee_id.isdigit():
        groupes = groupes.filter(annee_id=int(annee_id))
    if niveau_id.isdigit():
        groupes = groupes.filter(niveau_id=int(niveau_id))

    periodes = Periode.objects.select_related("annee").all().order_by("ordre")
    if annee_id.isdigit():
        periodes = periodes.filter(annee_id=int(annee_id))

    # ✅ Bonus : si tu veux aussi un filtre matière (facultatif)
    # matieres = Matiere.objects.filter(is_active=True).order_by("nom")

    return render(request, "admin/enseignants/list.html", {
        "enseignants": enseignants,

        "annees": annees,
        "annee_selected": annee_id,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "periodes": periodes,
        "periode_selected": periode_id,

        "matieres": matieres,
        "matiere_selected": matiere_id,

        "q": q,
        "statut": statut,
    })

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Q
from django.forms import modelformset_factory
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models.deletion import ProtectedError

from core.models import (
    Enseignant,
    EnseignantGroupe,
    ProfGroupe,          # optionnel (compat)
    AnneeScolaire,
    AbsenceProf,
    Seance,
)
from core.forms import EnseignantForm, EnseignantGroupeForm
from .utils_users import get_or_create_user_with_group


# =========================================================
# Helper: garantir la ligne group-only (matiere_fk NULL)
# =========================================================
def _ensure_group_only(enseignant: Enseignant, annee_id: int, groupe_id: int):
    """
    Crée (si absent) la ligne qui donne accès au groupe côté PROF:
    EnseignantGroupe(matiere_fk = NULL)
    """
    if not (enseignant and annee_id and groupe_id):
        return

    EnseignantGroupe.objects.get_or_create(
        enseignant=enseignant,
        annee_id=annee_id,
        groupe_id=groupe_id,
        matiere_fk=None,
    )


def _sync_prof_groupe(enseignant: Enseignant, groupe_id: int):
    """
    OPTIONNEL (compat): si ton portail prof utilise ProfGroupe quelque part.
    """
    if not (enseignant and enseignant.user_id and groupe_id):
        return

    ProfGroupe.objects.get_or_create(
        user_id=enseignant.user_id,
        groupe_id=groupe_id,
    )


def _cleanup_prof_groupe_if_unused(enseignant: Enseignant, groupe_id: int):
    """
    OPTIONNEL: si plus aucune affectation EnseignantGroupe sur ce groupe,
    on supprime ProfGroupe.
    """
    if not (enseignant and enseignant.user_id and groupe_id):
        return

    still = EnseignantGroupe.objects.filter(
        enseignant=enseignant,
        groupe_id=groupe_id,
    ).exists()

    if not still:
        ProfGroupe.objects.filter(
            user_id=enseignant.user_id,
            groupe_id=groupe_id,
        ).delete()


# =========================================================
# DETAIL
# =========================================================
@login_required
@permission_required("core.view_enseignant", raise_exception=True)
def enseignant_detail(request, pk):
    ens = get_object_or_404(Enseignant, pk=pk)
    annee = AnneeScolaire.objects.filter(is_active=True).first()

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    stats = None
    absences = AbsenceProf.objects.none()

    if annee:
        stats = stats_mensuelles_prof(enseignant=ens, annee=annee, year=year, month=month)
        absences = (
            AbsenceProf.objects
            .filter(annee=annee, enseignant=ens, date__year=year, date__month=month)
            .select_related("seance", "seance__groupe")
            .order_by("-date")
        )

    aff_qs = (
        EnseignantGroupe.objects
        .filter(enseignant=ens)
        .select_related("annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "matiere_fk")
        .order_by("-annee__date_debut", "groupe__niveau__ordre", "groupe__nom", "matiere_fk__nom")
    )

    affectations_active = aff_qs.filter(annee=annee) if annee else aff_qs.none()
    affectations_autres = aff_qs.exclude(annee=annee) if annee else aff_qs

    return render(request, "admin/enseignants/detail.html", {
        "ens": ens,
        "annee": annee,
        "year": year,
        "month": month,
        "stats": stats,
        "absences_profs": absences,
        "affectations_active": affectations_active,
        "affectations_autres": affectations_autres,
    })


# =========================================================
# CREATE
# =========================================================
@login_required
@permission_required("core.add_enseignant", raise_exception=True)
def enseignant_create(request):
    AffectFormSet = modelformset_factory(
        EnseignantGroupe,
        form=EnseignantGroupeForm,
        extra=1,
        can_delete=True
    )

    if request.method == "POST":
        form = EnseignantForm(request.POST, request.FILES)
        formset = AffectFormSet(request.POST, queryset=EnseignantGroupe.objects.none(), prefix="aff")

        if form.is_valid() and formset.is_valid():
            ens = form.save()

            # ✅ assurer user PROF
            try:
                user, pwd, created = get_or_create_user_with_group(ens.matricule, "PROF")
                if ens.user_id != user.id:
                    ens.user = user
                    ens.save(update_fields=["user"])

                if created:
                    messages.success(request, f"✅ Enseignant ajouté: {ens.matricule} | MDP temporaire: {pwd}")
                else:
                    messages.info(request, f"ℹ️ User existe déjà pour {ens.matricule}. Rôle PROF assuré.")
            except Exception as e:
                messages.warning(request, f"⚠️ Enseignant créé mais user non créé: {e}")

            saved = 0

            for f in formset.forms:
                if not getattr(f, "cleaned_data", None):
                    continue
                if f.cleaned_data.get("DELETE"):
                    continue

                annee = f.cleaned_data.get("annee")
                groupe = f.cleaned_data.get("groupe")
                matiere = f.cleaned_data.get("matiere_fk")

                # skip ligne extra vide
                if not annee and not groupe and not matiere:
                    continue

                obj = f.save(commit=False)
                obj.enseignant = ens
                obj.save()
                saved += 1

                # ✅ CRITIQUE: garantir group-only pour accès prof
                if obj.annee_id and obj.groupe_id:
                    _ensure_group_only(ens, obj.annee_id, obj.groupe_id)
                    _sync_prof_groupe(ens, obj.groupe_id)  # optionnel

            if saved:
                messages.success(request, f"✅ {saved} affectation(s) enregistrée(s).")
            else:
                messages.info(request, "ℹ️ Enseignant créé sans affectation.")

            return redirect("core:enseignant_detail", pk=ens.pk)

        messages.error(request, "⚠️ Formulaire invalide. Vérifie les champs et les affectations.")

    else:
        form = EnseignantForm()
        formset = AffectFormSet(queryset=EnseignantGroupe.objects.none(), prefix="aff")

    return render(request, "admin/enseignants/form.html", {
        "form": form,
        "formset": formset,
        "mode": "create",
    })


# =========================================================
# UPDATE
# =========================================================
@login_required
@permission_required("core.change_enseignant", raise_exception=True)
def enseignant_update(request, pk):
    ens = get_object_or_404(Enseignant, pk=pk)

    AffectFormSet = modelformset_factory(
        EnseignantGroupe,
        form=EnseignantGroupeForm,
        extra=1,
        can_delete=True
    )

    qs = (
        EnseignantGroupe.objects
        .filter(enseignant=ens)
        .select_related("annee", "groupe", "matiere_fk")
        .order_by("id")
    )

    if request.method == "POST":
        form = EnseignantForm(request.POST, request.FILES, instance=ens)
        formset = AffectFormSet(request.POST, queryset=qs, prefix="aff")

        if form.is_valid() and formset.is_valid():
            form.save()
            ens.refresh_from_db()

            # ✅ assurer user PROF
            try:
                user, pwd, created = get_or_create_user_with_group(ens.matricule, "PROF")
                if ens.user_id != user.id:
                    ens.user = user
                    ens.save(update_fields=["user"])
            except Exception:
                pass

            saved = 0
            deleted = 0

            # 1) DELETE (ligne par ligne)
            for f in formset.forms:
                if not getattr(f, "cleaned_data", None):
                    continue

                if f.cleaned_data.get("DELETE") and f.instance and f.instance.pk:
                    grp_id = f.instance.groupe_id
                    f.instance.delete()
                    deleted += 1

                    # optionnel cleanup ProfGroupe si plus aucune affectation
                    _cleanup_prof_groupe_if_unused(ens, grp_id)

            # 2) SAVE / CREATE
            for f in formset.forms:
                if not getattr(f, "cleaned_data", None):
                    continue
                if f.cleaned_data.get("DELETE"):
                    continue

                annee = f.cleaned_data.get("annee")
                groupe = f.cleaned_data.get("groupe")
                matiere = f.cleaned_data.get("matiere_fk")

                # skip ligne extra totalement vide
                if not annee and not groupe and not matiere and not (f.instance and f.instance.pk):
                    continue

                obj = f.save(commit=False)
                obj.enseignant = ens
                obj.save()
                saved += 1

                # ✅ CRITIQUE: garantir group-only pour accès prof
                if obj.annee_id and obj.groupe_id:
                    _ensure_group_only(ens, obj.annee_id, obj.groupe_id)
                    _sync_prof_groupe(ens, obj.groupe_id)  # optionnel

            messages.success(request, "✅ Enseignant mis à jour.")
            if saved:
                messages.success(request, f"✅ {saved} affectation(s) enregistrée(s).")
            if deleted:
                messages.success(request, f"🗑️ {deleted} affectation(s) supprimée(s).")

            return redirect("core:enseignant_detail", pk=ens.pk)

        messages.error(request, "⚠️ Formulaire invalide. Vérifie les champs et les affectations.")

    else:
        form = EnseignantForm(instance=ens)
        formset = AffectFormSet(queryset=qs, prefix="aff")

    return render(request, "admin/enseignants/form.html", {
        "form": form,
        "formset": formset,
        "mode": "update",
        "ens": ens,
    })


# =========================================================
# DELETE
# =========================================================
@login_required
@permission_required("core.delete_enseignant", raise_exception=True)
def enseignant_delete(request, pk):
    ens = get_object_or_404(Enseignant, pk=pk)

    seances_bloquantes = (
        Seance.objects
        .filter(enseignant=ens)
        .select_related("annee", "groupe", "groupe__niveau", "groupe__niveau__degre")
        .order_by("jour", "heure_debut")
    )

    if request.method == "POST":
        try:
            ens.delete()
            messages.success(request, "✅ Enseignant supprimé avec succès.")
            return redirect("core:enseignant_list")
        except ProtectedError:
            messages.error(request, "❌ Suppression impossible : cet enseignant est utilisé dans l'emploi du temps.")

    return render(request, "admin/enseignants/delete.html", {
        "ens": ens,
        "seances_bloquantes": seances_bloquantes,
    })



@login_required
@permission_required("core.view_enseignant", raise_exception=True)
def enseignant_affectations(request, pk):
    ens = get_object_or_404(Enseignant, pk=pk)

    affectations = (
        EnseignantGroupe.objects
        .select_related("annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "matiere_fk")
        .filter(enseignant=ens)   # ✅ plus de matiere_fk__isnull
        .order_by("-id")
    )

    form = EnseignantGroupeForm()

    return render(request, "admin/enseignants/affectations.html", {
        "ens": ens,
        "affectations": affectations,
        "form": form,
    })


@login_required
@permission_required("core.add_enseignantgroupe", raise_exception=True)
def enseignant_affectation_add(request, pk):
    ens = get_object_or_404(Enseignant, pk=pk)

    if request.method != "POST":
        return redirect("core:enseignant_affectations", pk=ens.pk)

    form = EnseignantGroupeForm(request.POST)
    if not form.is_valid():
        messages.error(request, "⚠️ Formulaire invalide. Vérifie les champs.")
        return redirect("core:enseignant_affectations", pk=ens.pk)

    annee = form.cleaned_data["annee"]
    groupe = form.cleaned_data["groupe"]
    matiere = form.cleaned_data.get("matiere_fk")

    if not matiere:
        messages.error(request, "⚠️ Choisis une matière.")
        return redirect("core:enseignant_affectations", pk=ens.pk)

    # ✅ 1) GARANTIR la ligne group-only (autorisation portail prof)
    EnseignantGroupe.objects.get_or_create(
        enseignant=ens,
        annee=annee,
        groupe=groupe,
        matiere_fk=None,
    )

    # ✅ 2) Créer/MAJ la ligne groupe + matière
    obj, created = EnseignantGroupe.objects.get_or_create(
        enseignant=ens,
        annee=annee,
        groupe=groupe,
        matiere_fk=matiere,
    )

    if created:
        messages.success(request, "✅ Groupe + matière affectés.")
    else:
        messages.success(request, "✅ Affectation déjà existante (OK).")

    return redirect("core:enseignant_affectations", pk=ens.pk)

@login_required
@permission_required("core.delete_enseignantgroupe", raise_exception=True)
def enseignant_affectation_delete(request, pk, aff_id):
    ens = get_object_or_404(Enseignant, pk=pk)
    aff = get_object_or_404(EnseignantGroupe, pk=aff_id, enseignant=ens)

    if request.method == "POST":
        aff.delete()
        messages.success(request, "🗑️ Affectation supprimée.")

    return redirect("core:enseignant_affectations", pk=ens.pk)

# ============================
# E2 — Emploi du temps (Séances)
# ============================


def _jour_code_from_date(date_str: str):
    """
    Convertit une date 'YYYY-MM-DD' vers code jour Seance (LUN..SAM).
    Retourne None si invalide.
    """
    if not date_str:
        return None
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None

    # weekday(): lundi=0 ... dimanche=6
    mapping = {0: "LUN", 1: "MAR", 2: "MER", 3: "JEU", 4: "VEN", 5: "SAM"}
    return mapping.get(d.weekday())  # dimanche => None


@login_required
@permission_required("core.view_seance", raise_exception=True)
def seance_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    annee_id = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    enseignant_id = (request.GET.get("enseignant") or "").strip()
    date_str = (request.GET.get("date") or "").strip()
    q = (request.GET.get("q") or "").strip()

    # ✅ année par défaut = active
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    jour_code = _jour_code_from_date(date_str)

    # ✅ Queryset de base
    seances = Seance.objects.select_related(
        "annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "enseignant"
    )

    # ✅ IMPORTANT : appliquer le filtre année (sinon => toutes les années)
    if annee_id and annee_id.isdigit():
        seances = seances.filter(annee_id=int(annee_id))

    # ✅ Filtre niveau/groupe (via groupe)
    if niveau_id and niveau_id.isdigit():
        seances = seances.filter(groupe__niveau_id=int(niveau_id))

    if groupe_id and groupe_id.isdigit():
        seances = seances.filter(groupe_id=int(groupe_id))

    # ✅ Filtre enseignant
    if enseignant_id and enseignant_id.isdigit():
        seances = seances.filter(enseignant_id=int(enseignant_id))

    # ✅ Filtre DATE -> convertit en jour (LUN/MAR/...)
    if date_str:
        if jour_code:
            seances = seances.filter(jour=jour_code)
        else:
            seances = seances.none()

    # ✅ Recherche
    if q:
        seances = seances.filter(
            Q(matiere__icontains=q) |
            Q(salle__icontains=q) |
            Q(groupe__nom__icontains=q) |
            Q(enseignant__nom__icontains=q) |
            Q(enseignant__prenom__icontains=q)
        )

    # Listes filtres
    annees = AnneeScolaire.objects.all()

    niveaux = Niveau.objects.select_related("degre").all()
    if annee_id and annee_id.isdigit():
        niveaux = niveaux.filter(groupes__annee_id=int(annee_id)).distinct()

    groupes = Groupe.objects.select_related("niveau", "niveau__degre", "annee").all()
    if annee_id and annee_id.isdigit():
        groupes = groupes.filter(annee_id=int(annee_id))
    if niveau_id and niveau_id.isdigit():
        groupes = groupes.filter(niveau_id=int(niveau_id))

    # ✅ Enseignants proposés : uniquement ceux qui ont des séances selon filtres
    ens_qs = Seance.objects.all()
    if annee_id and annee_id.isdigit():
        ens_qs = ens_qs.filter(annee_id=int(annee_id))
    if niveau_id and niveau_id.isdigit():
        ens_qs = ens_qs.filter(groupe__niveau_id=int(niveau_id))
    if groupe_id and groupe_id.isdigit():
        ens_qs = ens_qs.filter(groupe_id=int(groupe_id))
    if date_str and jour_code:
        ens_qs = ens_qs.filter(jour=jour_code)

    enseignants = Enseignant.objects.filter(
        id__in=ens_qs.values_list("enseignant_id", flat=True).distinct()
    )

    return render(request, "admin/seances/list.html", {
        "seances": seances,

        "annees": annees,
        "annee_selected": annee_id,
        "annee_active": annee_active,  # ✅ pour ton template

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "enseignants": enseignants,
        "enseignant_selected": enseignant_id,

        "date_selected": date_str,
        "q": q,
    })


@login_required
@permission_required("core.add_seance", raise_exception=True)
def seance_create(request):
    if request.method == "POST":
        form = SeanceForm(request.POST)
        if form.is_valid():
            s = form.save()
            messages.success(request, "✅ Séance ajoutée.")
            
            action = request.POST.get("action", "save")
            if action == "save_add":
                # on repart sur create avec mêmes filtres
                url = reverse("core:seance_create")
                params = f"?annee={s.annee_id}&groupe={s.groupe_id}&jour={s.jour}"
                return redirect(url + params)

            # sinon retour EDT semaine filtré
            return redirect(reverse("core:edt_week") + f"?annee={s.annee_id}&groupe={s.groupe_id}")
    else:
        initial = {}
        if request.GET.get("annee"):
            initial["annee"] = request.GET.get("annee")
        if request.GET.get("groupe"):
            initial["groupe"] = request.GET.get("groupe")
        if request.GET.get("jour"):
            initial["jour"] = request.GET.get("jour")

        form = SeanceForm(initial=initial)
    return render(request, "admin/seances/form.html", {"form": form, "mode": "create"})


@login_required
@permission_required("core.change_seance", raise_exception=True)
def seance_update(request, pk):
    s = get_object_or_404(Seance, pk=pk)
    if request.method == "POST":
        form = SeanceForm(request.POST, instance=s)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Séance mise à jour.")
            return redirect("core:seance_list")
    else:
        form = SeanceForm(instance=s)
    return render(request, "admin/seances/form.html", {"form": form, "mode": "update", "s": s})


@login_required
@permission_required("core.delete_seance", raise_exception=True)
def seance_delete(request, pk):
    s = get_object_or_404(
        Seance.objects.select_related(
            "annee", "groupe", "groupe__niveau", "groupe__niveau__degre", "enseignant"
        ),
        pk=pk
    )

    absences_bloquantes = (
        Absence.objects
        .filter(seance=s)
        .select_related("eleve", "groupe", "annee")
        .order_by("-date", "eleve__nom", "eleve__prenom")
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "delete").strip().lower()

        # ✅ CAS 1 — suppression normale
        if action == "delete":
            try:
                s.delete()
                messages.success(request, "🗑️ Séance supprimée.")
                return redirect("core:seance_list")

            except ProtectedError:
                messages.error(
                    request,
                    "❌ Suppression impossible : cette séance est liée à des absences. "
                    "Tu peux soit supprimer ces absences, soit cliquer sur 'Forcer la suppression'."
                )

        # ✅ CAS 2 — FORCER : supprime d’abord les absences liées puis la séance
        elif action == "force":
            count_abs = absences_bloquantes.count()

            try:
                with transaction.atomic():
                    # 1) supprimer toutes les absences liées à cette séance
                    absences_bloquantes.delete()

                    # 2) supprimer la séance
                    s.delete()

                messages.success(
                    request,
                    f"🗑️ Séance supprimée (FORCÉ) + {count_abs} absence(s) supprimée(s)."
                )
                return redirect("core:seance_list")

            except ProtectedError:
                # normalement après delete des absences ça passe,
                # mais on garde une sécurité
                messages.error(
                    request,
                    "❌ Forçage impossible : la séance est encore référencée ailleurs."
                )

    return render(request, "admin/seances/delete.html", {
        "s": s,
        "absences_bloquantes": absences_bloquantes,
    })

@login_required
@permission_required("core.view_seance", raise_exception=True)
def edt_week(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    # --- GET (toujours en str) ---
    annee_id = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()

    # --- Année par défaut : année active si annee_id vide/invalide ---
    if not annee_id.isdigit():
        annee_id = str(annee_active.id) if annee_active else ""

    # Listes
    annees = AnneeScolaire.objects.all()

    # ✅ Niveaux (filtrés par année)
    niveaux = Niveau.objects.select_related("degre").all()
    if annee_id.isdigit():
        niveaux = niveaux.filter(groupes__annee_id=int(annee_id)).distinct()

    # ✅ Groupes (filtrés par année + niveau)
    groupes = Groupe.objects.select_related("annee", "niveau", "niveau__degre").all()
    if annee_id.isdigit():
        groupes = groupes.filter(annee_id=int(annee_id))
    if niveau_id.isdigit():
        groupes = groupes.filter(niveau_id=int(niveau_id))

    # ✅ Groupe par défaut = 1er groupe du queryset
    if (not groupe_id.isdigit()) and groupes.exists():
        groupe_id = str(groupes.first().id)

    # ✅ Séances : ne filtre par annee/groupe que si IDs valides
    seances = Seance.objects.select_related("enseignant", "groupe").all()
    if annee_id.isdigit():
        seances = seances.filter(annee_id=int(annee_id))
    if groupe_id.isdigit():
        seances = seances.filter(groupe_id=int(groupe_id))

    days = [
        ("LUN", "Lundi"),
        ("MAR", "Mardi"),
        ("MER", "Mercredi"),
        ("JEU", "Jeudi"),
        ("VEN", "Vendredi"),
        ("SAM", "Samedi"),
    ]

    by_day = {code: [] for code, _ in days}
    for s in seances.order_by("jour", "heure_debut"):
        by_day[s.jour].append(s)

    return render(request, "admin/seances/week.html", {
        "annees": annees,
        "annee_selected": annee_id,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "days": days,
        "by_day": by_day,
    })


# ============================
# F1 — Absences
# ============================

@login_required
@permission_required("core.view_absence", raise_exception=True)
def absences_pratique(request):
    """
    Page 1:
    1) choisir niveau + groupe
    2) choisir date
    3) affiche séances du jour (via api_seances_par_groupe_date)
    """
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = request.GET.get("annee", "") or (str(annee_active.id) if annee_active else "")

    niveau_id = request.GET.get("niveau", "")
    groupe_id = request.GET.get("groupe", "")
    date_str = request.GET.get("date", "")

    if not date_str:
        date_str = timezone.now().date().isoformat()

    annees = AnneeScolaire.objects.all()

    # ✅ niveaux disponibles pour cette année (via groupes)
    niveaux = Niveau.objects.select_related("degre").all()
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    # ✅ groupes filtrés par année (+ niveau si choisi)
    groupes = Groupe.objects.select_related("niveau", "niveau__degre").all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    return render(request, "admin/absences/pratique.html", {
        "annees": annees,
        "annee_selected": annee_id,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "date_selected": date_str,
    })

@login_required
@permission_required("core.view_absence", raise_exception=True)
def absences_feuille(request):
    """
    Page 2:
    Affiche la feuille de présence (JS charge les données via api_feuille_presence)
    """
    seance_id = request.GET.get("seance_id") or request.GET.get("seance")
    date_str = request.GET.get("date")

    if not seance_id or not date_str:
        messages.error(request, "⚠️ Séance ou date manquante.")
        return redirect("core:absences_pratique")

    return render(request, "admin/absences/feuille.html", {
        "seance_id": seance_id,
        "date_selected": date_str,
    })


@require_GET
@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def api_feuille_presence(request):
    """
    GET /core/api/feuille-presence/?seance_id=..&date=YYYY-MM-DD
    Retour:
      seance: infos
      eleves: liste + statut PRESENT/ABSENT/RETARD (calculé depuis Absence)
    """
    seance_id = request.GET.get("seance_id")
    date_str = request.GET.get("date")

    if not seance_id or not date_str:
        return JsonResponse({"error": "missing_params"}, status=400)

    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"error": "invalid_date"}, status=400)

    seance = get_object_or_404(
        Seance.objects.select_related("enseignant", "groupe", "annee"),
        id=seance_id
    )

    # élèves du groupe via inscriptions (année de la séance)
    inscriptions = (
        Inscription.objects
        .select_related("eleve")
        .filter(annee=seance.annee, groupe=seance.groupe)
        .order_by("eleve__nom", "eleve__prenom")
    )
    eleves = [i.eleve for i in inscriptions]

    # absences existantes pour cette date + séance
    abs_qs = Absence.objects.filter(
        annee=seance.annee,
        groupe=seance.groupe,
        seance=seance,
        date=d,
        eleve__in=eleves
    ).values("eleve_id", "type")

    abs_map = {row["eleve_id"]: row["type"] for row in abs_qs}

    # build response
    eleves_data = []
    for e in eleves:
        t = abs_map.get(e.id)
        if t == "RET":
            statut = "RETARD"
        elif t == "ABS":
            statut = "ABSENT"
        else:
            statut = "PRESENT"

        eleves_data.append({
            "id": e.id,
            "matricule": e.matricule,
            "nom": e.nom,
            "prenom": e.prenom,
            "statut": statut,
            "is_active": e.is_active,  # ✅ utile côté JS pour désactiver ligne
        })

    seance_data = {
        "id": seance.id,
        "date": d.isoformat(),
        "jour": seance.jour,
        "heure_debut": seance.heure_debut.strftime("%H:%M") if seance.heure_debut else "",
        "heure_fin": seance.heure_fin.strftime("%H:%M") if seance.heure_fin else "",
        "matiere": seance.matiere or "",
        "salle": seance.salle or "",
        "prof": f"{seance.enseignant.nom} {seance.enseignant.prenom}",
        "groupe": seance.groupe.nom,
        "annee_id": seance.annee_id,
    }

    return JsonResponse({"seance": seance_data, "eleves": eleves_data})


@require_POST
@login_required
@permission_required("core.change_anneescolaire", raise_exception=True)
def api_feuille_presence_save(request):
    """
    POST JSON:
    {
      "seance_id": 15,
      "date": "2025-12-29",
      "items": [{"eleve_id":1,"statut":"PRESENT"|"ABSENT"|"RETARD"}, ...]
    }

    Règle:
    - PRESENT => supprime Absence si existe
    - ABSENT  => Absence(type="ABS") upsert
    - RETARD  => Absence(type="RET") upsert
    """

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "invalid_json"}, status=400)

    seance_id = payload.get("seance_id")
    date_str  = payload.get("date")
    items     = payload.get("items", [])

    if not seance_id or not date_str or not isinstance(items, list):
        return JsonResponse({"ok": False, "error": "missing_params"}, status=400)

    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"ok": False, "error": "invalid_date"}, status=400)

    seance = get_object_or_404(
        Seance.objects.select_related("annee", "groupe"),
        id=seance_id
    )

    # ids élèves soumis
    eleve_ids = []
    for x in items:
        if x.get("eleve_id"):
            try:
                eleve_ids.append(int(x["eleve_id"]))
            except Exception:
                pass

    # sécurité: seulement élèves inscrits dans ce groupe/année
    allowed_ids = set(
        Inscription.objects.filter(
            annee=seance.annee,
            groupe=seance.groupe,
            eleve_id__in=eleve_ids,
            eleve__is_active=True,  # ✅ BLOQUE inactifs
        ).values_list("eleve_id", flat=True)
    )


    saved = 0

    with transaction.atomic():
        for row in items:
            eleve_id = row.get("eleve_id")
            statut = (row.get("statut") or "PRESENT").upper().strip()

            if not eleve_id:
                continue
            try:
                eleve_id = int(eleve_id)
            except Exception:
                continue

            if eleve_id not in allowed_ids:
                continue

            # ✅ IMPORTANT : lookup aligné avec UniqueConstraint (eleve, date, seance)
            lookup = {"eleve_id": eleve_id, "date": d, "seance": seance}

            if statut == "PRESENT":
                Absence.objects.filter(**lookup).delete()
                saved += 1
                continue

            if statut == "ABSENT":
                Absence.objects.update_or_create(
                    **lookup,
                    defaults={
                        "annee": seance.annee,
                        "groupe": seance.groupe,
                        "type": "ABS",
                    }
                )
                saved += 1
                continue

            if statut == "RETARD":
                Absence.objects.update_or_create(
                    **lookup,
                    defaults={
                        "annee": seance.annee,
                        "groupe": seance.groupe,
                        "type": "RET",
                    }
                )
                saved += 1
                continue

            # statut inconnu => ignore

    return JsonResponse({"ok": True, "saved": saved})


@login_required
@permission_required("core.view_absence", raise_exception=True)
def absence_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    annee_id = request.GET.get("annee", "").strip()
    niveau_id = request.GET.get("niveau", "").strip()
    groupe_id = request.GET.get("groupe", "").strip()
    periode_id = request.GET.get("periode", "").strip()
    date_str = request.GET.get("date", "").strip()

    type_sel = request.GET.get("type", "").strip()
    q = request.GET.get("q", "").strip()

    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    absences = Absence.objects.select_related(
        "eleve", "groupe", "groupe__niveau", "groupe__niveau__degre",
        "seance", "seance__enseignant", "annee"
    )

    if annee_id:
        absences = absences.filter(annee_id=annee_id)

    if niveau_id:
        absences = absences.filter(groupe__niveau_id=niveau_id)

    if groupe_id:
        absences = absences.filter(groupe_id=groupe_id)

    if type_sel:
        absences = absences.filter(type=type_sel)

    if date_str:
        absences = absences.filter(date=date_str)

    # ✅ filtre période (2 modes)
    if periode_id:
        p = Periode.objects.filter(id=periode_id).first()
        if p and p.date_debut and p.date_fin:
            # meilleur : par dates
            absences = absences.filter(date__gte=p.date_debut, date__lte=p.date_fin)
        else:
            # fallback : via inscription.periode (si tes inscriptions ont la période)
            absences = absences.filter(
                eleve__inscriptions__annee_id=annee_id,
                eleve__inscriptions__periode_id=periode_id
            ).distinct()

    if q:
        absences = absences.filter(
            Q(eleve__matricule__icontains=q) |
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q)
        )

    annees = AnneeScolaire.objects.all()

    niveaux = Niveau.objects.select_related("degre").all()
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    groupes = Groupe.objects.select_related("niveau", "niveau__degre").all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    periodes = Periode.objects.all()
    if annee_id:
        periodes = periodes.filter(annee_id=annee_id)

    return render(request, "admin/absences/list.html", {
        "absences": absences,

        "annees": annees,
        "annee_selected": annee_id,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "periodes": periodes,
        "periode_selected": periode_id,

        "date_selected": date_str,
        "type_selected": type_sel,
        "q": q,
    })


def _get_annee_active():
    return AnneeScolaire.objects.filter(is_active=True).first()

@login_required
@permission_required("core.add_absence", raise_exception=True)
def absence_create(request):
    annee_active = _get_annee_active()
    if not annee_active:
        messages.error(request, "⚠️ Aucune année scolaire active.")
        return redirect("core:absence_list")

    # ✅ Pré-remplissage depuis l'URL
    initial = {"annee": annee_active.id}
    if request.GET.get("groupe"):
        initial["groupe"] = request.GET.get("groupe")
    if request.GET.get("date"):
        initial["date"] = request.GET.get("date")
    if request.GET.get("seance"):
        initial["seance"] = request.GET.get("seance")

    if request.method == "POST":
        form = AbsenceForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)

            # ✅ sécurité : force année active
            obj.annee = annee_active

            obj.save()
            messages.success(request, "✅ Absence enregistrée.")
            return redirect("core:absence_list")
        # ✅ si invalid: on retombe sur le render avec erreurs
    else:
        form = AbsenceForm(initial=initial)

    return render(request, "admin/absences/form.html", {
        "form": form,
        "mode": "create",
        "annee_active": annee_active,
    })

@login_required
@permission_required("core.change_absence", raise_exception=True)
def absence_update(request, pk):
    a = get_object_or_404(Absence, pk=pk)
    if request.method == "POST":
        form = AbsenceForm(request.POST, instance=a)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Absence mise à jour.")
            return redirect("core:absence_list")
    else:
        form = AbsenceForm(instance=a)
    return render(request, "admin/absences/form.html", {"form": form, "mode": "update", "a": a})


@login_required
@permission_required("core.delete_absence", raise_exception=True)
def absence_delete(request, pk):
    a = get_object_or_404(Absence, pk=pk)
    if request.method == "POST":
        a.delete()
        messages.success(request, "🗑️ Absence supprimée.")
        return redirect("core:absence_list")
    return render(request, "admin/absences/delete.html", {"a": a})

@login_required
@permission_required("core.view_groupe", raise_exception=True)
def api_seances_par_groupe_date(request):
    """
    Retourne les séances (JSON) filtrées par annee + groupe + date.
    URL: /core/api/seances/?annee_id=..&groupe_id=..&date=YYYY-MM-DD
    """
    annee_id = request.GET.get("annee_id")
    groupe_id = request.GET.get("groupe_id")
    date_str = request.GET.get("date")

    if not (annee_id and groupe_id and date_str):
        return JsonResponse({"results": []})

    # mapping jour python -> code modèle
    # Monday=0 .. Sunday=6
    from datetime import datetime
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"results": []})

    jour_map = {0: "LUN", 1: "MAR", 2: "MER", 3: "JEU", 4: "VEN", 5: "SAM"}
    if d.weekday() not in jour_map:
        # dimanche = pas de séances
        return JsonResponse({"results": []})

    jour_code = jour_map[d.weekday()]

    seances = (
        Seance.objects
        .select_related("enseignant")
        .filter(annee_id=annee_id, groupe_id=groupe_id, jour=jour_code)
        .order_by("heure_debut")
    )

    data = []
    for s in seances:
        label = f"{s.heure_debut}-{s.heure_fin} — {s.enseignant.nom} {s.enseignant.prenom}"
        if s.matiere:
            label += f" — {s.matiere}"
        if s.salle:
            label += f" (Salle {s.salle})"
        data.append({"id": s.id, "label": label})

    return JsonResponse({"results": data})


@login_required
@permission_required("core.view_absence", raise_exception=True)
def absences_jour(request):
    from datetime import date as ddate

    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = request.GET.get("annee", "")
    groupe_id = request.GET.get("groupe", "")
    date_str = request.GET.get("date", "")

    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    if not date_str:
        date_str = ddate.today().isoformat()

    absences = Absence.objects.select_related(
        "eleve", "groupe", "seance", "seance__enseignant"
    )
    if annee_id:
        absences = absences.filter(annee_id=annee_id)

    absences = absences.filter(date=date_str)
    if groupe_id:
        absences = absences.filter(groupe_id=groupe_id)

    annees = AnneeScolaire.objects.all()
    groupes = Groupe.objects.all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)

    return render(request, "admin/absences/jour.html", {
        "absences": absences,
        "annees": annees,
        "groupes": groupes,
        "annee_selected": annee_id,
        "groupe_selected": groupe_id,
        "date_selected": date_str,
    })


# ============================
# G1 — Parents + liens (Formset)
# ============================

@login_required
@permission_required("core.view_parent", raise_exception=True)
def parent_list(request):
    q = (request.GET.get("q") or "").strip()
    statut = (request.GET.get("statut") or "").strip()

    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()

    parents = Parent.objects.all()

    # ✅ filtre statut parent
    if statut == "actifs":
        parents = parents.filter(is_active=True)
    elif statut == "inactifs":
        parents = parents.filter(is_active=False)

    # ✅ recherche
    if q:
        parents = parents.filter(
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(telephone__icontains=q) |
            Q(email__icontains=q)
        )

    # ✅ défaut: année active
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    # =========================================================
    # ✅ Filtrage via enfants -> inscriptions VALIDEE/EN_COURS
    # ✅ + élève actif uniquement
    # =========================================================
    if annee_id:
        parents = parents.filter(
            liens__eleve__is_active=True,
            liens__eleve__inscriptions__annee_id=annee_id,
            liens__eleve__inscriptions__statut__in=["VALIDEE", "EN_COURS"],
        )

    if niveau_id:
        parents = parents.filter(
            liens__eleve__inscriptions__groupe__niveau_id=niveau_id,
        )

    if groupe_id:
        parents = parents.filter(
            liens__eleve__inscriptions__groupe_id=groupe_id,
        )

    # ✅ éviter doublons (un parent peut avoir plusieurs enfants)
    parents = parents.distinct()

    # ✅ dropdowns
    annees = AnneeScolaire.objects.all().order_by("-date_debut", "-id")

    niveaux = Niveau.objects.all()
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    groupes = Groupe.objects.all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    return render(request, "admin/parents/list.html", {
        "parents": parents,
        "q": q,
        "statut": statut,

        "annees": annees,
        "annee_selected": annee_id,
        "niveaux": niveaux,
        "niveau_selected": niveau_id,
        "groupes": groupes,
        "groupe_selected": groupe_id,
    })


@login_required
@permission_required("core.view_parent", raise_exception=True)
def parent_detail(request, pk):
    p = get_object_or_404(Parent, pk=pk)

    liens = (
        ParentEleve.objects
        .select_related("eleve")
        .filter(parent=p, eleve__is_active=True)   # ✅ ONLY actifs
        .order_by("eleve__nom", "eleve__prenom")
    )

    return render(request, "admin/parents/detail.html", {
        "p": p,
        "liens": liens,
    })

@login_required
@permission_required("core.add_parent", raise_exception=True)
def parent_create(request):
    if request.method == "POST":
        form = ParentForm(request.POST)
        if form.is_valid():
            p = form.save()
            formset = ParentEleveFormSet(request.POST, instance=p)
            if formset.is_valid():
                formset.save()
                messages.success(request, "✅ Parent ajouté + liens enregistrés.")
                return redirect("core:parent_detail", pk=p.pk)
            else:
                # rollback propre si formset invalide
                p.delete()
        else:
            formset = ParentEleveFormSet(request.POST)
    else:
        form = ParentForm()
        formset = ParentEleveFormSet()

    return render(request, "admin/parents/form.html", {
        "form": form,
        "formset": formset,
        "mode": "create",
    })


@login_required
@permission_required("core.change_parent", raise_exception=True)
def parent_update(request, pk):
    p = get_object_or_404(Parent, pk=pk)

    if request.method == "POST":
        form = ParentForm(request.POST, instance=p)
        formset = ParentEleveFormSet(request.POST, instance=p)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "✅ Parent mis à jour.")
            return redirect("core:parent_detail", pk=p.pk)
    else:
        form = ParentForm(instance=p)
        formset = ParentEleveFormSet(instance=p)

    return render(request, "admin/parents/form.html", {
        "form": form,
        "formset": formset,
        "mode": "update",
        "p": p,
    })


@login_required
@permission_required("core.delete_parent", raise_exception=True)
def parent_delete(request, pk):
    p = get_object_or_404(Parent, pk=pk)

    # (optionnel) compteur pour afficher dans la page de confirmation
    tx_count = TransactionFinance.objects.filter(parent=p).count()

    if request.method == "POST":
        try:
            p.delete()
            messages.success(request, "🗑️ Parent supprimé.")
            return redirect("core:parent_list")

        except ProtectedError:
            messages.error(
                request,
                f"⛔ Impossible de supprimer ce parent : il est lié à {tx_count} transaction(s) finance. "
                f"Supprime/réaffecte d'abord ces transactions, ou désactive le parent."
            )
            return redirect("core:parent_detail", pk=p.pk)  # adapte si ton url est différente

    return render(request, "admin/parents/delete.html", {"p": p, "tx_count": tx_count})


#================================================================

@login_required
@permission_required("core.view_echeancemensuelle", raise_exception=True)
def impayes_pdf_view(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = request.GET.get("annee", "") or (str(annee_active.id) if annee_active else "")

    inscriptions = Inscription.objects.select_related(
        "eleve", "annee", "groupe", "groupe__niveau", "groupe__niveau__degre"
    )

    annee_obj = None
    if annee_id:
        inscriptions = inscriptions.filter(annee_id=annee_id)
        annee_obj = AnneeScolaire.objects.filter(id=annee_id).first()


    inscriptions = inscriptions.annotate(
        total_paye=Coalesce(
            Sum("paiements__montant"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        )
    ).annotate(
        reste=ExpressionWrapper(
            Coalesce(F("montant_total"), Decimal("0.00")) - Coalesce(F("total_paye"), Decimal("0.00")),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        )
    ).filter(
        reste__gt=Decimal("0.00")
    )

    total_du = inscriptions.aggregate(
        s=Coalesce(Sum("montant_total"), Decimal("0.00"),
                output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"]

    total_encaisse = inscriptions.aggregate(
        s=Coalesce(Sum("paiements__montant"), Decimal("0.00"),
                output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"]

    total_impaye = inscriptions.aggregate(
        s=Coalesce(Sum("reste"), Decimal("0.00"),
                output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"]

    return pdf_utils.impayes_pdf(annee_obj, inscriptions, total_du, total_encaisse, total_impaye)

@login_required
@permission_required("core.view_paiement", raise_exception=True)
def paiement_recu_pdf_auto(request, pk: int):
    p = get_object_or_404(
        Paiement.objects.select_related(
            "inscription", "inscription__eleve", "inscription__annee",
            "inscription__groupe", "inscription__groupe__niveau", "inscription__groupe__niveau__degre",
            "echeance"
        ),
        pk=pk
    )

    # total payé (inscription) + reste (global)
    total_paye = p.inscription.paiements.aggregate(
        s=Coalesce(Sum("montant"), Decimal("0.00"), output_field=DecimalField(max_digits=10, decimal_places=2))
    )["s"] or Decimal("0.00")

    reste = p.inscription.total_reste  # (tu as déjà la property)

    # ✅ Si batch => PDF batch
    if p.batch_token:
        qs = (
            Paiement.objects
            .select_related("echeance", "inscription", "inscription__eleve", "inscription__annee", "inscription__groupe")
            .filter(batch_token=p.batch_token)
            .order_by("echeance__mois_index", "id")
        )
        if not qs.exists():
            raise Http404("Batch introuvable")

        total_batch = sum(((x.montant or Decimal("0.00")) for x in qs), Decimal("0.00"))
        first = qs.first()
        return paiement_recu_batch_pdf(first, qs, total_batch, p.batch_token)

    # ✅ Sinon => PDF single
    return paiement_recu_pdf(p, total_paye, reste)


@login_required
@permission_required("core.view_absence", raise_exception=True)
def absences_jour_pdf_view(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = request.GET.get("annee", "") or (str(annee_active.id) if annee_active else "")
    groupe_id = request.GET.get("groupe", "")
    date_str = request.GET.get("date", "")

    absences = Absence.objects.select_related("eleve", "groupe", "seance", "seance__enseignant")
    annee_obj = None
    groupe_label = ""

    if annee_id:
        absences = absences.filter(annee_id=annee_id)
        annee_obj = AnneeScolaire.objects.filter(id=annee_id).first()

    if date_str:
        absences = absences.filter(date=date_str)

    if groupe_id:
        absences = absences.filter(groupe_id=groupe_id)
        g = Groupe.objects.select_related("niveau", "niveau__degre").filter(id=groupe_id).first()
        if g:
            groupe_label = f"{g.niveau.degre.nom}/{g.niveau.nom}/{g.nom}"

    return pdf_utils.absences_jour_pdf(date_str or "—", annee_obj, groupe_label, absences.order_by("eleve__nom"))






@login_required
@permission_required("core.view_parent", raise_exception=True)
def parents_excel_export(request):
    wb = openpyxl.Workbook()

    # Feuille Parents
    ws1 = wb.active
    ws1.title = "Parents"
    headers1 = ["parent_id", "nom", "prenom", "telephone", "email", "adresse", "is_active"]
    ws1.append(headers1)
    for col in range(1, len(headers1)+1):
        c = ws1.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for p in Parent.objects.order_by("nom", "prenom"):
        ws1.append([
            p.id,
            p.nom,
            p.prenom,
            p.telephone or "",
            p.email or "",
            p.adresse or "",
            "1" if p.is_active else "0",
        ])

    # Feuille Liens
    ws2 = wb.create_sheet("Liens")
    headers2 = ["parent_id", "parent_email", "eleve_matricule", "lien"]
    ws2.append(headers2)
    for col in range(1, len(headers2)+1):
        c = ws2.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    liens = ParentEleve.objects.select_related("parent", "eleve").order_by("parent_id")
    for l in liens:
        ws2.append([
            l.parent.id,
            l.parent.email or "",
            l.eleve.matricule,
            l.lien
        ])

    # widths simple
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="parents_liens.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required("core.change_parent", raise_exception=True)
def parents_excel_import(request):
    errors = []
    parents_created = 0
    parents_updated = 0
    liens_created = 0

    if request.method == "POST":
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "⚠️ Aucun fichier uploadé.")
            return redirect("core:parent_list")

        try:
            wb = openpyxl.load_workbook(f)
        except Exception:
            messages.error(request, "⚠️ Fichier Excel invalide.")
            return redirect("core:parent_list")

        if "Parents" not in wb.sheetnames or "Liens" not in wb.sheetnames:
            messages.error(request, "⚠️ Le fichier doit contenir 2 feuilles: Parents + Liens.")
            return redirect("core:parent_list")

        ws1 = wb["Parents"]
        ws2 = wb["Liens"]

        # ---- Parents ----
        headers1 = [str(c.value).strip().lower() if c.value else "" for c in ws1[1]]
        expected1 = ["parent_id", "nom", "prenom", "telephone", "email", "adresse", "is_active"]
        if headers1[:7] != expected1:
            messages.error(request, f"⚠️ Feuille Parents: colonnes attendues {expected1}")
            return redirect("core:parent_list")

        # mapping utile (email -> parent, id -> parent)
        parent_by_email = {}
        parent_by_id = {}

        for idx, row in enumerate(ws1.iter_rows(min_row=2, values_only=True), start=2):
            parent_id, nom, prenom, telephone, email, adresse, is_active = row[:7]

            if not nom or not prenom:
                errors.append(f"Parents ligne {idx}: nom/prenom obligatoires.")
                continue

            nom = str(nom).strip()
            prenom = str(prenom).strip()
            telephone = str(telephone).strip() if telephone else ""
            email = str(email).strip() if email else ""
            adresse = str(adresse).strip() if adresse else ""
            is_active = str(is_active).strip() if is_active is not None else "1"
            active_bool = True if is_active in ["1", "true", "True", "OUI", "oui", "YES", "yes"] else False

            # règle: si parent_id fourni -> update_or_create sur id
            if parent_id:
                try:
                    pid = int(parent_id)
                except Exception:
                    errors.append(f"Parents ligne {idx}: parent_id invalide.")
                    continue

                obj = Parent.objects.filter(id=pid).first()
                if obj:
                    obj.nom = nom
                    obj.prenom = prenom
                    obj.telephone = telephone
                    obj.email = email
                    obj.adresse = adresse
                    obj.is_active = active_bool
                    obj.save()
                    parents_updated += 1
                else:
                    obj = Parent.objects.create(
                        id=pid,
                        nom=nom, prenom=prenom,
                        telephone=telephone, email=email, adresse=adresse,
                        is_active=active_bool
                    )
                    parents_created += 1
            else:
                # sinon, si email existe -> update_or_create par email (sinon create simple)
                if email:
                    obj, created_flag = Parent.objects.update_or_create(
                        email=email,
                        defaults={
                            "nom": nom,
                            "prenom": prenom,
                            "telephone": telephone,
                            "adresse": adresse,
                            "is_active": active_bool,
                        }
                    )
                    parents_created += 1 if created_flag else 0
                    parents_updated += 0 if created_flag else 1
                else:
                    obj = Parent.objects.create(
                        nom=nom, prenom=prenom,
                        telephone=telephone, email=email, adresse=adresse,
                        is_active=active_bool
                    )
                    parents_created += 1

            if obj.email:
                parent_by_email[obj.email.lower()] = obj
            parent_by_id[obj.id] = obj

        # ---- Liens ----
        headers2 = [str(c.value).strip().lower() if c.value else "" for c in ws2[1]]
        expected2 = ["parent_id", "parent_email", "eleve_matricule", "lien"]
        if headers2[:4] != expected2:
            messages.error(request, f"⚠️ Feuille Liens: colonnes attendues {expected2}")
            return redirect("core:parent_list")

        for idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
            parent_id, parent_email, eleve_matricule, lien = row[:4]

            eleve_matricule = str(eleve_matricule).strip() if eleve_matricule else ""
            lien = str(lien).strip().upper() if lien else ""

            if not eleve_matricule or not lien:
                errors.append(f"Liens ligne {idx}: eleve_matricule et lien obligatoires.")
                continue

            if lien not in ["PERE", "MERE", "TUTEUR", "AUTRE"]:
                errors.append(f"Liens ligne {idx}: lien invalide ({lien}).")
                continue

            # retrouver parent
            parent_obj = None
            if parent_id:
                try:
                    parent_obj = parent_by_id.get(int(parent_id)) or Parent.objects.filter(id=int(parent_id)).first()
                except Exception:
                    parent_obj = None

            if not parent_obj and parent_email:
                pe = str(parent_email).strip().lower()
                parent_obj = parent_by_email.get(pe) or Parent.objects.filter(email__iexact=pe).first()

            if not parent_obj:
                errors.append(f"Liens ligne {idx}: parent introuvable (id/email).")
                continue

            eleve = Eleve.objects.filter(matricule=eleve_matricule).first()
            if not eleve:
                errors.append(f"Liens ligne {idx}: élève introuvable ({eleve_matricule}).")
                continue

            obj, created_flag = ParentEleve.objects.update_or_create(
                parent=parent_obj,
                eleve=eleve,
                defaults={"lien": lien}
            )
            if created_flag:
                liens_created += 1

        if errors:
            messages.warning(request, f"⚠️ Import terminé avec erreurs ({len(errors)}).")
        messages.success(
            request,
            f"✅ Import terminé: Parents {parents_created} créés / {parents_updated} mis à jour — Liens ajoutés {liens_created}"
        )

        return render(request, "admin/parents/import_result.html", {
            "errors": errors,
            "parents_created": parents_created,
            "parents_updated": parents_updated,
            "liens_created": liens_created,
        })

    return render(request, "admin/parents/import.html")

@login_required
@permission_required("core.view_inscription", raise_exception=True)
def inscriptions_excel_export(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = request.GET.get("annee", "") or (str(annee_active.id) if annee_active else "")

    qs = Inscription.objects.select_related(
        "eleve", "annee", "groupe", "groupe__niveau", "groupe__niveau__degre"
    )

    if annee_id:
        qs = qs.filter(annee_id=annee_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inscriptions"

    headers = ["eleve_matricule", "annee_nom", "groupe_nom", "montant_total"]
    ws.append(headers)

    for col in range(1, len(headers)+1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for insc in qs.order_by("eleve__nom", "eleve__prenom"):
        ws.append([
            insc.eleve.matricule,
            insc.annee.nom,
            insc.groupe.nom,
            str(insc.montant_total or ""),
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="inscriptions.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required("core.change_inscription", raise_exception=True)
def inscriptions_excel_import(request):
    errors = []
    created = 0
    updated = 0

    if request.method == "POST":
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "⚠️ Aucun fichier uploadé.")
            return redirect("core:inscription_list")

        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
        except Exception:
            messages.error(request, "⚠️ Fichier Excel invalide.")
            return redirect("core:inscription_list")

        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        expected = ["eleve_matricule", "annee_nom", "groupe_nom", "montant_total"]
        if headers[:4] != expected:
            messages.error(request, f"⚠️ Colonnes attendues: {expected}")
            return redirect("core:inscription_list")

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            eleve_matricule, annee_nom, groupe_nom, montant_total = row[:4]

            eleve_matricule = str(eleve_matricule).strip() if eleve_matricule else ""
            annee_nom = str(annee_nom).strip() if annee_nom else ""
            groupe_nom = str(groupe_nom).strip() if groupe_nom else ""
            montant_total = str(montant_total).strip() if montant_total is not None else ""

            if not eleve_matricule or not annee_nom or not groupe_nom or not montant_total:
                errors.append(f"Ligne {idx}: champs obligatoires manquants.")
                continue

            eleve = Eleve.objects.filter(matricule=eleve_matricule).first()
            if not eleve:
                errors.append(f"Ligne {idx}: élève introuvable ({eleve_matricule}).")
                continue

            annee = AnneeScolaire.objects.filter(nom=annee_nom).first()
            if not annee:
                errors.append(f"Ligne {idx}: année introuvable ({annee_nom}).")
                continue

            groupe = Groupe.objects.filter(annee=annee, nom=groupe_nom).first()
            if not groupe:
                errors.append(f"Ligne {idx}: groupe introuvable ({groupe_nom}) pour année {annee_nom}.")
                continue

            try:
                mt = Decimal(montant_total)
            except Exception:
                errors.append(f"Ligne {idx}: montant_total invalide ({montant_total}).")
                continue

            obj, created_flag = Inscription.objects.update_or_create(
                eleve=eleve,
                annee=annee,
                defaults={
                    "groupe": groupe,
                    "montant_total": mt,
                }
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        if errors:
            messages.warning(request, f"⚠️ Import terminé avec erreurs ({len(errors)}).")
        messages.success(request, f"✅ Import terminé: {created} créées / {updated} mises à jour.")

        return render(request, "admin/inscriptions/import_result.html", {
            "errors": errors,
            "created": created,
            "updated": updated,
        })

    return render(request, "admin/inscriptions/import.html")



@login_required
@permission_required("core.change_paiement", raise_exception=True)
def paiements_excel_import(request):
    errors = []
    created = 0

    if request.method == "POST":
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "⚠️ Aucun fichier uploadé.")
            return redirect("core:paiement_list")

        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
        except Exception:
            messages.error(request, "⚠️ Fichier Excel invalide.")
            return redirect("core:paiement_list")

        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        expected = ["eleve_matricule", "annee_nom", "date_paiement", "montant", "mode", "reference"]
        if headers[:6] != expected:
            messages.error(request, f"⚠️ Colonnes attendues: {expected}")
            return redirect("core:paiement_list")

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            eleve_matricule, annee_nom, date_paiement, montant, mode, reference = row[:6]

            eleve_matricule = str(eleve_matricule).strip() if eleve_matricule else ""
            annee_nom = str(annee_nom).strip() if annee_nom else ""
            mode = str(mode).strip() if mode else ""
            reference = str(reference).strip() if reference else ""

            if not eleve_matricule or not annee_nom or not date_paiement or montant is None:
                errors.append(f"Ligne {idx}: champs obligatoires manquants.")
                continue

            eleve = Eleve.objects.filter(matricule=eleve_matricule).first()
            if not eleve:
                errors.append(f"Ligne {idx}: élève introuvable ({eleve_matricule}).")
                continue

            annee = AnneeScolaire.objects.filter(nom=annee_nom).first()
            if not annee:
                errors.append(f"Ligne {idx}: année introuvable ({annee_nom}).")
                continue

            insc = Inscription.objects.filter(eleve=eleve, annee=annee).first()
            if not insc:
                errors.append(f"Ligne {idx}: inscription introuvable pour {eleve_matricule} / {annee_nom}.")
                continue

            # parse date
            try:
                if isinstance(date_paiement, str):
                    dp = datetime.strptime(date_paiement.strip(), "%Y-%m-%d").date()
                else:
                    dp = date_paiement  # openpyxl peut fournir un datetime/date
                    if hasattr(dp, "date"):
                        dp = dp.date()
            except Exception:
                errors.append(f"Ligne {idx}: date_paiement invalide ({date_paiement}). Utilise YYYY-MM-DD.")
                continue

            # parse montant
            try:
                mt = Decimal(str(montant).strip())
            except Exception:
                errors.append(f"Ligne {idx}: montant invalide ({montant}).")
                continue

            p = Paiement.objects.create(
                inscription=insc,
                montant=mt,
                date_paiement=dp,
            )

            # champs optionnels (si existent)
            if hasattr(p, "mode"):
                p.mode = mode
            if hasattr(p, "reference"):
                p.reference = reference
            p.save()

            created += 1

        if errors:
            messages.warning(request, f"⚠️ Import terminé avec erreurs ({len(errors)}).")
        messages.success(request, f"✅ Import terminé: {created} paiements créés.")

        return render(request, "admin/paiements/import_result.html", {
            "errors": errors,
            "created": created,
        })

    return render(request, "admin/paiements/import.html")

# ============================
# O1 — Notes & Evaluations
# ============================


@login_required
@permission_required("core.view_matiere", raise_exception=True)
def matiere_list(request):
    degre_id = request.GET.get("degre", "")
    niveau_id = request.GET.get("niveau", "")

    matieres = Matiere.objects.prefetch_related("niveaux", "enseignants").all()

    if niveau_id:
        matieres = matieres.filter(niveaux__id=niveau_id)
    elif degre_id:
        matieres = matieres.filter(niveaux__degre_id=degre_id)

    degres = Degre.objects.all().order_by("ordre", "nom")
    niveaux = Niveau.objects.select_related("degre").all().order_by("degre__ordre", "ordre", "nom")

    if degre_id:
        niveaux = niveaux.filter(degre_id=degre_id)

    return render(request, "admin/matieres/list.html", {
        "matieres": matieres.distinct().order_by("nom"),
        "degres": degres,
        "niveaux": niveaux,
        "degre_selected": degre_id,
        "niveau_selected": niveau_id,
    })


@login_required
@permission_required("core.add_matiere", raise_exception=True)
def matiere_create(request):
    if request.method == "POST":
        form = MatiereForm(request.POST)
        if form.is_valid():
            matiere = form.save()

            transaction.on_commit(lambda: sync_enseignant_groupe_from_matiere(matiere))

            messages.success(request, "✅ Matière ajoutée.")
            return redirect("core:matiere_list")
    else:
        form = MatiereForm()

    return render(request, "admin/matieres/form.html", {"form": form, "mode": "create"})


@login_required
@permission_required("core.change_matiere", raise_exception=True)
def matiere_update(request, pk):
    matiere = get_object_or_404(Matiere, pk=pk)

    if request.method == "POST":
        form = MatiereForm(request.POST, instance=matiere)
        if form.is_valid():
            matiere = form.save()

            transaction.on_commit(lambda: sync_enseignant_groupe_from_matiere(matiere))

            messages.success(request, "✅ Matière modifiée.")
            return redirect("core:matiere_list")
    else:
        form = MatiereForm(instance=matiere)

    return render(request, "admin/matieres/form.html", {"form": form, "mode": "update", "matiere": matiere})

@login_required
@permission_required("core.delete_matiere", raise_exception=True)
def matiere_delete(request, pk):
    matiere = get_object_or_404(Matiere, pk=pk)

    if request.method == "POST":
        # ✅ Soft delete : on désactive au lieu de supprimer
        matiere.is_active = False
        matiere.save(update_fields=["is_active"])
        messages.success(request, "✅ Matière désactivée (historique conservé).")
        return redirect("core:matiere_list")

    return render(request, "admin/matieres/delete.html", {
        "matiere": matiere
    })


# core/views_notes.py (ou core/views.py selon ton projet)


# =========================
# AJAX — Élèves par groupe (pour Communication / SMS)
# =========================

@login_required
@permission_required("core.view_eleve", raise_exception=True)
def ajax_eleves_par_groupe(request):
    """
    AJAX — Élèves par groupe (Communication/SMS)
    ✅ Ne retourne QUE les élèves actifs, inscrits sur l'année active, sans doublons.
    Params:
      - groupe_id
    """
    groupe_id = (request.GET.get("groupe_id") or "").strip()
    if not groupe_id.isdigit():
        return JsonResponse({"results": []})

    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    qs = Eleve.objects.filter(is_active=True, inscriptions__groupe_id=int(groupe_id))

    if annee_active:
        qs = qs.filter(inscriptions__annee=annee_active)

    # si ton Inscription a un champ statut, on ESSAIE de filtrer, sinon on ne bloque pas
    if "statut" in [f.name for f in Inscription._meta.fields]:
        wanted = ["EN_COURS", "VALIDE"]
        qs_statut = qs.filter(inscriptions__statut__in=wanted)
        # si ça donne 0 résultats, on garde qs sans filtrer (évite liste vide)
        if qs_statut.exists():
            qs = qs_statut

    qs = (qs
          .distinct()
          .order_by("nom", "prenom")
          .values("id", "matricule", "nom", "prenom"))

    results = [
        {
            "id": e["id"],
            "label": f'{(e["matricule"] or "").strip()} — {e["nom"]} {e["prenom"]}'.strip(" —")
        }
        for e in qs
    ]
    return JsonResponse({"results": results})


@login_required
@require_GET
def ajax_enseignants(request):
    """
    ADMIN: Enseignants en cascade:
    - annee + groupe => enseignants affectés au groupe (EnseignantGroupe)
    Retour: {"results":[{"id":..,"label":"..."}]}
    """
    annee_id = (request.GET.get("annee") or "").strip()
    groupe_id = (request.GET.get("groupe") or request.GET.get("groupe_id") or "").strip()

    # fallback année active
    if not annee_id:
        annee_active = AnneeScolaire.objects.filter(is_active=True).first()
        annee_id = str(annee_active.id) if annee_active else ""

    if not (annee_id and groupe_id and groupe_id.isdigit()):
        return JsonResponse({"results": []})

    # sécurité cohérence année (optionnel mais propre)
    g = Groupe.objects.filter(id=groupe_id).values("id", "annee_id").first()
    if not g or (annee_id and str(g["annee_id"]) != str(annee_id)):
        return JsonResponse({"results": []})

    ens_qs = (
        Enseignant.objects
        .filter(is_active=True)
        .filter(
            id__in=(
                EnseignantGroupe.objects
                .filter(annee_id=annee_id, groupe_id=groupe_id)
                .values_list("enseignant_id", flat=True)
                .distinct()
            )
        )
        .order_by("nom", "prenom")
    )

    data = [{
        "id": e.id,
        "label": f"{e.matricule or ''} — {e.nom} {e.prenom}".strip(" —")
    } for e in ens_qs]

    return JsonResponse({"results": data})

@login_required
@require_GET
def ajax_matieres(request):
    """
    ADMIN: Matières en cascade:
    - si (annee + groupe + enseignant) => matières exactes via EnseignantGroupe.matiere_fk
    - sinon si groupe => matières du niveau du groupe
    - sinon si niveau => matières du niveau
    """
    annee_id = (request.GET.get("annee") or "").strip()
    groupe_id = (request.GET.get("groupe") or request.GET.get("groupe_id") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    enseignant_id = (request.GET.get("enseignant") or "").strip()

    # fallback année active
    if not annee_id:
        annee_active = AnneeScolaire.objects.filter(is_active=True).first()
        annee_id = str(annee_active.id) if annee_active else ""

    qs = Matiere.objects.filter(is_active=True)

    # ✅ CAS 1: enseignant + groupe + annee => matières exactes (AZ)
    if annee_id and groupe_id and enseignant_id and groupe_id.isdigit() and enseignant_id.isdigit():
        matiere_ids = (
            EnseignantGroupe.objects
            .filter(
                annee_id=annee_id,
                groupe_id=groupe_id,
                enseignant_id=enseignant_id,
                matiere_fk__isnull=False,
            )
            .values_list("matiere_fk_id", flat=True)
            .distinct()
        )

        qs = qs.filter(id__in=matiere_ids).order_by("nom")
        return JsonResponse({"results": [{"id": m.id, "label": m.nom} for m in qs]})

    # ✅ CAS 2: groupe => matières du niveau du groupe
    if groupe_id and groupe_id.isdigit():
        g = Groupe.objects.select_related("annee", "niveau").filter(id=groupe_id).first()
        if not g or not g.niveau_id:
            return JsonResponse({"results": []})

        # sécurité cohérence année
        if annee_id and str(g.annee_id) != str(annee_id):
            return JsonResponse({"results": []})

        qs = qs.filter(niveaux=g.niveau).distinct().order_by("nom")
        return JsonResponse({"results": [{"id": m.id, "label": m.nom} for m in qs]})

    # ✅ CAS 3: niveau => matières du niveau
    if niveau_id and niveau_id.isdigit():
        qs = qs.filter(niveaux__id=niveau_id).distinct().order_by("nom")
        return JsonResponse({"results": [{"id": m.id, "label": m.nom} for m in qs]})

    return JsonResponse({"results": []})


@login_required
@permission_required("core.view_evaluation", raise_exception=True)
def evaluation_list(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = request.GET.get("annee", "") or (str(annee_active.id) if annee_active else "")

    # ✅ filtres GET
    date_str = (request.GET.get("date") or "").strip()      # YYYY-MM-DD
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    enseignant_id = (request.GET.get("enseignant") or "").strip()
    matiere_id = (request.GET.get("matiere") or "").strip()
    periode_id = (request.GET.get("periode") or "").strip()

    # ✅ Si groupe choisi => force niveau cohérent
    groupe_obj = None
    if groupe_id:
        groupe_obj = (
            Groupe.objects
            .select_related("niveau", "niveau__degre", "annee")
            .filter(id=groupe_id)
            .first()
        )
        if groupe_obj:
            niveau_id = str(groupe_obj.niveau_id)

    # =========================
    # Query Evaluations
    # =========================
    qs = Evaluation.objects.select_related(
        "matiere", "enseignant",
        "periode", "periode__annee",
        "groupe", "groupe__niveau", "groupe__niveau__degre"
    )

    # année via période
    if annee_id:
        qs = qs.filter(periode__annee_id=annee_id)

    # date exacte
    if date_str:
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            qs = qs.filter(date=d)
        except ValueError:
            qs = qs.none()

    if periode_id:
        qs = qs.filter(periode_id=periode_id)

    if niveau_id:
        qs = qs.filter(groupe__niveau_id=niveau_id)

    if groupe_id:
        qs = qs.filter(groupe_id=groupe_id)

    if enseignant_id:
        qs = qs.filter(enseignant_id=enseignant_id)

    if matiere_id:
        qs = qs.filter(matiere_id=matiere_id)

    # =========================
    # Dropdowns (listes filtres)
    # =========================
    annees = AnneeScolaire.objects.all()

    niveaux = Niveau.objects.select_related("degre").all()
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    groupes = Groupe.objects.select_related("annee", "niveau", "niveau__degre").all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    periodes = Periode.objects.select_related("annee").all()
    if annee_id:
        periodes = periodes.filter(annee_id=annee_id)

    # ✅ Matières: filtrage correct (par groupe -> niveau)
    matieres = Matiere.objects.filter(is_active=True)
    if groupe_obj:
        matieres = matieres.filter(niveaux=groupe_obj.niveau).distinct()
    elif niveau_id:
        matieres = matieres.filter(niveaux__id=niveau_id).distinct()
    matieres = matieres.order_by("nom")

    if matiere_id and not matieres.filter(id=matiere_id).exists():
        matiere_id = ""

    # ✅ Enseignants: via EnseignantGroupe
    enseignants = Enseignant.objects.filter(is_active=True)

    if annee_id and groupe_id:
        enseignants = enseignants.filter(
            affectations_groupes__annee_id=annee_id,
            affectations_groupes__groupe_id=groupe_id
        ).distinct()
    elif annee_id and niveau_id:
        enseignants = enseignants.filter(
            affectations_groupes__annee_id=annee_id,
            affectations_groupes__groupe__niveau_id=niveau_id
        ).distinct()

    enseignants = enseignants.order_by("nom", "prenom")

    if enseignant_id and not enseignants.filter(id=enseignant_id).exists():
        enseignant_id = ""

    return render(request, "admin/notes/evaluations_list.html", {
        "evaluations": qs.order_by("-date", "-id"),

        "annees": annees,
        "annee_selected": annee_id,

        "date_selected": date_str,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "periodes": periodes,
        "periode_selected": periode_id,

        "enseignants": enseignants,
        "enseignant_selected": enseignant_id,

        "matieres": matieres,
        "matiere_selected": matiere_id,
    })

@login_required
@permission_required("core.add_evaluation", raise_exception=True)
def evaluation_create(request):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    if not annee_active:
        messages.error(request, "⚠️ Aucune année scolaire active.")
        return redirect("core:evaluation_list")  # adapte si besoin

    niveau_ui = (
        (request.POST.get("niveau_ui") if request.method == "POST" else request.GET.get("niveau_ui"))
        or ""
    ).strip()

    niveaux = (
        Niveau.objects
        .select_related("degre")
        .all()
        .order_by("degre__ordre", "ordre", "nom")
    )

    if request.method == "POST":
        form = EvaluationForm(request.POST, niveau_ui=niveau_ui)
        if form.is_valid():
            ev = form.save()
            messages.success(request, "✅ Évaluation créée.")
            return redirect("core:notes_saisie", evaluation_id=ev.id)
    else:
        form = EvaluationForm(niveau_ui=niveau_ui)

    return render(request, "admin/notes/evaluations_form.html", {
        "form": form,
        "mode": "create",
        "annee_active": annee_active,
        "niveaux": niveaux,
        "niveau_ui_selected": niveau_ui,
    })


def _clean_int(value):
    """
    Retourne '' si value n'est pas un entier (ex: 'ANGLAIS'),
    sinon retourne la valeur en string.
    """
    value = (value or "").strip()
    return value if value.isdigit() else ""


@login_required
@permission_required("core.view_note", raise_exception=True)
def notes_saisie_home(request):
    """
    Page 1: choisir une évaluation à saisir
    + filtres: q, niveau, groupe, periode, matiere
    + affiche le nom d'enseignant (direct ou fallback via EnseignantGroupe.matiere_fk)
    """
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    if not annee_active:
        messages.error(request, "⚠️ Aucune année scolaire active.")
        return redirect("core:dashboard")

    # -------------------------
    # ✅ GET (safe)
    # -------------------------
    q = (request.GET.get("q") or "").strip()

    niveau_id = _clean_int(request.GET.get("niveau"))
    groupe_id = _clean_int(request.GET.get("groupe"))
    periode_id = _clean_int(request.GET.get("periode"))

    # ⚠️ matiere peut arriver en texte => on corrige
    raw_matiere = (request.GET.get("matiere") or "").strip()
    matiere_id = raw_matiere if raw_matiere.isdigit() else ""

    if raw_matiere and not raw_matiere.isdigit():
        m = Matiere.objects.filter(is_active=True, nom__iexact=raw_matiere).first()
        matiere_id = str(m.id) if m else ""

    # -------------------------
    # ✅ Dropdowns
    # -------------------------
    niveaux = (
        Niveau.objects
        .select_related("degre")
        .all()
        .order_by("degre__ordre", "ordre", "nom")
    )

    groupes = (
        Groupe.objects
        .select_related("niveau", "niveau__degre")
        .filter(annee=annee_active)
    )
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    periodes = Periode.objects.filter(annee=annee_active).order_by("ordre")

    # matières filtrées selon niveau/groupe
    matieres = Matiere.objects.filter(is_active=True)

    if groupe_id:
        g = Groupe.objects.select_related("niveau").filter(id=groupe_id, annee=annee_active).first()
        if g:
            matieres = matieres.filter(niveaux=g.niveau).distinct()
        else:
            matieres = matieres.none()
    elif niveau_id:
        matieres = matieres.filter(niveaux__id=niveau_id).distinct()

    matieres = matieres.order_by("nom")

    if matiere_id and not matieres.filter(id=matiere_id).exists():
        matiere_id = ""

    # -------------------------
    # ✅ Evaluations
    # -------------------------
    evaluations = (
        Evaluation.objects
        .select_related("matiere", "enseignant", "periode", "groupe", "groupe__niveau", "groupe__niveau__degre")
        .filter(groupe__annee=annee_active, periode__annee=annee_active)
    )

    if niveau_id and not groupe_id:
        evaluations = evaluations.filter(groupe__niveau_id=niveau_id)

    if groupe_id:
        evaluations = evaluations.filter(groupe_id=groupe_id)

    if periode_id:
        evaluations = evaluations.filter(periode_id=periode_id)

    if matiere_id:
        evaluations = evaluations.filter(matiere_id=matiere_id)

    if q:
        evaluations = evaluations.filter(
            Q(titre__icontains=q) |
            Q(matiere__nom__icontains=q) |
            Q(enseignant__nom__icontains=q) |
            Q(enseignant__prenom__icontains=q)
        )

    evaluations = evaluations.order_by("-date", "matiere__nom", "titre")

    # -------------------------
    # ✅ Fallback enseignant via EnseignantGroupe
    #   clé robuste: (groupe_id, matiere_id) -> enseignant
    # -------------------------
    ens_map = {}

    eg_qs = (
        EnseignantGroupe.objects
        .select_related("enseignant", "groupe", "matiere_fk")
        .filter(annee=annee_active, matiere_fk__isnull=False)
    )

    for eg in eg_qs:
        # ✅ mapping par IDs (propre)
        ens_map[(eg.groupe_id, eg.matiere_fk_id)] = eg.enseignant

    # ✅ label prêt pour template
    for ev in evaluations:
        if ev.enseignant:
            ev.enseignant_label = f"{ev.enseignant.nom} {ev.enseignant.prenom}"
        else:
            ens = ens_map.get((ev.groupe_id, ev.matiere_id))
            ev.enseignant_label = f"{ens.nom} {ens.prenom}" if ens else "—"

    return render(request, "admin/notes/notes_saisie_home.html", {
        "annee_active": annee_active,

        "q": q,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "periodes": periodes,
        "periode_selected": periode_id,

        "matieres": matieres,
        "matiere_selected": matiere_id,

        "evaluations": evaluations,
    })


@login_required
@permission_required("core.view_note", raise_exception=True)
def notes_saisie(request, evaluation_id):
    """
    Page 2: saisie notes
    ✅ Après Enregistrer => on reste sur page et on revoit les anciennes notes
    ✅ Champ vide => supprime la note
    """
    ev = get_object_or_404(
        Evaluation.objects.select_related("groupe", "periode", "matiere", "enseignant"),
        id=evaluation_id
    )

    inscriptions = (
        Inscription.objects
        .select_related("eleve")
        .filter(annee=ev.periode.annee, groupe=ev.groupe)
        .order_by("eleve__nom", "eleve__prenom")
    )
    eleves = [i.eleve for i in inscriptions]

    if request.method == "POST":
        saved = 0
        deleted = 0

        for e in eleves:
            key = f"note_{e.id}"
            raw = (request.POST.get(key) or "").strip()

            if raw == "":
                Note.objects.filter(evaluation=ev, eleve=e).delete()
                deleted += 1
                continue

            try:
                val = Decimal(raw.replace(",", "."))
            except Exception:
                messages.error(request, f"⚠️ Note invalide pour {e.matricule} : {raw}")
                continue

            if val < 0 or val > Decimal(ev.note_max):
                messages.error(request, f"⚠️ Note hors limite pour {e.matricule} (0–{ev.note_max})")
                continue

            Note.objects.update_or_create(
                evaluation=ev,
                eleve=e,
                defaults={"valeur": val}
            )
            saved += 1

        messages.success(request, f"✅ Enregistré: {saved} note(s) | supprimé: {deleted} (vides)")
        return redirect("core:notes_saisie", evaluation_id=ev.id)

    notes_map = {
        n.eleve_id: n
        for n in Note.objects.filter(evaluation=ev, eleve__in=eleves)
    }

    return render(request, "admin/notes/notes_saisie.html", {
        "ev": ev,
        "eleves": eleves,
        "notes_map": notes_map,
    })


# si tu as déjà group_required, garde le tien
 # adapte le chemin si besoin

# si tu as ton pdf_utils existant


# Views
# =========================
@login_required
@permission_required("core.view_note", raise_exception=True)
def bulletin_view(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)

    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annee_id = (request.GET.get("annee", "") or "").strip() or (str(annee_active.id) if annee_active else "")
    periode_id = (request.GET.get("periode", "") or "").strip()

    annees = AnneeScolaire.objects.all().order_by("-date_debut")

    periodes = Periode.objects.select_related("annee").all()
    if annee_id:
        periodes = periodes.filter(annee_id=annee_id).order_by("ordre")

    periode = None
    if periode_id:
        qs = Periode.objects.filter(id=periode_id)
        if annee_id:
            qs = qs.filter(annee_id=annee_id)
        periode = qs.first()
    else:
        periode = periodes.first()

    data = {"rows": [], "moyenne_generale": None, "best": None, "nb_notes": 0}
    groupe = None
    rank = effectif = None
    moyenne_cls = None

    if periode:
        # groupe via inscription
        insc = (
            Inscription.objects
            .select_related("groupe", "groupe__niveau", "annee")
            .filter(eleve=eleve, annee=periode.annee)
            .first()
        )
        groupe = insc.groupe if insc else None

        data = bulletin_data(eleve, periode)

        if groupe:
            rank, effectif = rang_eleve(periode, groupe, eleve)
            moyenne_cls = moyenne_classe(periode, groupe)

    return render(request, "admin/notes/bulletin.html", {
        "eleve": eleve,
        "annees": annees,
        "annee_selected": annee_id,
        "periodes": periodes,
        "periode_selected": str(periode.id) if periode else "",
        "periode": periode,
        "data": data,
        "groupe": groupe,
        "rank": rank,
        "effectif": effectif,
        "moyenne_classe": moyenne_cls,
    })


@login_required
@permission_required("core.view_note", raise_exception=True)
def bulletin_pdf_view(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)
    periode_id = request.GET.get("periode")

    if not periode_id:
        messages.error(request, "⚠️ Période manquante.")
        return redirect("core:bulletin_view", eleve_id=eleve.id)

    periode = get_object_or_404(Periode, id=periode_id)

    insc = (
        Inscription.objects
        .select_related("groupe")
        .filter(eleve=eleve, annee=periode.annee)
        .first()
    )
    if not insc:
        messages.error(request, "⚠️ Cet élève n'est pas inscrit dans l'année de cette période.")
        return redirect("core:bulletin_view", eleve_id=eleve.id)

    groupe = insc.groupe
    data = bulletin_data(eleve, periode)

    rank, effectif = rang_eleve(periode, groupe, eleve) if groupe else (None, None)
    moyenne_cls = moyenne_classe(periode, groupe) if groupe else None

    # adapte si ton pdf_utils ne prend pas moyenne_classe
    return pdf_utils.bulletin_pdf(
        eleve, periode, data,
        groupe=groupe, rank=rank, effectif=effectif,
        moyenne_classe=moyenne_cls
    )


@login_required
def parent_dashboard(request):
    user = request.user

    # ✅ uniquement PARENT
    if not user.groups.filter(name="PARENT").exists():
        return render(request, "parent/forbidden.html", status=403)

    # ✅ Parent lié au user
    parent = getattr(user, "parent_profile", None)
    if parent is None:
        return render(
            request,
            "parent/forbidden.html",
            {"message": "Votre compte n'est pas encore lié à un parent. Contactez l'administration."},
            status=403,
        )

    # ✅ récupérer les liens + enfants
    liens = (
        ParentEleve.objects
        .select_related("eleve")
        .filter(parent=parent)
        .order_by("eleve__nom", "eleve__prenom")
    )

    enfants = [l.eleve for l in liens]

    return render(request, "parent/dashboard.html", {
        "parent": parent,
        "enfants": enfants,  # ✅ ton template attend ça
        "liens": liens,      # (optionnel, utile plus tard)
    })


# Si ton modèle Absence existe, décommente:
# from .models import Absence


def _pct_change(current, previous):
    """Retourne une variation % arrondie (int)."""
    current = float(current or 0)
    previous = float(previous or 0)
    if previous <= 0:
        return 0
    return int(round(((current - previous) / previous) * 100))


# ============================
# U1 — Users (liste + export)
# ============================

@login_required
@permission_required("auth.view_user", raise_exception=True)
def users_list(request):
    User = get_user_model()

    q = (request.GET.get("q") or "").strip()
    group_filter = (request.GET.get("group") or "").strip()  # PROF / PARENT / ...

    users = User.objects.all().prefetch_related("groups").order_by("username")

    if q:
        users = users.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q)
        )

    if group_filter:
        users = users.filter(groups__name=group_filter)

    # groups pour dropdown filtre
    from django.contrib.auth.models import Group
    groups = Group.objects.all().order_by("name")

    return render(request, "admin/users/list.html", {
        "users": users,
        "q": q,
        "group_filter": group_filter,
        "groups": groups,
    })


@login_required
@permission_required("auth.view_user", raise_exception=True)
def users_export_csv(request):
    User = get_user_model()
    q = (request.GET.get("q") or "").strip()
    group_filter = (request.GET.get("group") or "").strip()

    users = User.objects.all().prefetch_related("groups").order_by("username")

    if q:
        users = users.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q)
        )
    if group_filter:
        users = users.filter(groups__name=group_filter)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="users_export.csv"'

    writer = csv.writer(response)
    writer.writerow(["matricule", "nom", "role", "is_active"])

    for u in users:
        nom = f"{(u.last_name or '').strip()} {(u.first_name or '').strip()}".strip() or "—"
        roles = [g.name for g in u.groups.all()]
        role = roles[0] if roles else "—"
        writer.writerow([u.username, nom, role, "1" if u.is_active else "0"])

    return response


@login_required
@permission_required("auth.change_user", raise_exception=True)
def users_reset_passwords_export_csv(request):
    """
    Reset + export CSV des mots de passe temporaires
    Format: matricule | nom | mdp_temp | role
    """
    User = get_user_model()

    q = (request.GET.get("q") or "").strip()
    group_filter = (request.GET.get("group") or "").strip()

    users = User.objects.all().prefetch_related("groups").order_by("username")

    if q:
        users = users.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q)
        )

    if group_filter:
        users = users.filter(groups__name=group_filter)

    # sécurité : jamais toucher aux superusers
    users = users.filter(is_superuser=False)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="users_passwords_temp.csv"'

    writer = csv.writer(response)
    writer.writerow(["matricule", "nom", "mdp_temp", "role"])

    for u in users:
        temp_pwd = reset_password(u)

        # stocker le mdp temporaire
        TempPassword.objects.update_or_create(
            user=u,
            defaults={"password": temp_pwd}
        )

        nom = f"{u.last_name or ''} {u.first_name or ''}".strip() or "—"
        roles = [g.name for g in u.groups.all()]
        role = roles[0] if roles else "—"

        writer.writerow([u.username, nom, temp_pwd, role])

    return response

# ============================
# U1 — Users (CRUD + mdp + roles)
# ============================


@login_required
@permission_required("auth.view_user", raise_exception=True)
def users_detail(request, user_id):
    User = get_user_model()
    u = get_object_or_404(User.objects.prefetch_related("groups"), id=user_id)

    temp = TempPassword.objects.filter(user=u).first()
    roles = [g.name for g in u.groups.all()]

    return render(request, "admin/users/detail.html", {
        "u": u,
        "roles": roles,
        "temp": temp,
    })


@login_required
@permission_required("auth.add_user", raise_exception=True)
def users_create(request):
    User = get_user_model()

    if request.method == "POST":
        form = UserCreateForm(request.POST)
        if form.is_valid():
            auto = form.cleaned_data["auto_password"]
            pwd = (form.cleaned_data.get("password") or "").strip()

            # créer user
            user = User(
                username=form.cleaned_data["username"].strip(),
                last_name=form.cleaned_data.get("last_name", "").strip(),
                first_name=form.cleaned_data.get("first_name", "").strip(),
                email=form.cleaned_data.get("email", "").strip(),
                is_active=form.cleaned_data["is_active"],
            )

            # mdp
            if auto:
                pwd = generate_temp_password()
            user.set_password(pwd)
            user.save()

            # groupes
            groups = form.cleaned_data.get("groups")
            if groups:
                user.groups.set(groups)

            # stocker mdp temporaire (tu le veux pour export)
            TempPassword.objects.update_or_create(
                user=user,
                defaults={"password": pwd, "updated_at": timezone.now()}
            )

            messages.success(request, f"✅ User créé: {user.username} | MDP: {pwd}")
            return redirect("core:users_detail", user_id=user.id)
    else:
        form = UserCreateForm()

    return render(request, "admin/users/form.html", {
        "form": form,
        "mode": "create",
    })


@login_required
@permission_required("auth.change_user", raise_exception=True)
def users_update(request, user_id):
    User = get_user_model()
    u = get_object_or_404(User.objects.prefetch_related("groups"), id=user_id)

    if u.is_superuser:
        messages.error(request, "⛔ Modification superuser bloquée ici.")
        return redirect("core:users_detail", user_id=u.id)

    if request.method == "POST":
        form = UserUpdateForm(request.POST, instance=u)
        if form.is_valid():
            obj = form.save()
            groups = form.cleaned_data.get("groups")
            obj.groups.set(groups)
            messages.success(request, "✅ User mis à jour.")
            return redirect("core:users_detail", user_id=obj.id)
    else:
        form = UserUpdateForm(instance=u)
        form.fields["groups"].initial = u.groups.all()

    return render(request, "admin/users/form.html", {
        "form": form,
        "mode": "update",
        "u": u,
    })


@login_required
@permission_required("auth.view_user", raise_exception=True)
def users_password(request, user_id):
    User = get_user_model()
    u = get_object_or_404(User, id=user_id)

    # Sécurité: on bloque superuser
    if u.is_superuser:
        messages.error(request, "⛔ Reset mdp superuser bloqué.")
        return redirect("core:users_detail", pk=u.id)  # ✅ adapte si ton url est pk

    if request.method == "POST":
        form = PasswordChangeForm(request.POST)
        if form.is_valid():
            auto = form.cleaned_data["auto_password"]
            pwd_input = (form.cleaned_data.get("password") or "").strip()

            if auto:
                # ✅ reset_password() fait déjà TempPassword.update_or_create()
                pwd = reset_password(u)
            else:
                pwd = pwd_input
                u.set_password(pwd)
                u.save(update_fields=["password"])

                TempPassword.objects.update_or_create(
                    user=u,
                    defaults={"password": pwd, "updated_at": timezone.now()}
                )

            messages.success(request, f"✅ Mot de passe mis à jour: {u.username} | MDP: {pwd}")
            return redirect("core:users_detail", user_id=u.id)  # ✅ adapte si ton url est user_id
    else:
        form = PasswordChangeForm(initial={"auto_password": True})

    return render(request, "admin/users/password.html", {
        "u": u,
        "form": form,
    })

def _has_linked_profile(user):
    """
    Empêche suppression si lié à un profil Parent/Enseignant.
    (sécurise FK/OneToOne)
    """
    # Parent: tu utilises user.parent_profile
    try:
        _ = user.parent_profile
        return True
    except Exception:
        pass

    # Enseignant: on ne connait pas ton related_name exact, donc on tente plusieurs
    for attr in ["enseignant", "enseignant_profile", "enseignant_set"]:
        try:
            obj = getattr(user, attr)
            if obj:
                return True
        except Exception:
            continue

    return False


@login_required
@permission_required("auth.delete_user", raise_exception=True)
def users_delete(request, user_id):
    User = get_user_model()
    u = get_object_or_404(User, id=user_id)

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    # 1) Sécurité superuser
    if u.is_superuser:
        msg = "⛔ Suppression superuser bloquée."
        if is_ajax:
            return JsonResponse({"ok": False, "redirect": None, "message": msg}, status=403)
        messages.error(request, msg)
        return redirect("core:users_detail", user_id=u.id)

    # 2) Profils liés
    parent = Parent.objects.select_related("user").filter(user=u).first()
    ens = Enseignant.objects.select_related("user").filter(user=u).first()

    if request.method == "POST":
        try:
            with transaction.atomic():

                # === CAS A: lié à Enseignant => on désactive (pas de suppression)
                if ens:
                    if u.is_active:
                        u.is_active = False
                        u.save(update_fields=["is_active"])
                    msg = "✅ User désactivé (lié à un Enseignant)."
                    if is_ajax:
                        return JsonResponse({"ok": True, "redirect": "detail", "user_id": u.id, "message": msg})
                    messages.success(request, msg)
                    return redirect("core:users_detail", user_id=u.id)

                # === CAS B: lié à Parent ===
                if parent:
                    # Parent a encore des élèves ? => désactive uniquement
                    if parent.liens.exists():
                        if u.is_active:
                            u.is_active = False
                            u.save(update_fields=["is_active"])
                        msg = "✅ User désactivé (Parent encore lié à des élèves)."
                        if is_ajax:
                            return JsonResponse({"ok": True, "redirect": "detail", "user_id": u.id, "message": msg})
                        messages.success(request, msg)
                        return redirect("core:users_detail", user_id=u.id)

                    # Parent orphelin => supprimer Parent + User
                    parent.delete()
                    u.delete()
                    msg = "🗑️ Parent orphelin + User supprimés."
                    if is_ajax:
                        return JsonResponse({"ok": True, "redirect": "list", "message": msg})
                    messages.success(request, msg)
                    return redirect("core:users_list")

                # === CAS C: aucun profil lié => suppression
                u.delete()
                msg = "🗑️ User supprimé."
                if is_ajax:
                    return JsonResponse({"ok": True, "redirect": "list", "message": msg})
                messages.success(request, msg)
                return redirect("core:users_list")

        except ProtectedError:
            # FK PROTECT => désactivation fallback
            if u.is_active:
                u.is_active = False
                u.save(update_fields=["is_active"])
            msg = "⛔ Suppression impossible (références existantes). User désactivé."
            if is_ajax:
                return JsonResponse({"ok": True, "redirect": "detail", "user_id": u.id, "message": msg})
            messages.error(request, msg)
            return redirect("core:users_detail", user_id=u.id)

    # ✅ GET : si AJAX => renvoyer le fragment modal, sinon page normale
    # (ton delete.html doit être le HTML modal AZ, sans <style> ni <script>)
    return render(request, "admin/users/delete.html", {"u": u, "parent": parent, "ens": ens})

@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def api_enseignants(request):
    annee_id = request.GET.get("annee_id")
    groupe_id = request.GET.get("groupe_id")

    if not (annee_id and groupe_id):
        return JsonResponse({"results": []})

    qs = (
        Enseignant.objects.filter(
            affectations_groupes__annee_id=annee_id,
            affectations_groupes__groupe_id=groupe_id,
            is_active=True,
        )
        .distinct()
        .order_by("nom", "prenom")
    )

    data = [
        {"id": e.id, "label": f"{e.nom} {e.prenom} ({e.matricule})"}
        for e in qs
    ]
    return JsonResponse({"results": data})



@login_required
def api_matieres_par_groupe(request):
    groupe_id = (request.GET.get("groupe_id") or "").strip()
    if not groupe_id.isdigit():
        return JsonResponse({"results": []})

    gid = int(groupe_id)

    # ✅ STAFF/SUPERUSER: accès total
    if request.user.is_staff or request.user.is_superuser:
        g = Groupe.objects.filter(id=gid).select_related("niveau").first()
    else:
        allowed = _allowed_groupes_for_user(request.user)
        g = allowed.filter(id=gid).select_related("niveau").first()

    if not g or not getattr(g, "niveau_id", None):
        return JsonResponse({"results": []})

    qs = (
        Matiere.objects
        .filter(is_active=True, niveaux=g.niveau)
        .distinct()
        .order_by("nom")
    )

    # ✅ option PROF (uniquement si l’utilisateur est un prof, et pas admin)
    ens = getattr(request.user, "enseignant_profile", None)
    if ens and not (request.user.is_staff or request.user.is_superuser):
        qs_prof = qs.filter(enseignants=ens)
        # LOOSE (recommandé)
        qs = qs_prof if qs_prof.exists() else qs

    data = [{"id": m.id, "label": m.nom} for m in qs]
    return JsonResponse({"results": data})



@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def ajax_niveaux(request):
    """
    Retourne les niveaux disponibles pour une année (annee_id).
    GET: /ajax/niveaux/?annee=<id>
    """
    annee_id = request.GET.get("annee", "").strip()

    qs = Niveau.objects.select_related("degre").all()

    # Filtrer uniquement les niveaux qui ont des groupes dans cette année
    if annee_id:
        qs = qs.filter(groupes__annee_id=annee_id).distinct()

    data = [
        {
            "id": n.id,
            "label": f"{n.degre.nom} — {n.nom}",
        }
        for n in qs.order_by("degre__ordre", "ordre", "nom")
    ]
    return JsonResponse({"results": data})


@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def ajax_groupes(request):
    """
    Retourne les groupes selon année + niveau.
    GET: /ajax/groupes/?annee=<id>&niveau=<id>
    Si annee vide => année active.
    """
    annee_id = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()

    # ✅ fallback annee => année active
    if not annee_id:
        annee_active = AnneeScolaire.objects.filter(is_active=True).first()
        if annee_active:
            annee_id = str(annee_active.id)

    qs = Groupe.objects.select_related("niveau", "niveau__degre", "annee").all()

    if annee_id:
        qs = qs.filter(annee_id=annee_id)

    if niveau_id:
        qs = qs.filter(niveau_id=niveau_id)

    qs = qs.order_by("niveau__degre__ordre", "niveau__ordre", "nom")

    data = [
        {
            "id": g.id,
            "label": f"{g.niveau.nom} / {g.nom}"
        }
        for g in qs
    ]
    return JsonResponse({"results": data})

@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def ajax_periodes(request):
    """
    GET: /ajax/periodes/?annee=<id>
    """
    annee_id = request.GET.get("annee", "").strip()

    qs = Periode.objects.select_related("annee").all()
    if annee_id:
        qs = qs.filter(annee_id=annee_id)

    data = [{"id": p.id, "label": p.nom} for p in qs.order_by("ordre")]
    return JsonResponse({"results": data})


# =========================
# I — Communication (Avis)
# =========================


@login_required
@permission_required("core.view_avis", raise_exception=True)
def avis_list(request):
    q = (request.GET.get("q") or "").strip()
    cible = (request.GET.get("cible") or "").strip()      # cible_type
    period = (request.GET.get("period") or "").strip()    # 7d, 30d, ytd, all

    items = Avis.objects.all()

    # 🔎 Search
    if q:
        items = items.filter(Q(titre__icontains=q) | Q(contenu__icontains=q))

    # 🎯 Cible type
    if cible:
        items = items.filter(cible_type=cible)

    # ⏱️ Période (date_publication)
    now = timezone.now()
    if period == "7d":
        items = items.filter(date_publication__gte=now - timedelta(days=7))
    elif period == "30d":
        items = items.filter(date_publication__gte=now - timedelta(days=30))
    elif period == "ytd":
        start_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        items = items.filter(date_publication__gte=start_year)
    # "all" ou vide => rien

    items = items.order_by("-date_publication", "-id")

    # ✅ pour remplir le select des cibles dynamiquement
    # (Avis.CIBLE_CHOICES si tu l'as dans le model)
    cible_choices = getattr(Avis, "CIBLE_CHOICES", None)

    return render(request, "admin/avis/list.html", {
        "items": items,
        "q": q,
        "cible": cible,
        "period": period,
        "cible_choices": cible_choices,
    })

@login_required
@permission_required("core.add_avis", raise_exception=True)
def avis_create(request):
    form = AvisForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Avis enregistré.")
            return redirect("core:avis_list")
        else:
            # Debug: tu verras l’erreur exacte dans console serveur
            print("AVIS FORM ERRORS =>", form.errors)

    return render(request, "admin/avis/form.html", {"form": form, "mode": "create"})

@login_required
def avis_detail(request, pk):
    obj = get_object_or_404(Avis, pk=pk)
    return render(request, "admin/avis/detail.html", {"obj": obj})


@login_required
def avis_delete(request, pk):
    obj = get_object_or_404(Avis, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "🗑️ Avis supprimé.")
        return redirect("core:avis_list")
    return render(request, "admin/avis/delete.html", {"obj": obj})


# =========================
# I — Communication (SMS)
# =========================

def _parents_from_cible(cible_type: str, degre_id=None, niveau_id=None, groupe_id=None, eleve_id=None):
    """
    Retourne une QuerySet de Parent ciblés via ParentEleve.
    """
    liens = ParentEleve.objects.select_related("parent", "eleve")

    if cible_type == "TOUS":
        pass

    elif cible_type == "DEGRE":
        liens = liens.filter(eleve__inscriptions__groupe__niveau__degre_id=degre_id)

    elif cible_type == "NIVEAU":
        liens = liens.filter(eleve__inscriptions__groupe__niveau_id=niveau_id)

    elif cible_type == "GROUPE":
        liens = liens.filter(eleve__inscriptions__groupe_id=groupe_id)

    elif cible_type == "ELEVE":
        liens = liens.filter(eleve_id=eleve_id)

    # parents uniques
    parent_ids = liens.values_list("parent_id", flat=True).distinct()
    return Parent.objects.filter(id__in=parent_ids, is_active=True)


@login_required
def sms_send(request):
    """
    Envoi réel SMS + historique.
    """
    if request.method == "POST":
        form = SmsSendForm(request.POST)
        if form.is_valid():
            msg = form.cleaned_data["message"]
            cible_type = form.cleaned_data["cible_type"]

            degre_id = form.cleaned_data.get("degre")
            niveau_id = form.cleaned_data.get("niveau")
            groupe_id = form.cleaned_data.get("groupe")
            eleve_id = form.cleaned_data.get("eleve")

            parents = _parents_from_cible(cible_type, degre_id, niveau_id, groupe_id, eleve_id)

            sent = 0
            failed = 0

            for p in parents:
                tel = normalize_phone(p.telephone)
                if not tel:
                    # on log quand même, en FAILED
                    SmsHistorique.objects.create(
                        parent=p,
                        telephone="",
                        message=msg,
                        status="FAILED",
                        provider="twilio",
                        error_message="Téléphone parent vide.",
                    )
                    failed += 1
                    continue

                hist = SmsHistorique.objects.create(
                    parent=p,
                    telephone=tel,
                    message=msg,
                    status="PENDING",
                    provider="twilio",
                )

                ok, provider_id, err = send_sms_via_bulksms_ma(tel, msg)

                if ok:
                    hist.status = "SENT"
                    hist.provider_message_id = provider_id
                    hist.sent_at = timezone.now()
                    hist.error_message = ""
                    sent += 1
                else:
                    hist.status = "FAILED"
                    hist.error_message = err
                    failed += 1

                hist.save(update_fields=["status", "provider_message_id", "error_message", "sent_at", "updated_at", "updated_by"])

            messages.success(request, f"📩 SMS terminé — Envoyés: {sent} | Échecs: {failed}")
            return redirect("core:sms_history")
    else:
        form = SmsSendForm()

    # Pour aider l’UI : listes (select)
    degres = Degre.objects.all()
    niveaux = Niveau.objects.select_related("degre").all()
    groupes = Groupe.objects.select_related("niveau", "annee").all()
    eleves = Eleve.objects.all()

    return render(request, "admin/communication/sms_send.html", {
        "form": form,
        "degres": degres,
        "niveaux": niveaux,
        "groupes": groupes,
        "eleves": eleves,
    })


@login_required
def sms_history(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()     # SENT/FAILED/PENDING/...
    period = (request.GET.get("period") or "").strip()     # 7d / 30d / ytd / all

    qs = SmsHistorique.objects.select_related("parent").all()

    if q:
        qs = qs.filter(
            Q(parent__nom__icontains=q) |
            Q(parent__prenom__icontains=q) |
            Q(telephone__icontains=q) |
            Q(message__icontains=q) |
            Q(error_message__icontains=q)
        )

    if status:
        qs = qs.filter(status=status)

    now = timezone.now()
    if period == "7d":
        qs = qs.filter(created_at__gte=now - timedelta(days=7))
    elif period == "30d":
        qs = qs.filter(created_at__gte=now - timedelta(days=30))
    elif period == "ytd":
        start_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        qs = qs.filter(created_at__gte=start_year)

    qs = qs.order_by("-created_at", "-id")

    # ✅ liste statuts pour select
    status_choices = ["PENDING", "SENT", "DELIVERED", "FAILED", "ERROR"]

    return render(request, "admin/communication/sms_history.html", {
        "items": qs,
        "q": q,
        "status_selected": status,
        "period_selected": period,
        "status_choices": status_choices,
    })


User = get_user_model()

def users_detail_modal(request, pk):
    u = get_object_or_404(User, pk=pk)

    # roles (selon ton système : groups, role field, etc.)
    roles = list(u.groups.values_list("name", flat=True)) if hasattr(u, "groups") else []

    # temp password (si tu as un modèle pour ça)
    temp = getattr(u, "temp_password", None)  # adapte selon ton code
    # ou mets temp=None si tu n’as pas de modèle

    return render(request, "admin/users/_detail_modal.html", {
        "u": u,
        "roles": roles,
        "temp": temp,
    })


@login_required
@permission_required("core.view_eleve", raise_exception=True)
def eleve_reinscrire(request, pk):
    eleve = get_object_or_404(Eleve, pk=pk)

    # ✅ année active (par défaut)
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    annees = AnneeScolaire.objects.all().order_by("-date_debut")

    # ✅ filtres (GET) pour remplir les dropdowns
    annee_id = request.GET.get("annee", "")
    niveau_id = request.GET.get("niveau", "")
    groupe_id = request.GET.get("groupe", "")
    periode_id = request.GET.get("periode", "")

    # ✅ par défaut : année active si rien choisi
    if not annee_id and annee_active:
        annee_id = str(annee_active.id)

    # ✅ dropdowns dépendants
    niveaux = Niveau.objects.select_related("degre").all()
    if annee_id:
        niveaux = niveaux.filter(groupes__annee_id=annee_id).distinct()

    groupes = Groupe.objects.select_related("niveau", "annee").all()
    if annee_id:
        groupes = groupes.filter(annee_id=annee_id)
    if niveau_id:
        groupes = groupes.filter(niveau_id=niveau_id)

    periodes = Periode.objects.all()
    if annee_id:
        periodes = periodes.filter(annee_id=annee_id)

    # ✅ POST = créer la nouvelle inscription
    if request.method == "POST":
        annee_id_post = request.POST.get("annee")
        groupe_id_post = request.POST.get("groupe")
        periode_id_post = request.POST.get("periode") or None

        if not annee_id_post or not groupe_id_post:
            messages.error(request, "Année et groupe sont obligatoires.")
            return redirect("core:eleve_reinscrire", pk=eleve.id)

        # ✅ sécurité : le groupe doit exister
        groupe = get_object_or_404(Groupe, pk=groupe_id_post)

        try:
            Inscription.objects.create(
                eleve=eleve,
                annee_id=annee_id_post,
                groupe=groupe,
                periode_id=periode_id_post,
                statut="VALIDEE",
            )
            messages.success(request, "Réinscription effectuée ✅")
            return redirect("core:eleve_list")
        except IntegrityError:
            # ✅ ta contrainte unique (eleve, annee) déclenche ici
            messages.error(request, "Cet élève est déjà inscrit sur cette année.")
            return redirect("core:eleve_reinscrire", pk=eleve.id)

    return render(request, "admin/eleves/reinscrire.html", {
        "eleve": eleve,
        "annees": annees,
        "annee_selected": annee_id,
        "niveaux": niveaux,
        "niveau_selected": niveau_id,
        "groupes": groupes,
        "groupe_selected": groupe_id,
        "periodes": periodes,
        "periode_selected": periode_id,
    })


@login_required
def api_periodes_par_groupe(request):
    gid = (request.GET.get("groupe_id") or "").strip()
    if not gid.isdigit():
        return JsonResponse([], safe=False)

    groupes = _allowed_groupes_for_user(request.user)
    g = groupes.filter(id=int(gid)).select_related("annee").first()
    if not g:
        return JsonResponse([], safe=False)

    periodes = Periode.objects.filter(annee=g.annee).order_by("ordre")
    return JsonResponse([{"id": p.id, "label": p.nom} for p in periodes], safe=False)


@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def ajax_transport_echeances(request):
    """
    GET /core/ajax/transport-echeances/?inscription=ID
    -> {enabled:bool, items:[...], tarif_mensuel:"0.00"}
    """
    insc_id = (request.GET.get("inscription") or "").strip()
    if not insc_id.isdigit():
        return JsonResponse({"enabled": False, "items": [], "tarif_mensuel": "0.00"}, status=200)

    insc = get_object_or_404(
        Inscription.objects.select_related("eleve", "annee", "groupe"),
        pk=int(insc_id)
    )

    tr = getattr(insc.eleve, "transport", None)
    if not tr or not tr.enabled:
        return JsonResponse({"enabled": False, "items": [], "tarif_mensuel": "0.00"}, status=200)

    qs = (EcheanceTransportMensuelle.objects
          .filter(eleve_id=insc.eleve_id, annee_id=insc.annee_id)
          .order_by("mois_index"))

    items = []
    for e in qs:
        du = e.montant_du or Decimal("0.00")
        is_paye = (e.statut == "PAYE")
        items.append({
            "id": e.id,
            "mois_index": int(e.mois_index),
            "mois_nom": e.mois_nom,
            "du": str(du),
            "is_paye": bool(is_paye),
            "statut": e.statut,
            "date_echeance": str(e.date_echeance),
        })

    return JsonResponse({
        "enabled": True,
        "tarif_mensuel": str(tr.tarif_mensuel or Decimal("0.00")),
        "items": items
    }, status=200)


def _d(x) -> Decimal:
    try:
        return Decimal(str(x or "0").replace(",", "."))
    except Exception:
        return Decimal("0.00")


@login_required
@permission_required("core.change_anneescolaire", raise_exception=True)
def transport_toggle(request, eleve_id: int):
    eleve = get_object_or_404(Eleve, pk=eleve_id)
    obj, _ = EleveTransport.objects.get_or_create(eleve=eleve)

    obj.enabled = not obj.enabled

    # règle: enabled => tarif > 0
    if obj.enabled and (obj.tarif_mensuel or Decimal("0.00")) <= Decimal("0.00"):
        obj.enabled = False
        obj.save(update_fields=["enabled"])
        messages.error(request, "Transport non activé : tarif mensuel = 0.00. Défini d’abord le tarif.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    obj.save(update_fields=["enabled"])

    # sync année active (inscription active)
    insc = (Inscription.objects
            .select_related("annee")
            .filter(eleve_id=eleve.id, annee__is_active=True)
            .first())
    if insc:
        sync_transport_echeances_for_inscription(insc.id)

    messages.success(request, f"Transport {'activé' if obj.enabled else 'désactivé'} pour {eleve.matricule}.")
    return redirect(request.META.get("HTTP_REFERER", "/"))


@login_required
@permission_required("core.change_anneescolaire", raise_exception=True)
@require_POST
def transport_set_tarif(request, eleve_id: int):
    eleve = get_object_or_404(Eleve, pk=eleve_id)
    obj, _ = EleveTransport.objects.get_or_create(eleve=eleve)

    tarif = _d(request.POST.get("tarif_mensuel"))

    if tarif <= Decimal("0.00"):
        obj.tarif_mensuel = Decimal("0.00")
        obj.enabled = False
        obj.save(update_fields=["tarif_mensuel", "enabled"])

        insc = Inscription.objects.filter(eleve_id=eleve.id, annee__is_active=True).first()
        if insc:
            sync_transport_echeances_for_inscription(insc.id)

        messages.success(request, "Tarif transport mis à 0.00 → transport désactivé.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    obj.tarif_mensuel = tarif
    # si tu veux auto-enable quand tarif>0 :
    if not obj.enabled:
        obj.enabled = True
    obj.save(update_fields=["tarif_mensuel", "enabled"])

    insc = Inscription.objects.filter(eleve_id=eleve.id, annee__is_active=True).first()
    if insc:
        sync_transport_echeances_for_inscription(insc.id)

    messages.success(request, f"Transport activé + tarif mis à jour : {tarif:.2f} MAD.")
    return redirect(request.META.get("HTTP_REFERER", "/"))


@login_required
@permission_required("core.view_anneescolaire", raise_exception=True)
def ajax_transport_status(request):
    eleve_id = (request.GET.get("eleve_id") or "").strip()
    if not eleve_id.isdigit():
        return JsonResponse({"enabled": False, "tarif_mensuel": "0.00"})

    eleve = get_object_or_404(Eleve, pk=int(eleve_id))
    obj = getattr(eleve, "transport", None)

    if not obj:
        return JsonResponse({"enabled": False, "tarif_mensuel": "0.00"})

    return JsonResponse({
        "enabled": bool(obj.enabled),
        "tarif_mensuel": f"{(obj.tarif_mensuel or Decimal('0.00')):.2f}",
    })


# core/views_finance.py


def _D(v, default=Decimal("0.00")):
    try:
        s = str(v or "").strip()
        if s == "":
            return default
        s = s.replace(" ", "").replace(",", ".")
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return default


@login_required
@permission_required("core.add_remboursementfinance", raise_exception=True)
def transaction_remboursement_create(request, group_key: str):
    """
    ✅ Remboursement PAR PAIEMENT (batch):
    - group_key="TX-<id>" => transaction unique
    - group_key="<id>" (digit) => transaction unique (compat ancien URL)
    - sinon group_key=batch_token => batch (distribution auto)
    """

    base_qs = (
        TransactionFinance.objects
        .select_related(
            "parent",  # ✅ AJOUT
            "inscription",
            "inscription__eleve",
            "inscription__annee",
            "inscription__groupe",
            "inscription__groupe__niveau",
            "inscription__groupe__niveau__degre",
        )
        .prefetch_related(
            "lignes",
            "lignes__echeance",
            "lignes__echeance_transport",
            "remboursements",
        )
    )


    batch_token = ""

    # ✅ CAS 1: "TX-43"
    if group_key.startswith("TX-"):
        try:
            tx_id = int(group_key.replace("TX-", "").strip())
        except ValueError:
            raise Http404
        txs = base_qs.filter(id=tx_id).order_by("id")

    # ✅ CAS 2: "43" (compat)
    elif group_key.isdigit():
        tx_id = int(group_key)
        txs = base_qs.filter(id=tx_id).order_by("id")

    # ✅ CAS 3: batch_token
    else:
        batch_token = group_key
        txs = base_qs.filter(batch_token=batch_token).order_by("id")

    if not txs.exists():
        messages.error(request, "Paiement introuvable.")
        return redirect("core:paiement_list")

    # -----------------------------
    # 2) Calculs paiement (batch)
    # -----------------------------
    total_brut = Decimal("0.00")
    total_deja = Decimal("0.00")
    total_restant = Decimal("0.00")

    zero_count = 0
    zero_annulables = 0

    first = txs.first()
    annee_label = getattr(getattr(first.inscription, "annee", None), "nom", "—")

    # ✅ parent label (si tu passes parent_label depuis list via annotation)
    parent_label = "—"
    if getattr(first, "parent", None):
        parent_label = f"{first.parent.nom} {first.parent.prenom}"

    items = []

    for tx in txs:
        total = tx.montant_total or Decimal("0.00")

        deja = tx.remboursements.aggregate(s=Sum("montant"))["s"] or Decimal("0.00")
        restant = max(total - deja, Decimal("0.00"))

        already_annulated = tx.remboursements.filter(is_annulation=True).exists()
        is_zero_tx = (total == Decimal("0.00"))

        can_refund_tx = (not already_annulated) if is_zero_tx else (restant > Decimal("0.00"))

        total_brut += total
        total_deja += deja
        total_restant += restant

        if is_zero_tx:
            zero_count += 1
            if not already_annulated:
                zero_annulables += 1

        # statut UI
        if is_zero_tx and already_annulated:
            status_code = "ANNULE"
            status_label = "Annulé (0 MAD)"
        elif (not is_zero_tx) and total > Decimal("0.00") and restant == Decimal("0.00") and deja >= total:
            status_code = "REMBOURSE"
            status_label = "Remboursé"
        elif deja > Decimal("0.00"):
            status_code = "PARTIEL"
            status_label = f"Partiel ({deja} MAD)"
        else:
            status_code = "PAYE"
            status_label = "Payé"

        items.append({
            "tx": tx,
            "total": total,
            "deja": deja,
            "restant": restant,
            "is_zero_tx": is_zero_tx,
            "already_annulated": already_annulated,
            "can_refund": can_refund_tx,
            "status_code": status_code,
            "status_label": status_label,
        })

    # ✅ batch remboursable si:
    # - il reste du montant sur au moins une tx >0
    # - ou il existe une tx=0 pas encore annulée
    can_refund_batch = (total_restant > Decimal("0.00")) or (zero_annulables > 0)

    if not can_refund_batch:
        messages.error(request, "Ce paiement est déjà totalement remboursé / annulé.")
        return redirect("core:paiement_list")

    # -----------------------------
    # 3) POST: créer remboursements
    # -----------------------------
    if request.method == "POST":
        mode = (request.POST.get("mode") or "ESPECES").strip()
        raison = (request.POST.get("raison") or "").strip()

        remaining = total_restant if total_restant > Decimal("0.00") else Decimal("0.00")

        try:
            with transaction.atomic():

                # 3.1 Annuler toutes les tx=0 non annulées (si présentes)
                for it in items:
                    tx = it["tx"]
                    if it["is_zero_tx"] and (not it["already_annulated"]):
                        r0 = RemboursementFinance(
                            transaction=tx,
                            montant=Decimal("0.00"),
                            is_annulation=True,
                            mode=mode,
                            raison=raison or "Annulation transaction 0",
                            created_by=request.user,
                        )
                        r0.save()

                # 3.2 Remboursement sur tx>0 (distribution)
                if remaining > Decimal("0.00"):
                    refundable = [it for it in items if (not it["is_zero_tx"]) and it["restant"] > Decimal("0.00")]
                    # choix: rembourser les plus récentes d'abord
                    refundable.sort(key=lambda x: x["tx"].id, reverse=True)

                    for it in refundable:
                        if remaining <= Decimal("0.00"):
                            break

                        tx = it["tx"]
                        take = min(it["restant"], remaining)

                        r = RemboursementFinance(
                            transaction=tx,
                            montant=take,
                            mode=mode,
                            raison=raison,
                            created_by=request.user,
                        )
                        r.save()
                        remaining -= take

                messages.success(request, "✅ Remboursement (paiement) enregistré.")
                return redirect("core:paiement_list")

        except Exception as e:
            messages.error(request, f"⚠️ {e}")

    return render(request, "admin/paiements/remboursement_form.html", {
        "group_key": group_key,
        "batch_token": batch_token,
        "annee_label": annee_label,
        "parent_label": parent_label,

        "items": items,

        "total_brut": total_brut,
        "total_deja": total_deja,
        "total_restant": total_restant,

        "zero_count": zero_count,
        "zero_annulables": zero_annulables,

        "can_refund": can_refund_batch,
        "modes": RemboursementFinance.MODE_CHOICES,
    })

@login_required
@permission_required("core.view_seance", raise_exception=True)
def edt_prof_week(request, pk=None):
    annee_active = AnneeScolaire.objects.filter(is_active=True).first()

    # --- GET (toujours en str) ---
    annee_id = (request.GET.get("annee") or "").strip()
    niveau_id = (request.GET.get("niveau") or "").strip()
    groupe_id = (request.GET.get("groupe") or "").strip()
    enseignant_id = (request.GET.get("enseignant") or "").strip()

    # ✅ si accès depuis fiche enseignant (/enseignants/<pk>/edt/)
    if pk and not enseignant_id:
        enseignant_id = str(pk)

    # ✅ Année par défaut = active
    if not annee_id.isdigit():
        annee_id = str(annee_active.id) if annee_active else ""

    # Listes filtres
    annees = AnneeScolaire.objects.all()

    niveaux = Niveau.objects.select_related("degre").all()
    if annee_id.isdigit():
        niveaux = niveaux.filter(groupes__annee_id=int(annee_id)).distinct()

    groupes = Groupe.objects.select_related("annee", "niveau", "niveau__degre").all()
    if annee_id.isdigit():
        groupes = groupes.filter(annee_id=int(annee_id))
    if niveau_id.isdigit():
        groupes = groupes.filter(niveau_id=int(niveau_id))

    # ✅ Enseignants possibles (selon filtres) = ceux qui ont des séances
    ens_qs = Seance.objects.all()
    if annee_id.isdigit():
        ens_qs = ens_qs.filter(annee_id=int(annee_id))
    if niveau_id.isdigit():
        ens_qs = ens_qs.filter(groupe__niveau_id=int(niveau_id))
    if groupe_id.isdigit():
        ens_qs = ens_qs.filter(groupe_id=int(groupe_id))

    enseignants = Enseignant.objects.filter(
        id__in=ens_qs.values_list("enseignant_id", flat=True).distinct()
    ).order_by("nom", "prenom", "matricule")

    # ✅ si enseignant_id vide, on prend le 1er (UX pratique)
    if (not enseignant_id.isdigit()) and enseignants.exists():
        enseignant_id = str(enseignants.first().id)

    # =========================
    # Séances (EDT prof)
    # =========================
    seances = Seance.objects.select_related("enseignant", "groupe", "groupe__niveau").all()

    if annee_id.isdigit():
        seances = seances.filter(annee_id=int(annee_id))

    if niveau_id.isdigit():
        seances = seances.filter(groupe__niveau_id=int(niveau_id))

    if groupe_id.isdigit():
        seances = seances.filter(groupe_id=int(groupe_id))

    if enseignant_id.isdigit():
        seances = seances.filter(enseignant_id=int(enseignant_id))
    else:
        # aucun enseignant => aucune séance
        seances = seances.none()

    seances = list(seances)

    # =========================
    # Grille semaine (même algo que prof_edt)
    # =========================
    days = [("LUN", "Lu"), ("MAR", "Ma"), ("MER", "Me"), ("JEU", "Je"), ("VEN", "Ve"), ("SAM", "Sa")]
    slots = [
        (time(8, 30),  time(9, 30)),
        (time(9, 30),  time(10, 30)),
        (time(10, 30), time(11, 30)),
        (time(11, 30), time(12, 30)),
        (time(14, 30), time(15, 30)),
        (time(15, 30), time(16, 30)),
        (time(16, 30), time(17, 30)),
        (time(17, 30), time(18, 30)),
    ]

    def fmt(t): return t.strftime("%H:%M") if t else ""

    def find_slot_index_start(t_):
        for i, (a, b) in enumerate(slots):
            if a <= t_ < b:
                return i
        return None

    def find_slot_index_end(t_):
        for i, (a, b) in enumerate(slots):
            if a < t_ <= b:
                return i
        return None

    slot_labels = [{"num": i, "start": fmt(a), "end": fmt(b)} for i, (a, b) in enumerate(slots, start=1)]
    n = len(slots)

    seances_by_day = {}
    for se in seances:
        seances_by_day.setdefault(se.jour, []).append(se)

    rows = []
    for code, short in days:
        occ = [None] * n

        for se in seances_by_day.get(code, []):
            i0 = find_slot_index_start(se.heure_debut)
            i1 = find_slot_index_end(se.heure_fin)
            if i0 is None or i1 is None:
                continue

            colspan = max(1, (i1 - i0 + 1))
            if isinstance(occ[i0], dict):
                occ[i0]["items"].append(se)
                occ[i0]["colspan"] = max(occ[i0]["colspan"], colspan)
            else:
                occ[i0] = {"items": [se], "colspan": colspan}

            for k in range(i0 + 1, min(i0 + colspan, n)):
                occ[k] = "SKIP"

        cells_render = []
        i = 0
        while i < n:
            if occ[i] == "SKIP":
                i += 1
                continue

            if isinstance(occ[i], dict):
                cell = occ[i]
                span = min(cell["colspan"], n - i)
                cells_render.append({"kind": "event", "colspan": span, "items": cell["items"]})
                i += span
            else:
                cells_render.append({"kind": "empty", "colspan": 1, "items": []})
                i += 1

        rows.append({"code": code, "short": short, "cells": cells_render})

    # pour header (nom enseignant sélectionné)
    ens_obj = None
    if enseignant_id.isdigit():
        ens_obj = Enseignant.objects.filter(id=int(enseignant_id)).first()

    return render(request, "admin/seances/edt_prof_week.html", {
        "annee_active": annee_active,

        "annees": annees,
        "annee_selected": annee_id,

        "niveaux": niveaux,
        "niveau_selected": niveau_id,

        "groupes": groupes,
        "groupe_selected": groupe_id,

        "enseignants": enseignants,
        "enseignant_selected": enseignant_id,
        "enseignant_obj": ens_obj,

        "slot_labels": slot_labels,
        "rows": rows,
        "has_slots": bool(seances),
    })
    
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required

@login_required
@group_required("SUPER_ADMIN", "ADMIN", "COMPTABLE", "SCOLARITE")
def ajax_eleves_search(request):
    """
    AJAX — Autocomplete élèves (wizard)
    ✅ Elèves actifs + inscription sur année active (VALIDEE/EN_COURS)
    Retour compatible TomSelect: {items: [...], results: [...]}
    """
    q = (request.GET.get("q") or "").strip()

    annee_active = AnneeScolaire.objects.filter(is_active=True).first()
    if not annee_active:
        return JsonResponse({"items": [], "results": []})

    # ✅ IMPORTANT: filtre direct (évite helper qui casse en prod)
    insc_qs = (
        Inscription.objects
        .select_related("eleve")
        .filter(
            annee=annee_active,
            statut__in=["VALIDEE", "EN_COURS"],
            eleve__is_active=True,   # ✅ filtre safe
        )
    )

    if q:
        insc_qs = insc_qs.filter(
            Q(eleve__matricule__icontains=q) |
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q)
        )

    insc_qs = insc_qs.order_by("eleve__nom", "eleve__prenom")[:50]

    items = []
    for insc in insc_qs:
        e = insc.eleve
        label = f'{(e.matricule or "").strip()} — {e.nom} {e.prenom}'.strip(" —")
        items.append({
            "id": str(e.id),
            "label": label,
            "inscription_id": str(insc.id),
        })

    return JsonResponse({"items": items, "results": items})
